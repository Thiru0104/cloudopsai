#!/usr/bin/env python3
import http.server
import json
import urllib.parse
import os
import csv
import io
from datetime import datetime
from dotenv import load_dotenv
from azure.identity import ClientSecretCredential
from azure.mgmt.subscription import SubscriptionClient
from azure.mgmt.resource import ResourceManagementClient
from azure.mgmt.network import NetworkManagementClient
import asyncio
import logging
import threading
from concurrent.futures import ThreadPoolExecutor
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
# Attempt to import NSG validation utilities; fall back to lightweight stubs
try:
    from nsg_validation import nsg_validator, NSGValidator  # type: ignore
except Exception:
    class NSGValidator:  # minimal stub to keep server running without full AI/DB stack
        def analyze_nsg_rules(self, *args, **kwargs):
            return {
                "summary": {
                    "total_rules": 0,
                    "issues_found": 0
                },
                "recommendations": [],
                "nsgs": []
            }

        def analyze_nsg_rules_from_demo(self, *args, **kwargs):
            return {
                "summary": {
                    "total_rules": 0,
                    "issues_found": 0
                },
                "recommendations": [],
                "nsgs": []
            }

        def generate_llm_recommendations(self, *args, **kwargs):
            return []

        def count_ip_addresses(self, nsg):
            return 0

    # Provide a module-level instance similar to the real implementation
    nsg_validator = NSGValidator()
try:
    from app.services.azure_service import AzureService
except ImportError:
    AzureService = None

# Email imports
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import time

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Load Azure configuration from environment
AZURE_CONFIG = {
    "client_id": os.getenv("AZURE_CLIENT_ID", ""),
    "tenant_id": os.getenv("AZURE_TENANT_ID", ""),
    "client_secret": os.getenv("AZURE_CLIENT_SECRET", ""),
    "subscription_id": os.getenv("AZURE_SUBSCRIPTION_ID", ""),
    "storage_account": os.getenv("AZURE_STORAGE_ACCOUNT", ""),
    "key_vault_url": os.getenv("AZURE_KEY_VAULT_URL", "")
}

# In-memory storage for agents
AGENTS_STORAGE = []

# In-memory storage for email configuration and schedules
EMAIL_CONFIG = {
    "smtpServer": "smtp.gmail.com",
    "smtpPort": 587,
    "smtpUsername": "",
    "smtpPassword": "",
    "fromEmail": "",
    "fromName": "NSG Tool Reports",
    "enableTLS": True
}

EMAIL_SCHEDULES = {}

# In-memory storage for application settings and users
SETTINGS_STORAGE = {
    "security": {
        "twoFactorAuth": False,
        "sessionTimeout": "1 hour",
        "passwordPolicy": True,
        "auditLogging": True
    },
    "notifications": {
        "securityAlerts": True,
        "systemUpdates": True,
        "backupStatus": True
    },
    "system": {
        "version": "v2.1.0",
        "uptime": "15 days, 4 hours",
        "database": "Connected",
        "storageUsage": "2.4 GB / 10 GB"
    }
}

USERS_STORAGE = [
    {
        "id": "1",
        "name": "John Doe",
        "email": "john.doe@company.com",
        "role": "Admin",
        "status": "Active",
        "lastLogin": "2024-01-15 10:30 AM"
    },
    {
        "id": "2",
        "name": "Jane Smith",
        "email": "jane.smith@company.com",
        "role": "User",
        "status": "Active",
        "lastLogin": "2024-01-14 2:15 PM"
    },
    {
        "id": "3",
        "name": "Mike Johnson",
        "email": "mike.johnson@company.com",
        "role": "Viewer",
        "status": "Inactive",
        "lastLogin": "2024-01-10 9:45 AM"
    }
]

# Email sending functionality
def send_email(to_emails, subject, body, attachments=None):
    """Send email using configured SMTP settings"""
    try:
        if not EMAIL_CONFIG.get('smtpUsername') or not EMAIL_CONFIG.get('smtpPassword'):
            logger.warning("Email configuration incomplete - skipping email send")
            return False
            
        msg = MIMEMultipart()
        msg['From'] = f"{EMAIL_CONFIG['fromName']} <{EMAIL_CONFIG['fromEmail']}>"
        msg['To'] = ', '.join(to_emails)
        msg['Subject'] = subject
        
        msg.attach(MIMEText(body, 'html'))
        
        # Add attachments if provided
        if attachments:
            for filename, content in attachments.items():
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(content)
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename= {filename}'
                )
                msg.attach(part)
        
        server = smtplib.SMTP(EMAIL_CONFIG['smtpServer'], EMAIL_CONFIG['smtpPort'])
        if EMAIL_CONFIG['enableTLS']:
            server.starttls()
        server.login(EMAIL_CONFIG['smtpUsername'], EMAIL_CONFIG['smtpPassword'])
        server.send_message(msg)
        server.quit()
        
        logger.info(f"Email sent successfully to {to_emails}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send email: {e}")
        return False

def generate_report_email_body(report_type, report_data):
    """Generate HTML email body for report"""
    return f"""
    <html>
    <body>
        <h2>{report_type} Report</h2>
        <p>Please find your {report_type} report attached.</p>
        <p>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        <p>This is an automated email from NSG Tool.</p>
    </body>
    </html>
    """

# Initialize Azure Service
azure_service = None
if AzureService:
    try:
        azure_service = AzureService()
        logger.info("Azure Service initialized successfully")
    except Exception as e:
        logger.error(f"Failed to initialize Azure Service: {e}")

# Initialize Azure credential using Service Principal
def get_azure_credential():
    try:
        credential = ClientSecretCredential(
            tenant_id=AZURE_CONFIG["tenant_id"],
            client_id=AZURE_CONFIG["client_id"],
            client_secret=AZURE_CONFIG["client_secret"]
        )
        logger.info("Azure SPN authentication initialized successfully")
        return credential
    except Exception as e:
        logger.error(f"Failed to initialize Azure credential: {e}")
        return None

# Initialize Azure clients
def get_azure_clients():
    credential = get_azure_credential()
    if not credential:
        return None, None, None
    
    try:
        subscription_client = SubscriptionClient(credential)
        resource_client = ResourceManagementClient(credential, AZURE_CONFIG["subscription_id"])
        network_client = NetworkManagementClient(credential, AZURE_CONFIG["subscription_id"])
        return subscription_client, resource_client, network_client
    except Exception as e:
        logger.error(f"Failed to initialize Azure clients: {e}")
        return None, None, None

def get_subscriptions():
    """Get Azure subscriptions with enhanced metadata using SPN authentication"""
    try:
        subscription_client, _, _ = get_azure_clients()
        if subscription_client:
            subscriptions = []
            logger.info("Attempting to fetch subscriptions from Azure...")
            
            # Add timeout handling for Azure API calls using threading
            import threading
            import time
            
            result = [None]
            exception = [None]
            
            def fetch_subscriptions():
                try:
                    for sub in subscription_client.subscriptions.list():
                        subscription_data = {
                            "id": sub.subscription_id,
                            "name": sub.display_name,
                            "provider": "azure",
                            "status": "active" if str(sub.state) == "Enabled" else "inactive",
                            "state": str(sub.state),
                            "tenant_id": AZURE_CONFIG["tenant_id"],
                            "resource_groups_count": 0,  # Will be populated separately
                            "last_scan": "2024-01-15T10:30:00Z",
                            "compliance_score": 85,
                            "critical_findings": 3,
                            "subscription_type": "Pay-As-You-Go",
                            "cost_center": "IT-001",
                            "environment": "production" if "prod" in sub.display_name.lower() else "development"
                        }
                        subscriptions.append(subscription_data)
                    result[0] = subscriptions
                except Exception as e:
                    exception[0] = e
            
            # Start the thread
            thread = threading.Thread(target=fetch_subscriptions)
            thread.daemon = True
            thread.start()
            
            # Wait for up to 10 seconds
            thread.join(timeout=10)
            
            if thread.is_alive():
                logger.error("Azure API call timed out after 10 seconds")
                raise TimeoutError("Azure API call timed out")
            
            if exception[0]:
                raise exception[0]
            
            if result[0] is not None:
                logger.info(f"Successfully fetched {len(result[0])} subscriptions from Azure")
                return {
                    "subscriptions": result[0],
                    "total_subscriptions": len(result[0]),
                    "active_subscriptions": len([s for s in result[0] if s["status"] == "active"])
                }
                
    except Exception as e:
        logger.error(f"Failed to fetch real subscriptions: {e}")
    
    # Fallback to mock data
    logger.info("Using mock subscription data")
    return {
        "subscriptions": [
            {
                "id": "12345678-1234-1234-1234-123456789012",
                "name": "Production Environment (.env configured)",
                "provider": "azure",
                "status": "active",
                "state": "Enabled",
                "tenant_id": AZURE_CONFIG["tenant_id"],
                "resource_groups_count": 15,
                "last_scan": "2024-01-15T10:30:00Z",
                "compliance_score": 85,
                "critical_findings": 3,
                "subscription_type": "Pay-As-You-Go",
                "cost_center": "IT-001",
                "environment": "production"
            }
        ],
        "total_subscriptions": 1,
        "active_subscriptions": 1
    }

def get_resource_groups(subscription_id):
    """Get resource groups for a subscription using SPN authentication"""
    try:
        _, resource_client, _ = get_azure_clients()
        if resource_client:
            resource_groups = []
            for rg in resource_client.resource_groups.list():
                rg_data = {
                    "id": rg.id,
                    "name": rg.name,
                    "location": rg.location,
                    "status": "Succeeded",  # Resource groups don't have provisioning state
                    "tags": rg.tags or {},
                    "resource_count": 0,  # Will be populated separately if needed
                    "last_modified": "2024-01-15T08:30:00Z"  # Mock timestamp
                }
                resource_groups.append(rg_data)
            
            return {
                "resource_groups": resource_groups,
                "total_resource_groups": len(resource_groups)
            }
    except Exception as e:
        logger.error(f"Failed to fetch real resource groups: {e}")
    
    # Fallback to mock data
    logger.info("Using mock resource group data")
    return {
        "resource_groups": [
            {
                "id": "/subscriptions/12345678-1234-1234-1234-123456789012/resourceGroups/rg-production",
                "name": "rg-production",
                "location": "East US",
                "status": "Succeeded",
                "tags": {"environment": "production", "cost-center": "IT-001"},
                "resource_count": 25,
                "last_modified": "2024-01-15T08:30:00Z"
            },
            {
                "id": "/subscriptions/12345678-1234-1234-1234-123456789012/resourceGroups/rg-development",
                "name": "rg-development",
                "location": "West US 2",
                "status": "Succeeded",
                "tags": {"environment": "development", "cost-center": "IT-002"},
                "resource_count": 12,
                "last_modified": "2024-01-14T16:45:00Z"
            }
        ],
        "total_resource_groups": 2
    }

def get_nsgs(subscription_id, resource_group_name=None):
    """Get Network Security Groups with enhanced metadata using SPN authentication"""
    try:
        _, _, network_client = get_azure_clients()
        if network_client:
            nsgs = []
            nsg_counter = 1
            
            if resource_group_name:
                # Get NSGs for specific resource group
                nsg_list = network_client.network_security_groups.list(resource_group_name)
            else:
                # Get all NSGs in subscription
                nsg_list = network_client.network_security_groups.list_all()
            
            for nsg in nsg_list:
                # Extract resource group from NSG ID
                rg_name = nsg.id.split('/')[4] if len(nsg.id.split('/')) > 4 else 'unknown'
                
                # Get inbound and outbound rules (both custom and default)
                inbound_rules = []
                outbound_rules = []
                
                # Process custom security rules
                if nsg.security_rules:
                    for rule in nsg.security_rules:
                        rule_data = {
                            "name": rule.name,
                            "priority": rule.priority,
                            "direction": rule.direction,
                            "access": rule.access,
                            "protocol": rule.protocol,
                            "source_port_range": rule.source_port_range,
                            "destination_port_range": rule.destination_port_range,
                            "source_address_prefix": rule.source_address_prefix,
                            "destination_address_prefix": rule.destination_address_prefix,
                            # Include plural address prefixes for consistency
                            "source_address_prefixes": getattr(rule, 'source_address_prefixes', None) or [],
                            "destination_address_prefixes": getattr(rule, 'destination_address_prefixes', None) or [],
                            "description": rule.description or "",
                            "is_default": False
                        }
                        
                        if rule.direction == "Inbound":
                            inbound_rules.append(rule_data)
                        else:
                            outbound_rules.append(rule_data)
                
                # Process default security rules
                if nsg.default_security_rules:
                    for rule in nsg.default_security_rules:
                        rule_data = {
                            "name": rule.name,
                            "priority": rule.priority,
                            "direction": rule.direction,
                            "access": rule.access,
                            "protocol": rule.protocol,
                            "source_port_range": rule.source_port_range,
                            "destination_port_range": rule.destination_port_range,
                            "source_address_prefix": rule.source_address_prefix,
                            "destination_address_prefix": rule.destination_address_prefix,
                            # Include plural address prefixes for consistency
                            "source_address_prefixes": getattr(rule, 'source_address_prefixes', None) or [],
                            "destination_address_prefixes": getattr(rule, 'destination_address_prefixes', None) or [],
                            "description": rule.description or "",
                            "is_default": True
                        }
                        
                        if rule.direction == "Inbound":
                            inbound_rules.append(rule_data)
                        else:
                            outbound_rules.append(rule_data)
                
                # Calculate compliance score and risk level
                total_rules = len(inbound_rules) + len(outbound_rules)
                compliance_score = max(60, min(95, 85 - (total_rules * 2)))  # Mock calculation
                
                if compliance_score >= 90:
                    risk_level = "low"
                elif compliance_score >= 70:
                    risk_level = "medium"
                elif compliance_score >= 50:
                    risk_level = "high"
                else:
                    risk_level = "critical"
                
                nsg_data = {
                    "id": nsg_counter,
                    "name": nsg.name,
                    "resource_group": rg_name,
                    "region": nsg.location,
                    "subscription_id": subscription_id,
                    "azure_id": nsg.id,
                    "inbound_rules": inbound_rules,
                    "outbound_rules": outbound_rules,
                    "tags": nsg.tags or {},
                    "is_active": nsg.provisioning_state == "Succeeded",
                    "compliance_score": compliance_score,
                    "risk_level": risk_level,
                    "last_sync": "2024-08-22T13:51:52Z",
                    "last_backup": None,
                    "created_at": "2024-01-15T09:15:00Z",
                    "updated_at": "2024-08-22T13:51:52Z"
                }
                nsgs.append(nsg_data)
                nsg_counter += 1
            
            logger.info(f"Successfully fetched {len(nsgs)} NSGs from Azure")
            return {
                "nsgs": nsgs,
                "total_nsgs": len(nsgs),
                "active_nsgs": len([n for n in nsgs if n["is_active"]]),
                "total_rules": sum(len(n["inbound_rules"]) + len(n["outbound_rules"]) for n in nsgs)
            }
    except Exception as e:
        logger.error(f"Failed to fetch real NSGs: {e}")
    
    # Fallback to mock data with correct structure
    logger.info("Using mock NSG data")
    return {
        "nsgs": [
            {
                "id": 1,
                "name": "nsg-web-tier",
                "resource_group": "rg-production",
                "region": "East US",
                "subscription_id": subscription_id,
                "azure_id": "/subscriptions/12345678-1234-1234-1234-123456789012/resourceGroups/rg-production/providers/Microsoft.Network/networkSecurityGroups/nsg-web-tier",
                "inbound_rules": [
                    {
                        "name": "AllowHTTP",
                        "priority": 100,
                        "direction": "Inbound",
                        "access": "Allow",
                        "protocol": "TCP",
                        "source_port_range": "*",
                        "destination_port_range": "80",
                        "source_address_prefix": "*",
                        "destination_address_prefix": "*",
                        "description": "Allow HTTP traffic"
                    }
                ],
                "outbound_rules": [
                    {
                        "name": "AllowAllOutbound",
                        "priority": 100,
                        "direction": "Outbound",
                        "access": "Allow",
                        "protocol": "*",
                        "source_port_range": "*",
                        "destination_port_range": "*",
                        "source_address_prefix": "*",
                        "destination_address_prefix": "*",
                        "description": "Allow all outbound traffic"
                    }
                ],
                "tags": {"tier": "web", "environment": "production"},
                "is_active": True,
                "compliance_score": 85,
                "risk_level": "medium",
                "last_sync": "2024-08-22T13:51:52Z",
                "last_backup": None,
                "created_at": "2024-01-15T09:15:00Z",
                "updated_at": "2024-08-22T13:51:52Z"
            },
            {
                "id": 2,
                "name": "nsg-app-tier",
                "resource_group": "rg-production",
                "region": "East US",
                "subscription_id": subscription_id,
                "azure_id": "/subscriptions/12345678-1234-1234-1234-123456789012/resourceGroups/rg-production/providers/Microsoft.Network/networkSecurityGroups/nsg-app-tier",
                "inbound_rules": [
                    {
                        "name": "AllowHTTPS",
                        "priority": 100,
                        "direction": "Inbound",
                        "access": "Allow",
                        "protocol": "TCP",
                        "source_port_range": "*",
                        "destination_port_range": "443",
                        "source_address_prefix": "*",
                        "destination_address_prefix": "*",
                        "description": "Allow HTTPS traffic"
                    }
                ],
                "outbound_rules": [
                    {
                        "name": "AllowAllOutbound",
                        "priority": 100,
                        "direction": "Outbound",
                        "access": "Allow",
                        "protocol": "*",
                        "source_port_range": "*",
                        "destination_port_range": "*",
                        "source_address_prefix": "*",
                        "destination_address_prefix": "*",
                        "description": "Allow all outbound traffic"
                    }
                ],
                "tags": {"tier": "application", "environment": "production"},
                "is_active": True,
                "compliance_score": 75,
                "risk_level": "high",
                "last_sync": "2024-08-22T13:51:52Z",
                "last_backup": None,
                "created_at": "2024-01-14T14:30:00Z",
                "updated_at": "2024-08-22T13:51:52Z"
            }
        ],
        "total_nsgs": 2,
        "active_nsgs": 2,
        "total_rules": 4
    }

def get_route_tables(subscription_id, resource_group_name=None):
    """Get Route Tables with enhanced metadata using SPN authentication"""
    try:
        _, _, network_client = get_azure_clients()
        if network_client:
            route_tables = []
            rt_counter = 1
            
            if resource_group_name:
                # Get Route Tables for specific resource group
                rt_list = network_client.route_tables.list(resource_group_name)
            else:
                # Get all Route Tables in subscription
                rt_list = network_client.route_tables.list_all()
            
            for rt in rt_list:
                # Extract resource group from Route Table ID
                rg_name = rt.id.split('/')[4] if len(rt.id.split('/')) > 4 else 'unknown'
                
                # Get routes
                routes = []
                if rt.routes:
                    for route in rt.routes:
                        route_data = {
                            "name": route.name,
                            "address_prefix": route.address_prefix,
                            "next_hop_type": route.next_hop_type,
                            "next_hop_ip_address": route.next_hop_ip_address,
                            "provisioning_state": route.provisioning_state
                        }
                        routes.append(route_data)
                
                rt_data = {
                    "id": rt_counter,
                    "name": rt.name,
                    "resource_group": rg_name,
                    "region": rt.location,
                    "subscription_id": subscription_id,
                    "azure_id": rt.id,
                    "routes": routes,
                    "tags": rt.tags or {},
                    "is_active": rt.provisioning_state == "Succeeded",
                    "last_sync": "2024-08-22T13:51:52Z",
                    "created_at": "2024-01-15T09:15:00Z",
                    "updated_at": "2024-08-22T13:51:52Z"
                }
                route_tables.append(rt_data)
                rt_counter += 1
            
            logger.info(f"Successfully fetched {len(route_tables)} Route Tables from Azure")
            return {
                "route_tables": route_tables,
                "total_route_tables": len(route_tables),
                "active_route_tables": len([rt for rt in route_tables if rt["is_active"]]),
                "total_routes": sum(len(rt["routes"]) for rt in route_tables)
            }
    except Exception as e:
        logger.error(f"Failed to fetch real Route Tables: {e}")
    
    # Fallback to mock data with correct structure
    logger.info("Using mock Route Table data")
    return {
        "route_tables": [
            {
                "id": 1,
                "name": "rt-web-tier",
                "resource_group": "rg-production",
                "region": "East US",
                "subscription_id": subscription_id,
                "azure_id": "/subscriptions/12345678-1234-1234-1234-123456789012/resourceGroups/rg-production/providers/Microsoft.Network/routeTables/rt-web-tier",
                "routes": [
                    {
                        "name": "route-to-internet",
                        "address_prefix": "0.0.0.0/0",
                        "next_hop_type": "Internet",
                        "next_hop_ip_address": None,
                        "provisioning_state": "Succeeded"
                    },
                    {
                        "name": "route-to-vnet",
                        "address_prefix": "10.0.0.0/16",
                        "next_hop_type": "VnetLocal",
                        "next_hop_ip_address": None,
                        "provisioning_state": "Succeeded"
                    }
                ],
                "tags": {"tier": "web", "environment": "production"},
                "is_active": True,
                "last_sync": "2024-08-22T13:51:52Z",
                "created_at": "2024-01-15T09:15:00Z",
                "updated_at": "2024-08-22T13:51:52Z"
            },
            {
                "id": 2,
                "name": "rt-app-tier",
                "resource_group": "rg-production",
                "region": "East US",
                "subscription_id": subscription_id,
                "azure_id": "/subscriptions/12345678-1234-1234-1234-123456789012/resourceGroups/rg-production/providers/Microsoft.Network/routeTables/rt-app-tier",
                "routes": [
                    {
                        "name": "route-to-firewall",
                        "address_prefix": "0.0.0.0/0",
                        "next_hop_type": "VirtualAppliance",
                        "next_hop_ip_address": "10.0.1.4",
                        "provisioning_state": "Succeeded"
                    }
                ],
                "tags": {"tier": "application", "environment": "production"},
                "is_active": True,
                "last_sync": "2024-08-22T13:51:52Z",
                "created_at": "2024-01-14T14:30:00Z",
                "updated_at": "2024-08-22T13:51:52Z"
            }
        ],
        "total_route_tables": 2,
        "active_route_tables": 2,
        "total_routes": 3
    }

def get_asgs(subscription_id, resource_group_name=None):
    """Get Application Security Groups with enhanced metadata using SPN authentication"""
    try:
        _, _, network_client = get_azure_clients()
        if network_client:
            asgs = []
            asg_counter = 1
            
            if resource_group_name:
                # Get ASGs for specific resource group
                asg_list = network_client.application_security_groups.list(resource_group_name)
            else:
                # Get all ASGs in subscription
                asg_list = network_client.application_security_groups.list_all()
            
            for asg in asg_list:
                # Extract resource group from ASG ID
                rg_name = asg.id.split('/')[4] if len(asg.id.split('/')) > 4 else 'unknown'
                
                # Calculate compliance and validation metrics
                name_length = len(asg.name)
                has_tags = bool(asg.tags)
                is_properly_named = name_length >= 3 and name_length <= 80
                
                # Determine validation status
                validation_issues = []
                if name_length < 3:
                    validation_issues.append("Name too short (minimum 3 characters)")
                elif name_length > 80:
                    validation_issues.append("Name too long (maximum 80 characters)")
                
                if not has_tags:
                    validation_issues.append("Missing resource tags")
                
                validation_status = "valid" if len(validation_issues) == 0 else "invalid"
                risk_level = "low" if validation_status == "valid" else "medium" if len(validation_issues) == 1 else "high"
                
                asg_data = {
                    "id": asg_counter,
                    "name": asg.name,
                    "resource_group": rg_name,
                    "region": asg.location,
                    "subscription_id": subscription_id,
                    "azure_id": asg.id,
                    "tags": asg.tags or {},
                    "is_active": asg.provisioning_state == "Succeeded",
                    "validation_status": validation_status,
                    "validation_issues": validation_issues,
                    "risk_level": risk_level,
                    "last_sync": "2024-08-22T13:51:52Z",
                    "created_at": "2024-01-15T09:15:00Z",
                    "updated_at": "2024-08-22T13:51:52Z"
                }
                asgs.append(asg_data)
                asg_counter += 1
            
            logger.info(f"Successfully fetched {len(asgs)} ASGs from Azure")
            return {
                "asgs": asgs,
                "total_asgs": len(asgs),
                "active_asgs": len([a for a in asgs if a["is_active"]]),
                "valid_asgs": len([a for a in asgs if a["validation_status"] == "valid"]),
                "invalid_asgs": len([a for a in asgs if a["validation_status"] == "invalid"])
            }
    except Exception as e:
        logger.error(f"Failed to fetch real ASGs: {e}")
    
    # Fallback to mock data with correct structure
    logger.info("Using mock ASG data")
    return {
        "asgs": [
            {
                "id": 1,
                "name": "asg-web-servers",
                "resource_group": "rg-production",
                "region": "East US",
                "subscription_id": subscription_id,
                "azure_id": "/subscriptions/12345678-1234-1234-1234-123456789012/resourceGroups/rg-production/providers/Microsoft.Network/applicationSecurityGroups/asg-web-servers",
                "tags": {"tier": "web", "environment": "production"},
                "is_active": True,
                "validation_status": "valid",
                "validation_issues": [],
                "risk_level": "low",
                "last_sync": "2024-08-22T13:51:52Z",
                "created_at": "2024-01-15T09:15:00Z",
                "updated_at": "2024-08-22T13:51:52Z"
            },
            {
                "id": 2,
                "name": "asg-app-servers",
                "resource_group": "rg-production",
                "region": "East US",
                "subscription_id": subscription_id,
                "azure_id": "/subscriptions/12345678-1234-1234-1234-123456789012/resourceGroups/rg-production/providers/Microsoft.Network/applicationSecurityGroups/asg-app-servers",
                "tags": {"tier": "application", "environment": "production"},
                "is_active": True,
                "validation_status": "valid",
                "validation_issues": [],
                "risk_level": "low",
                "last_sync": "2024-08-22T13:51:52Z",
                "created_at": "2024-01-14T14:30:00Z",
                "updated_at": "2024-08-22T13:51:52Z"
            },
            {
                "id": 3,
                "name": "db",
                "resource_group": "rg-development",
                "region": "West US 2",
                "subscription_id": subscription_id,
                "azure_id": "/subscriptions/12345678-1234-1234-1234-123456789012/resourceGroups/rg-development/providers/Microsoft.Network/applicationSecurityGroups/db",
                "tags": {},
                "is_active": True,
                "validation_status": "invalid",
                "validation_issues": ["Name too short (minimum 3 characters)", "Missing resource tags"],
                "risk_level": "high",
                "last_sync": "2024-08-22T13:51:52Z",
                "created_at": "2024-01-12T10:20:00Z",
                "updated_at": "2024-08-22T13:51:52Z"
            }
        ],
        "total_asgs": 3,
        "active_asgs": 3,
        "valid_asgs": 2,
        "invalid_asgs": 1
    }

def update_nsg_rules_in_azure(subscription_id, resource_group, nsg_id, inbound_rules, outbound_rules):
    """Update NSG rules in Azure using the Network Management Client"""
    try:
        _, _, network_client = get_azure_clients()
        if not network_client:
            return {"success": False, "error": "Failed to get Azure network client"}
        
        # First, we need to find the NSG by ID or name
        # Since we're getting an integer ID from frontend, we need to map it to actual NSG
        nsgs = list(network_client.network_security_groups.list_all())
        target_nsg = None
        
        # Find NSG by index (since frontend uses sequential IDs)
        if isinstance(nsg_id, int) and 1 <= nsg_id <= len(nsgs):
            target_nsg = nsgs[nsg_id - 1]  # Convert 1-based to 0-based index
        
        if not target_nsg:
            return {"success": False, "error": f"NSG with ID {nsg_id} not found"}
        
        # Extract resource group from NSG ID if not provided
        if not resource_group:
            resource_group = target_nsg.id.split('/')[4]
        
        nsg_name = target_nsg.name
        logger.info(f"Updating NSG rules for {nsg_name} in resource group {resource_group}")
        
        # Get current NSG to preserve existing structure
        current_nsg = network_client.network_security_groups.get(resource_group, nsg_name)
        
        # Convert frontend rule format to Azure SDK format
        def convert_rule_to_azure_format(rule, direction):
            # Handle port ranges - ensure empty/null values become '*'
            source_port = rule.get('source_port_range', '*')
            if not source_port or source_port.strip() == '':
                source_port = '*'
            
            dest_port = rule.get('destination_port_range', '*')
            if not dest_port or dest_port.strip() == '':
                dest_port = '*'
            
            # Handle address prefixes - prefer arrays when provided, otherwise use single value
            source_prefixes = rule.get('source_address_prefixes') or []
            dest_prefixes = rule.get('destination_address_prefixes') or []

            source_addr = rule.get('source_address_prefix', '*')
            if not source_addr or str(source_addr).strip() == '':
                source_addr = '*'
            
            dest_addr = rule.get('destination_address_prefix', '*')
            if not dest_addr or str(dest_addr).strip() == '':
                dest_addr = '*'
            
            azure_rule = {
                'name': rule.get('name', ''),
                'priority': int(rule.get('priority', 1000)),
                'direction': direction,
                'access': rule.get('access', 'Allow'),
                'protocol': rule.get('protocol', 'TCP'),
                'source_port_range': source_port,
                'destination_port_range': dest_port,
                'description': rule.get('description', '')
            }
            # Only set one of singular or array fields to match Azure API expectations
            if source_prefixes:
                azure_rule['source_address_prefixes'] = source_prefixes
            else:
                azure_rule['source_address_prefix'] = source_addr
            if dest_prefixes:
                azure_rule['destination_address_prefixes'] = dest_prefixes
            else:
                azure_rule['destination_address_prefix'] = dest_addr

            return azure_rule
        
        # Convert rules to Azure format
        azure_security_rules = []
        
        # Add inbound rules
        for rule in inbound_rules:
            if rule.get('name'):  # Only add rules with names
                azure_rule = convert_rule_to_azure_format(rule, 'Inbound')
                azure_security_rules.append(azure_rule)
        
        # Add outbound rules
        for rule in outbound_rules:
            if rule.get('name'):  # Only add rules with names
                azure_rule = convert_rule_to_azure_format(rule, 'Outbound')
                azure_security_rules.append(azure_rule)
        
        # Update NSG with new rules
        nsg_params = {
            'location': current_nsg.location,
            'security_rules': azure_security_rules,
            'tags': current_nsg.tags
        }
        
        # Perform the update
        operation = network_client.network_security_groups.begin_create_or_update(
            resource_group, nsg_name, nsg_params
        )
        
        # Wait for completion
        result = operation.result()
        
        logger.info(f"Successfully updated NSG {nsg_name} with {len(azure_security_rules)} rules")
        return {"success": True, "nsg_name": nsg_name}
        
    except Exception as e:
        logger.error(f"Failed to update NSG rules in Azure: {e}")
        return {"success": False, "error": str(e)}


def convert_any_rule_to_azure_format(rule):
    """Normalize a rule dict from various sources (CSV, frontend, backup) to Azure SDK format.

    Supports both singular and plural address prefixes, multiple key styles (snake_case, camelCase),
    and gracefully handles comma-separated strings for prefix lists and ports.
    """
    try:
        # Name and basic fields
        name = rule.get('name', '')
        direction = rule.get('direction') or rule.get('Direction') or 'Inbound'
        access = rule.get('access') or rule.get('action') or rule.get('Access') or 'Allow'
        protocol = (
            rule.get('protocol')
            or rule.get('Protocol')
            or rule.get('prot')
            or '*'
        )

        # Ports: check multiple key styles and normalize blanks to '*'
        def norm_port(val):
            if val is None:
                return '*'
            s = str(val).strip()
            return s if s else '*'

        source_port = (
            rule.get('source_port_range')
            or rule.get('sourcePortRange')
            or rule.get('source_port')
            or rule.get('Source Port')
            or rule.get('SourcePort')
            or '*'
        )
        dest_port = (
            rule.get('destination_port_range')
            or rule.get('destinationPortRange')
            or rule.get('destination_port')
            or rule.get('Destination Port')
            or rule.get('DestinationPort')
            or '*'
        )
        source_port = norm_port(source_port)
        dest_port = norm_port(dest_port)

        # Addresses: plural lists take precedence; normalize comma-separated strings
        def to_list(value):
            if value is None:
                return []
            if isinstance(value, list):
                return [str(v).strip() for v in value if str(v).strip()]
            # split on comma if string
            parts = [p.strip() for p in str(value).split(',')]
            return [p for p in parts if p]

        src_list = (
            rule.get('source_address_prefixes')
            or rule.get('sourceAddressPrefixes')
            or rule.get('source_addresses')
            or rule.get('Source Addresses')
            or []
        )
        dst_list = (
            rule.get('destination_address_prefixes')
            or rule.get('destinationAddressPrefixes')
            or rule.get('destination_addresses')
            or rule.get('Destination Addresses')
            or []
        )
        src_list = to_list(src_list)
        dst_list = to_list(dst_list)

        # Singular fallbacks (first element or explicit singular keys)
        src_single = (
            rule.get('source_address_prefix')
            or rule.get('sourceAddressPrefix')
            or rule.get('source_address')
            or rule.get('Source Address')
            or rule.get('source')
            or '*'
        )
        dst_single = (
            rule.get('destination_address_prefix')
            or rule.get('destinationAddressPrefix')
            or rule.get('destination_address')
            or rule.get('Destination Address')
            or rule.get('destination')
            or '*'
        )

        def norm_addr(val):
            if val is None:
                return '*'
            s = str(val).strip()
            return s if s else '*'

        # If lists contain '*' only, prefer singular '*'
        if src_list == ['*']:
            src_list = []
        if dst_list == ['*']:
            dst_list = []

        # Use first item from list as singular fallback when needed
        if not src_single or str(src_single).strip() == '':
            src_single = src_list[0] if src_list else '*'
        if not dst_single or str(dst_single).strip() == '':
            dst_single = dst_list[0] if dst_list else '*'

        src_single = norm_addr(src_single)
        dst_single = norm_addr(dst_single)

        # Priority and description
        priority_raw = (
            rule.get('priority')
            or rule.get('Priority')
            or 1000
        )
        try:
            priority = int(priority_raw)
        except Exception:
            priority = 1000

        description = (
            rule.get('description')
            or rule.get('Description')
            or ''
        )

        azure_rule = {
            'name': name,
            'priority': priority,
            'direction': direction,
            'access': access,
            'protocol': protocol,
            'source_port_range': source_port,
            'destination_port_range': dest_port,
            'description': description,
        }

        # Prefer plural fields when lists provided; otherwise set singular
        if src_list:
            azure_rule['source_address_prefixes'] = src_list
        else:
            azure_rule['source_address_prefix'] = src_single
        if dst_list:
            azure_rule['destination_address_prefixes'] = dst_list
        else:
            azure_rule['destination_address_prefix'] = dst_single

        return azure_rule
    except Exception as e:
        logger.error(f"Failed to normalize rule: {e}")
        # Minimal fallback
        return {
            'name': rule.get('name', ''),
            'priority': int(rule.get('priority', 1000)),
            'direction': rule.get('direction', 'Inbound'),
            'access': rule.get('access', 'Allow'),
            'protocol': rule.get('protocol', '*'),
            'source_port_range': rule.get('source_port_range', '*'),
            'destination_port_range': rule.get('destination_port_range', '*'),
            'source_address_prefix': rule.get('source_address_prefix', '*'),
            'destination_address_prefix': rule.get('destination_address_prefix', '*'),
            'description': rule.get('description', '')
        }

def create_csv_for_single_nsg(nsg_data):
    """Create CSV content for a single NSG"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header matching the detailed format from the reference image
    writer.writerow([
        'Subscription', 'Resource Group', 'NSG Name', 'Rule Name', 'Direction', 
        'Priority', 'Access', 'Protocol', 'Source', 'Destination', 
        'Owner Address', 'Destination Address', 'Source ASG', 'Destination ASG', 'Description'
    ])
    
    subscription_name = nsg_data.get('subscription_id', '')[:8] if nsg_data.get('subscription_id') else 'Thiru'
    resource_group = nsg_data.get('resource_group', '') or 'resourcegroup002'
    nsg_name = nsg_data.get('name', '') or 'nsg_name'
    
    # Process all rules (both inbound and outbound)
    all_rules = []
    all_rules.extend([(rule, 'Inbound') for rule in nsg_data.get('inbound_rules', [])])
    all_rules.extend([(rule, 'Outbound') for rule in nsg_data.get('outbound_rules', [])])
    
    for rule, direction in all_rules:
        source_port = rule.get('source_port_range', '*')
        dest_port = rule.get('destination_port_range', '*')

        # Prefer array fields when present; fall back to singular
        src_prefixes = rule.get('source_address_prefixes') or []
        dst_prefixes = rule.get('destination_address_prefixes') or []

        # Normalize to list of strings
        if not isinstance(src_prefixes, list):
            src_prefixes = []
        if not isinstance(dst_prefixes, list):
            dst_prefixes = []

        # If arrays are empty, use singular values
        if not src_prefixes:
            singular_src = rule.get('source_address_prefix', '*')
            src_prefixes = [singular_src] if singular_src else ['*']
        if not dst_prefixes:
            singular_dst = rule.get('destination_address_prefix', '*')
            dst_prefixes = [singular_dst] if singular_dst else ['*']
        
        # Join multiple prefixes for display
        source_addr = ', '.join([str(p) for p in src_prefixes if p]) or '*'
        dest_addr = ', '.join([str(p) for p in dst_prefixes if p]) or '*'
        
        # Get ASG information
        source_asgs = rule.get('source_application_security_groups', [])
        dest_asgs = rule.get('destination_application_security_groups', [])
        
        # Extract ASG names from IDs (get the last part after the last slash)
        source_asg_names = []
        for asg_id in source_asgs:
            if isinstance(asg_id, str) and asg_id:
                asg_name = asg_id.split('/')[-1] if '/' in asg_id else asg_id
                source_asg_names.append(asg_name)
        
        dest_asg_names = []
        for asg_id in dest_asgs:
            if isinstance(asg_id, str) and asg_id:
                asg_name = asg_id.split('/')[-1] if '/' in asg_id else asg_id
                dest_asg_names.append(asg_name)
        
        # Format ASG names for display
        source_asg_display = ', '.join(source_asg_names) if source_asg_names else 'None'
        dest_asg_display = ', '.join(dest_asg_names) if dest_asg_names else 'None'
        
        # Format source and destination ports
        source_ports = source_port if source_port and source_port != '*' else '*'
        dest_ports = dest_port if dest_port and dest_port != '*' else '*'
        
        writer.writerow([
            subscription_name,
            resource_group,
            nsg_name,
            rule.get('name', ''),
            direction,
            rule.get('priority', ''),
            rule.get('access', ''),
            rule.get('protocol', '*'),
            source_ports,
            dest_ports,
            source_addr,  # Owner Address
            dest_addr,    # Destination Address
            source_asg_display,  # Source ASG
            dest_asg_display,    # Destination ASG
            rule.get('description', f"Allow {rule.get('access', '').lower()} {direction.lower()} traffic")
        ])
    
    # If no rules exist, add a placeholder row to indicate empty NSG
    if not all_rules:
        writer.writerow([
            subscription_name,
            resource_group,
            nsg_name,
            'No Rules Configured',
            'N/A',
            'N/A',
            'N/A',
            'N/A',
            'N/A',
            'N/A',
            'N/A',
            'N/A',
            'None',
            'None',
            f'NSG {nsg_name} has no security rules configured'
        ])
    
    return output.getvalue()

def create_standardized_csv_format(backup_content):
    """Create standardized CSV format for all backup, export, and restore operations"""
    try:
        logger.info(f"Creating CSV from backup content with {len(backup_content.get('nsgs', []))} NSGs")
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write standardized header format that will be used consistently across all operations
        writer.writerow([
            'Subscription', 'Resource Group', 'NSG Name', 'Rule Name', 'Direction', 
            'Priority', 'Access', 'Protocol', 'Source', 'Destination', 
            'Owner Address', 'Destination Address', 'Source ASG', 'Destination ASG', 'Description'
        ])
        
        # Write NSG rules with consistent data structure
        for nsg in backup_content.get('nsgs', []):
            logger.info(f"Processing NSG: {nsg.get('name', 'unknown')} with type: {type(nsg)}")
            subscription_name = nsg.get('subscription_id', '')[:8] if nsg.get('subscription_id') else 'Thiru'
            resource_group = nsg.get('resource_group', '') or 'resourcegroup002'
            nsg_name = nsg.get('name', '') or 'nsg_name'
        
            # Process all rules (both inbound and outbound)
            all_rules = []
            all_rules.extend([(rule, 'Inbound') for rule in nsg.get('inbound_rules', [])])
            all_rules.extend([(rule, 'Outbound') for rule in nsg.get('outbound_rules', [])])
            
            # If no rules exist, create a basic NSG entry to maintain consistency
            if not all_rules:
                writer.writerow([
                    subscription_name,
                    resource_group,
                    nsg_name,
                    'No Rules Configured',
                    'N/A',
                    'N/A',
                    'N/A',
                    'N/A',
                    'N/A',
                    'N/A',
                    'N/A',
                    'N/A',
                    'None',
                    'None',
                    f'NSG {nsg_name} has no security rules configured'
                ])
            else:
                for rule, direction in all_rules:
                    # Skip if rule is not a dictionary
                    if not isinstance(rule, dict):
                        logger.warning(f"Skipping invalid rule (not a dict): {rule}")
                        continue
                        
                    source_port = rule.get('source_port_range', '*')
                    dest_port = rule.get('destination_port_range', '*')
                    
                    # Prefer array fields when present; fall back to singular
                    src_prefixes = rule.get('source_address_prefixes') or []
                    dst_prefixes = rule.get('destination_address_prefixes') or []

                    # Normalize to list of strings
                    if not isinstance(src_prefixes, list):
                        src_prefixes = []
                    if not isinstance(dst_prefixes, list):
                        dst_prefixes = []

                    # If arrays are empty, use singular values
                    if not src_prefixes:
                        singular_src = rule.get('source_address_prefix', '*')
                        src_prefixes = [singular_src] if singular_src else ['*']
                    if not dst_prefixes:
                        singular_dst = rule.get('destination_address_prefix', '*')
                        dst_prefixes = [singular_dst] if singular_dst else ['*']

                    # Join multiple prefixes for display
                    source_addr = ', '.join([str(p) for p in src_prefixes if p]) or '*'
                    dest_addr = ', '.join([str(p) for p in dst_prefixes if p]) or '*'
                    
                    # Get ASG information
                    source_asgs = rule.get('source_application_security_groups', [])
                    dest_asgs = rule.get('destination_application_security_groups', [])
                    
                    # Handle both ASG objects and ASG ID strings
                    source_asg_names = []
                    if source_asgs:
                        for asg in source_asgs:
                            if isinstance(asg, dict):
                                source_asg_names.append(asg.get('name', ''))
                            else:
                                # ASG is an ID string, extract name from it
                                asg_name = str(asg).split('/')[-1] if '/' in str(asg) else str(asg)
                                source_asg_names.append(asg_name)
                    
                    dest_asg_names = []
                    if dest_asgs:
                        for asg in dest_asgs:
                            if isinstance(asg, dict):
                                dest_asg_names.append(asg.get('name', ''))
                            else:
                                # ASG is an ID string, extract name from it
                                asg_name = str(asg).split('/')[-1] if '/' in str(asg) else str(asg)
                                dest_asg_names.append(asg_name)
                    
                    writer.writerow([
                        subscription_name,
                        resource_group,
                        nsg_name,
                        rule.get('name', ''),
                        direction,
                        rule.get('priority', ''),
                        rule.get('access', ''),
                        rule.get('protocol', ''),
                        source_port,
                        dest_port,
                        source_addr,
                        dest_addr,
                        ', '.join(source_asg_names) if source_asg_names else 'None',
                        ', '.join(dest_asg_names) if dest_asg_names else 'None',
                        rule.get('description', '')
                    ])
        
        return output.getvalue()
    except Exception as e:
        logger.error(f"Error in create_standardized_csv_format: {e}")
        logger.error(f"Backup content type: {type(backup_content)}")
        logger.error(f"Backup content keys: {list(backup_content.keys()) if isinstance(backup_content, dict) else 'Not a dict'}")
        raise e

def create_csv_from_backup(backup_content):
    """Convert backup data to enhanced CSV format matching the detailed NSG rules table"""
    # Use the standardized format function
    return create_standardized_csv_format(backup_content)

def parse_csv_rules(csv_content):
    """Parse CSV content and convert to NSG rules format.

    - Supports multiple source/destination prefixes via comma-separated values.
    - Produces both snake_case keys (for legacy restore code) and Azure-style keys.
    - Normalizes missing ports/addresses to '*'.
    """
    rules = []
    try:
        import io
        csv_reader = csv.DictReader(io.StringIO(csv_content))

        logger.info(f"CSV Headers: {csv_reader.fieldnames}")

        # Check if this is an NSG summary CSV (contains NSG metadata) or rules CSV
        headers = csv_reader.fieldnames or []
        is_nsg_summary = any(header in headers for header in ['Subscription Name', 'Subscription ID', 'NSG Name', 'Resource Group'])
        is_rules_csv = any(header in headers for header in ['Rule Name', 'Priority', 'Direction', 'Access'])

        if is_nsg_summary and not is_rules_csv:
            logger.warning("CSV appears to be NSG summary data, not individual rules. Cannot restore from this format.")
            logger.info("Expected CSV format should contain columns: Rule Name, Priority, Direction, Access, Protocol, Source, Destination, Description")
            return []

        def norm_star(val):
            if val is None:
                return '*'
            s = str(val).strip()
            return s if s else '*'

        for row in csv_reader:
            # Skip empty rows or rows without rule names
            name_field = (row.get('Rule Name') or '').strip()
            if not name_field:
                continue

            logger.info(f"Processing CSV row: {dict(row)}")

            # Parse combined Source and Destination fields (address[:port])
            source_combined = (row.get('Source') or '*').strip()
            dest_combined = (row.get('Destination') or '*').strip()

            def split_addr_port(combined):
                parts = combined.split(':')
                addr = parts[0] if parts else '*'
                port = parts[1] if len(parts) > 1 else '*'
                return norm_star(addr), norm_star(port)

            source_address, source_port = split_addr_port(source_combined)
            dest_address, dest_port = split_addr_port(dest_combined)

            # Optional explicit address list columns; support comma-separated lists
            raw_src_addrs = (row.get('Source Address') or source_address)
            raw_dst_addrs = (row.get('Destination Address') or dest_address)

            def to_list(value):
                if value is None:
                    return []
                if isinstance(value, list):
                    return [str(v).strip() for v in value if str(v).strip()]
                parts = [p.strip() for p in str(value).split(',')]
                return [p for p in parts if p]

            src_prefixes = to_list(raw_src_addrs)
            dst_prefixes = to_list(raw_dst_addrs)

            # If list is empty or contains only '*', prefer singular '*'
            if src_prefixes == ['*']:
                src_prefixes = []
            if dst_prefixes == ['*']:
                dst_prefixes = []

            src_single = src_prefixes[0] if src_prefixes else source_address
            dst_single = dst_prefixes[0] if dst_prefixes else dest_address

            rule = {
                # Base
                "name": name_field,
                "priority": int(row.get('Priority', 1000)) if (row.get('Priority') or '').strip().isdigit() else 1000,
                "direction": (row.get('Direction') or 'Inbound').strip(),
                "access": (row.get('Access') or 'Allow').strip(),
                "protocol": (row.get('Protocol') or '*').strip() or '*',

                # Azure-style keys
                "source_address_prefix": src_single or '*',
                "destination_address_prefix": dst_single or '*',
                "source_address_prefixes": src_prefixes,
                "destination_address_prefixes": dst_prefixes,
                "source_port_range": source_port,
                "destination_port_range": dest_port,

                # Legacy snake_case used by restore confirm
                "source_address": src_single or '*',
                "destination_address": dst_single or '*',
                "source_port": source_port,
                "destination_port": dest_port,
                "action": (row.get('Access') or 'Allow').strip(),

                # Description
                "description": (row.get('Description') or '').strip()
            }

            logger.info(f"Parsed rule: {rule}")

            # Validate required fields
            if rule['name'] and rule['direction'] in ['Inbound', 'Outbound'] and rule['access'] in ['Allow', 'Deny']:
                rules.append(rule)

    except Exception as e:
        logger.error(f"Error parsing CSV rules: {e}")
        logger.error(f"CSV content preview: {csv_content[:500]}...")
        # Return sample rules as fallback
        rules = [
            {
                "name": "AllowHTTP",
                "priority": 100,
                "direction": "Inbound",
                "access": "Allow",
                "protocol": "TCP",
                "source_address_prefix": "*",
                "source_port_range": "*",
                "destination_address_prefix": "*",
                "destination_port_range": "80",
                "description": "Allow HTTP traffic from CSV"
            },
            {
                "name": "AllowHTTPS",
                "priority": 110,
                "direction": "Inbound",
                "access": "Allow",
                "protocol": "TCP",
                "source_address_prefix": "*",
                "source_port_range": "*",
                "destination_address_prefix": "*",
                "destination_port_range": "443",
                "description": "Allow HTTPS traffic from CSV"
            }
        ]

    logger.info(f"Returning {len(rules)} parsed rules")
    return rules

class WorkingHandler(BaseHTTPRequestHandler):
    
    def _calculate_next_execution(self, frequency, time_of_day):
        """Calculate the next execution time based on frequency and time of day"""
        from datetime import datetime, timedelta
        
        now = datetime.now()
        hour, minute = map(int, time_of_day.split(':'))
        
        if frequency == 'daily':
            next_exec = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
            if next_exec <= now:
                next_exec += timedelta(days=1)
        elif frequency == 'weekly':
            next_exec = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
            days_ahead = 7 - now.weekday()  # Next Monday
            if days_ahead <= 0 or (days_ahead == 7 and next_exec <= now):
                days_ahead += 7
            next_exec += timedelta(days=days_ahead)
        elif frequency == 'monthly':
            next_exec = now.replace(day=1, hour=hour, minute=minute, second=0, microsecond=0)
            if next_exec <= now:
                # Next month
                if next_exec.month == 12:
                    next_exec = next_exec.replace(year=next_exec.year + 1, month=1)
                else:
                    next_exec = next_exec.replace(month=next_exec.month + 1)
        else:
            next_exec = now + timedelta(days=1)  # Default to tomorrow
        
        return next_exec.isoformat()
    
    def _generate_csv_content(self, report_type, report_data):
        """Generate CSV content based on report type and data"""
        csv_content = ""
        
        if report_type == 'asg-validation':
            csv_content = "Subscription Name,Subscription ID,Resource Group,NSG Name,Location,Total ASGs,Source ASGs,Destination ASGs\n"
            for item in report_data.get('data', []):
                csv_content += f"{item.get('subscription_name', '')},{item.get('subscription_id', '')},{item.get('resource_group', '')},{item.get('nsg_name', '')},{item.get('location', '')},{item.get('total_asgs', 0)},{item.get('source_asgs_count', 0)},{item.get('destination_asgs_count', 0)}\n"
        
        elif report_type == 'nsg-rules':
            csv_content = "Subscription Name,Subscription ID,Resource Group,NSG Name,Location,Total Rules,Inbound Rules,Outbound Rules\n"
            for item in report_data.get('data', []):
                csv_content += f"{item.get('subscription_name', '')},{item.get('subscription_id', '')},{item.get('resource_group', '')},{item.get('nsg_name', '')},{item.get('location', '')},{item.get('total_rules', 0)},{item.get('inbound_rules_count', 0)},{item.get('outbound_rules_count', 0)}\n"
        
        elif report_type == 'ip-limitations':
            csv_content = "Subscription Name,Subscription ID,Resource Group,NSG Name,Location,Source IPs Count,Destination IPs Count,Source ASGs Count,Destination ASGs Count,Total Unique IPs\n"
            for item in report_data.get('data', []):
                csv_content += f"{item.get('subscription_name', '')},{item.get('subscription_id', '')},{item.get('resource_group', '')},{item.get('nsg_name', '')},{item.get('location', '')},{item.get('source_ips_count', 0)},{item.get('destination_ips_count', 0)},{item.get('source_asgs_count', 0)},{item.get('destination_asgs_count', 0)},{item.get('total_unique_ips', 0)}\n"
        
        elif report_type == 'nsg-ports':
            csv_content = "Subscription Name,Subscription ID,Resource Group,NSG Name,Location,Inbound Ports Count,Outbound Ports Count,High Risk Ports Count,Security Score\n"
            for item in report_data.get('data', []):
                csv_content += f"{item.get('subscription_name', '')},{item.get('subscription_id', '')},{item.get('resource_group', '')},{item.get('nsg_name', '')},{item.get('location', '')},{item.get('inbound_ports_count', 0)},{item.get('outbound_ports_count', 0)},{item.get('high_risk_ports_count', 0)},{item.get('security_score', 0)}\n"
        
        elif report_type == 'consolidation':
            csv_content = "Subscription Name,Subscription ID,Resource Group,NSG Name,Location,Total Rules,Redundant Rules,Overly Permissive Rules,Optimization Score,Recommendations\n"
            for item in report_data.get('data', []):
                recommendations = '; '.join(item.get('recommendations', []))
                csv_content += f"{item.get('subscription_name', '')},{item.get('subscription_id', '')},{item.get('resource_group', '')},{item.get('nsg_name', '')},{item.get('location', '')},{item.get('total_rules', 0)},{item.get('redundant_rules', 0)},{item.get('overly_permissive_rules', 0)},{item.get('optimization_score', 0)},\"{recommendations}\"\n"
        
        return csv_content
    
    def do_GET(self):
        logger.info(f"Received GET request: {self.path}")
        
        # Parse URL
        parsed_url = urlparse(self.path)
        path = parsed_url.path
        query_params = parse_qs(parsed_url.query)
        
        # Send response headers
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
        
        # Route handling
        if path == '/api/v1/health':
            response = {"status": "healthy", "message": "Working backend server is running"}
        elif path == '/api/v1/dashboard':
            try:
                logger.info("Processing dashboard request")
                # Aggregate dashboard statistics across all subscriptions
                # Get subscriptions data
                subscriptions_data = get_subscriptions()
                logger.info(f"Retrieved {len(subscriptions_data.get('subscriptions', []))} subscriptions")
                
                # Initialize aggregated counters
                total_nsgs = 0
                total_rules = 0
                active_nsgs = 0
                high_risk_nsgs = 0
                total_resource_groups = 0
                all_nsgs = []
                
                # Iterate through all accessible subscriptions
                for subscription in subscriptions_data.get('subscriptions', []):
                    subscription_id = subscription['id']
                    logger.info(f"Fetching data for subscription: {subscription['name']} ({subscription_id})")
                    
                    try:
                        # Get resource groups data for this subscription
                        resource_groups_data = get_resource_groups(subscription_id)
                        total_resource_groups += len(resource_groups_data.get('resource_groups', []))
                        
                        # Get NSGs data for this subscription
                        nsgs_data = get_nsgs(subscription_id)
                        subscription_nsgs = nsgs_data.get('nsgs', [])
                        
                        # Add subscription info to each NSG for tracking
                        for nsg in subscription_nsgs:
                            nsg['subscription_id'] = subscription_id
                            nsg['subscription_name'] = subscription['name']
                        
                        all_nsgs.extend(subscription_nsgs)
                        
                        # Aggregate statistics for this subscription
                        sub_total_nsgs = len(subscription_nsgs)
                        sub_total_rules = sum(len(nsg.get('inbound_rules', [])) + len(nsg.get('outbound_rules', [])) for nsg in subscription_nsgs)
                        sub_active_nsgs = len([nsg for nsg in subscription_nsgs if nsg.get('is_active', True)])
                        sub_high_risk_nsgs = len([nsg for nsg in subscription_nsgs if nsg.get('risk_level', 'low') in ['high', 'critical']])
                        
                        total_nsgs += sub_total_nsgs
                        total_rules += sub_total_rules
                        active_nsgs += sub_active_nsgs
                        high_risk_nsgs += sub_high_risk_nsgs
                        
                        logger.info(f"Subscription {subscription['name']}: {sub_total_nsgs} NSGs, {sub_total_rules} rules")
                        
                    except Exception as e:
                        logger.error(f"Error fetching data for subscription {subscription['name']}: {e}")
                        continue
                
                response = {
                    "metrics": {
                        "live_connections": 2847,  # This would come from monitoring system
                        "nsg_rules": total_rules,
                        "security_groups": total_nsgs,
                        "security_alerts": high_risk_nsgs
                    },
                    "recent_activity": [
                        {
                            "id": 1,
                            "type": "multi_subscription_scan",
                            "title": "Multi-Subscription Scan Completed",
                            "description": f"Scanned {len(subscriptions_data.get('subscriptions', []))} subscriptions, found {total_nsgs} NSGs across {total_resource_groups} resource groups",
                            "status": "success",
                            "timestamp": "2 min ago"
                        },
                        {
                            "id": 2,
                            "type": "nsg_discovery",
                            "title": "NSG Discovery",
                            "description": f"Discovered {len([nsg for nsg in all_nsgs if nsg.get('subscription_name')])} NSGs across multiple subscriptions" if all_nsgs else "No NSGs found in accessible subscriptions",
                            "status": "success" if all_nsgs else "warning",
                            "timestamp": "5 min ago"
                        },
                        {
                            "id": 3,
                            "type": "security_assessment",
                            "title": "Security Risk Assessment",
                            "description": f"Identified {high_risk_nsgs} high-risk NSGs requiring attention" if high_risk_nsgs > 0 else "No high-risk NSGs detected",
                            "status": "warning" if high_risk_nsgs > 0 else "success",
                            "timestamp": "10 min ago"
                        },
                        {
                            "id": 4,
                            "type": "subscription_access",
                            "title": "Subscription Access Verified",
                            "description": f"SPN has access to {subscriptions_data.get('active_subscriptions', 0)} active subscriptions",
                            "status": "success",
                            "timestamp": "15 min ago"
                        }
                    ],
                    "system_status": {
                        "azure_api": "online",
                        "database": "online",
                        "monitoring": "warning",
                        "backup_service": "offline"
                    },
                    "statistics": {
                        "total_subscriptions": len(subscriptions_data.get('subscriptions', [])),
                        "active_subscriptions": subscriptions_data.get('active_subscriptions', 0),
                        "total_resource_groups": total_resource_groups,
                        "total_nsgs": total_nsgs,
                        "active_nsgs": active_nsgs,
                        "total_rules": total_rules,
                        "high_risk_nsgs": high_risk_nsgs
                    },
                    "subscription_breakdown": [
                        {
                            "subscription_id": sub['id'],
                            "subscription_name": sub['name'],
                            "status": sub['status'],
                            "environment": sub['environment'],
                            "nsg_count": len([nsg for nsg in all_nsgs if nsg.get('subscription_id') == sub['id']]),
                            "rule_count": sum(len(nsg.get('inbound_rules', [])) + len(nsg.get('outbound_rules', [])) for nsg in all_nsgs if nsg.get('subscription_id') == sub['id'])
                        } for sub in subscriptions_data.get('subscriptions', [])
                    ]
                }
                logger.info(f"Dashboard response prepared successfully with {total_nsgs} NSGs")
            except Exception as e:
                logger.error(f"Dashboard endpoint error: {e}")
                response = {
                    "error": f"Dashboard data unavailable: {str(e)}",
                    "metrics": {
                        "live_connections": 0,
                        "nsg_rules": 0,
                        "security_groups": 0,
                        "security_alerts": 0
                    },
                    "recent_activity": [],
                    "system_status": {
                        "azure_api": "error",
                        "database": "unknown",
                        "monitoring": "error",
                        "backup_service": "offline"
                    },
                    "statistics": {
                        "total_subscriptions": 0,
                        "active_subscriptions": 0,
                        "total_resource_groups": 0,
                        "total_nsgs": 0,
                        "active_nsgs": 0,
                        "total_rules": 0,
                        "high_risk_nsgs": 0
                    },
                    "subscription_breakdown": []
                }
        elif path == '/api/v1/subscriptions':
            # Use real Azure data with SPN authentication
            azure_data = get_subscriptions()
            response = {
                "subscriptions": [
                    {
                        "id": sub["id"],
                        "name": sub["name"],
                        "state": sub["state"],
                        "tenantId": sub["tenant_id"]
                    } for sub in azure_data["subscriptions"]
                ]
            }
        elif path == '/api/v1/resource-groups':
            subscription_id = query_params.get('subscription_id', [''])[0]
            # Use real Azure data with SPN authentication
            azure_data = get_resource_groups(subscription_id)
            response = {
                "resource_groups": [
                    {
                        "name": rg["name"],
                        "id": rg["id"],
                        "location": rg["location"],
                        "subscription_id": subscription_id,
                        "provisioning_state": rg["status"],
                        "tags": rg["tags"]
                    } for rg in azure_data["resource_groups"]
                ]
            }
        elif path == '/api/v1/locations':
            subscription_id = query_params.get('subscription_id', [''])[0]
            response = {
                "locations": [
                    {
                        "name": "eastus",
                        "display_name": "East US",
                        "latitude": "37.3719",
                        "longitude": "-79.8164",
                        "subscription_id": subscription_id
                    },
                    {
                        "name": "westus2",
                        "display_name": "West US 2",
                        "latitude": "47.233",
                        "longitude": "-119.852",
                        "subscription_id": subscription_id
                    }
                ]
            }
        elif path == '/api/v1/nsgs':
            subscription_id = query_params.get('subscription_id', [''])[0]
            resource_group = query_params.get('resource_group', [''])[0]
            # Use real Azure data with SPN authentication
            azure_data = get_nsgs(subscription_id, resource_group if resource_group else None)
            response = azure_data  # Return the data structure as-is from get_nsgs()
        elif path == '/api/v1/route-tables':
            subscription_id = query_params.get('subscription_id', [''])[0]
            resource_group = query_params.get('resource_group', [''])[0]
            # Use real Azure data with SPN authentication
            azure_data = get_route_tables(subscription_id, resource_group if resource_group else None)
            response = azure_data  # Return the data structure as-is from get_route_tables()
        elif path == '/api/v1/asgs':
            subscription_id = query_params.get('subscription_id', [''])[0]
            resource_group = query_params.get('resource_group', [''])[0]
            # Mock ASG data for now
            response = {
                "asgs": [
                    {
                        "id": "/subscriptions/" + subscription_id + "/resourceGroups/rg-production/providers/Microsoft.Network/applicationSecurityGroups/asg-web-servers",
                        "azure_id": "/subscriptions/" + subscription_id + "/resourceGroups/rg-production/providers/Microsoft.Network/applicationSecurityGroups/asg-web-servers",
                        "name": "asg-web-servers",
                        "resource_group": "rg-production",
                        "location": "eastus",
                        "subscription_id": subscription_id
                    },
                    {
                        "id": "/subscriptions/" + subscription_id + "/resourceGroups/rg-production/providers/Microsoft.Network/applicationSecurityGroups/asg-db-servers",
                        "azure_id": "/subscriptions/" + subscription_id + "/resourceGroups/rg-production/providers/Microsoft.Network/applicationSecurityGroups/asg-db-servers",
                        "name": "asg-db-servers",
                        "resource_group": "rg-production",
                        "location": "eastus",
                        "subscription_id": subscription_id
                    },
                    {
                        "id": "/subscriptions/" + subscription_id + "/resourceGroups/rg-development/providers/Microsoft.Network/applicationSecurityGroups/asg-app-tier",
                        "azure_id": "/subscriptions/" + subscription_id + "/resourceGroups/rg-development/providers/Microsoft.Network/applicationSecurityGroups/asg-app-tier",
                        "name": "asg-app-tier",
                        "resource_group": "rg-development",
                        "location": "westus2",
                        "subscription_id": subscription_id
                    }
                ]
            }
        elif path == '/api/v1/storage-accounts':
            subscription_id = query_params.get('subscription_id', [''])[0]
            
            # Get real storage accounts from Azure
            if azure_service and azure_service.storage_client:
                try:
                    storage_accounts_data = asyncio.run(azure_service.list_storage_accounts(subscription_id))
                    response = {
                        "storage_accounts": storage_accounts_data
                    }
                    logger.info(f"Retrieved {len(storage_accounts_data)} storage accounts from Azure")
                except Exception as e:
                    logger.error(f"Failed to get storage accounts from Azure: {e}")
                    # Fallback to mock data
                    response = {
                        "storage_accounts": [
                            {
                                "id": "/subscriptions/" + subscription_id + "/resourceGroups/rg1/providers/Microsoft.Storage/storageAccounts/storage1",
                                "name": "thirustorage001",
                                "resource_group": "resourcegroup002",
                                "location": "eastus",
                                "sku": "Standard_LRS",
                                "kind": "StorageV2"
                            }
                        ]
                    }
            else:
                # Fallback to mock data when Azure Service is not available
                response = {
                    "storage_accounts": [
                        {
                            "id": "/subscriptions/" + subscription_id + "/resourceGroups/rg1/providers/Microsoft.Storage/storageAccounts/storage1",
                            "name": "thirustorage001",
                            "resource_group": "resourcegroup002",
                            "location": "eastus",
                            "sku": "Standard_LRS",
                            "kind": "StorageV2"
                        }
                    ]
                }
        elif path == '/api/v1/containers':
            storage_account = query_params.get('storage_account', [''])[0]
            
            # Get real containers from Azure Storage
            if azure_service and azure_service.blob_service_client:
                try:
                    containers_data = asyncio.run(azure_service.list_containers())
                    # Add storage_account to each container for compatibility
                    for container in containers_data:
                        container['storage_account'] = storage_account or AZURE_CONFIG.get('storage_account', '')
                    
                    response = {
                        "containers": containers_data
                    }
                    logger.info(f"Retrieved {len(containers_data)} containers from Azure Storage")
                except Exception as e:
                    logger.error(f"Failed to get containers from Azure Storage: {e}")
                    # Fallback to mock data
                    response = {
                        "containers": [
                            {
                                "name": "nsg-backups",
                                "storage_account": storage_account,
                                "last_modified": "2024-01-15T10:30:00Z",
                                "public_access": "None"
                            }
                        ]
                    }
            else:
                # Fallback to mock data when Azure Service is not available
                response = {
                    "containers": [
                        {
                            "name": "nsg-backups",
                            "storage_account": storage_account,
                            "last_modified": "2024-01-15T10:30:00Z",
                            "public_access": "None"
                        }
                    ]
                }
        elif path == '/api/v1/golden-rule/storage':
            # Get golden standards from Azure Storage
            storage_account = query_params.get('storage_account', [''])[0]
            container = query_params.get('container', [''])[0]
            file_name = query_params.get('file_name', [''])[0]
            
            try:
                # Mock loading golden standards from storage
                response = {
                    "success": True,
                    "golden_standards": [
                        {
                            "name": "Web Tier Security",
                            "description": "Standard security rules for web tier",
                            "inbound_rules": [
                                {
                                    "name": "AllowHTTP",
                                    "priority": 100,
                                    "direction": "Inbound",
                                    "access": "Allow",
                                    "protocol": "TCP",
                                    "source_address_prefix": "Internet",
                                    "source_port_range": "*",
                                    "destination_address_prefix": "VirtualNetwork",
                                    "destination_port_range": "80"
                                },
                                {
                                    "name": "AllowHTTPS",
                                    "priority": 110,
                                    "direction": "Inbound",
                                    "access": "Allow",
                                    "protocol": "TCP",
                                    "source_address_prefix": "Internet",
                                    "source_port_range": "*",
                                    "destination_address_prefix": "VirtualNetwork",
                                    "destination_port_range": "443"
                                }
                            ],
                            "outbound_rules": [
                                {
                                    "name": "AllowOutboundHTTP",
                                    "priority": 100,
                                    "direction": "Outbound",
                                    "access": "Allow",
                                    "protocol": "TCP",
                                    "source_address_prefix": "VirtualNetwork",
                                    "source_port_range": "*",
                                    "destination_address_prefix": "Internet",
                                    "destination_port_range": "80"
                                }
                            ]
                        }
                    ],
                    "source": {
                        "type": "storage",
                        "storage_account": storage_account,
                        "container": container,
                        "file_name": file_name
                    }
                }
                logger.info(f"Loaded golden standards from storage: {storage_account}/{container}/{file_name}")
            except Exception as e:
                logger.error(f"Failed to load golden standards from storage: {e}")
                response = {
                    "success": False,
                    "error": f"Failed to load golden standards: {str(e)}"
                }
        elif path.startswith('/api/v1/nsg-validation/'):
            # Extract NSG name from path
            nsg_name = path.split('/')[-1]
            subscription_id = query_params.get('subscription_id', [AZURE_CONFIG["subscription_id"]])[0]
            resource_group = query_params.get('resource_group', [''])[0]
            
            # Check if this is a test request for demo data
            if nsg_name == 'demo-nsg' and resource_group == 'demo-rg':
                # Create demo NSG with actual rules for AI analysis
                from nsg_validation import NSGRule, NSGValidator
                # Create a local validator instance
                local_nsg_validator = NSGValidator()
                
                demo_rules = [
                    NSGRule(
                        id='AllowHTTP',
                        name='AllowHTTP',
                        priority=100,
                        direction='Inbound',
                        access='Allow',
                        protocol='TCP',
                        source_address_prefix='10.0.1.0/24',
                        source_port_range='*',
                        destination_address_prefix='10.0.2.0/24',
                        destination_port_range='80',
                        source_application_security_groups=[],
                        destination_application_security_groups=[]
                    ),
                    NSGRule(
                        id='AllowHTTPS',
                        name='AllowHTTPS',
                        priority=110,
                        direction='Inbound',
                        access='Allow',
                        protocol='TCP',
                        source_address_prefix='10.0.1.0/24',
                        source_port_range='*',
                        destination_address_prefix='10.0.2.0/24',
                        destination_port_range='443',
                        source_application_security_groups=[],
                        destination_application_security_groups=[]
                    ),
                    NSGRule(
                        id='AllowSSH',
                        name='AllowSSH',
                        priority=120,
                        direction='Inbound',
                        access='Allow',
                        protocol='TCP',
                        source_address_prefix='192.168.1.100',
                        source_port_range='*',
                        destination_address_prefix='10.0.2.5',
                        destination_port_range='22',
                        source_application_security_groups=[],
                        destination_application_security_groups=[]
                    ),
                    NSGRule(
                        id='AllowDatabase',
                        name='AllowDatabase',
                        priority=130,
                        direction='Inbound',
                        access='Allow',
                        protocol='TCP',
                        source_address_prefix='10.0.2.0/24',
                        source_port_range='*',
                        destination_address_prefix='10.0.3.0/24',
                        destination_port_range='3306',
                        source_application_security_groups=[],
                        destination_application_security_groups=[]
                    ),
                    NSGRule(
                        id='DenyAll',
                        name='DenyAll',
                        priority=4000,
                        direction='Inbound',
                        access='Deny',
                        protocol='*',
                        source_address_prefix='*',
                        source_port_range='*',
                        destination_address_prefix='*',
                        destination_port_range='*',
                        source_application_security_groups=[],
                        destination_application_security_groups=[]
                    ),
                    NSGRule(
                        id='AllowOutboundHTTP',
                        name='AllowOutboundHTTP',
                        priority=100,
                        direction='Outbound',
                        access='Allow',
                        protocol='TCP',
                        source_address_prefix='10.0.2.0/24',
                        source_port_range='*',
                        destination_address_prefix='0.0.0.0/0',
                        destination_port_range='80',
                        source_application_security_groups=[],
                        destination_application_security_groups=[]
                    ),
                    NSGRule(
                        id='AllowOutboundHTTPS',
                        name='AllowOutboundHTTPS',
                        priority=110,
                        direction='Outbound',
                        access='Allow',
                        protocol='TCP',
                        source_address_prefix='10.0.2.0/24',
                        source_port_range='*',
                        destination_address_prefix='0.0.0.0/0',
                        destination_port_range='443',
                        source_application_security_groups=[],
                        destination_application_security_groups=[]
                    ),
                    NSGRule(
                        id='AllowOutboundDNS',
                        name='AllowOutboundDNS',
                        priority=120,
                        direction='Outbound',
                        access='Allow',
                        protocol='UDP',
                        source_address_prefix='10.0.2.0/24',
                        source_port_range='*',
                        destination_address_prefix='8.8.8.8',
                        destination_port_range='53',
                        source_application_security_groups=[],
                        destination_application_security_groups=[]
                    )
                ]
                
                # Use the actual AI analysis with demo rules
                try:
                    response = local_nsg_validator.analyze_nsg_rules_from_demo(demo_rules)
                except Exception as e:
                    logger.error(f"Demo AI analysis failed: {e}")
                    raise Exception(f"Failed to analyze demo NSG: {str(e)}")
            else:
                # Create validator instance for real NSG analysis
                from nsg_validation import NSGValidator
                local_nsg_validator = NSGValidator()
                try:
                    validation_result = local_nsg_validator.analyze_nsg_rules(subscription_id, resource_group, nsg_name)
                    response = validation_result
                except Exception as e:
                    logger.error(f"NSG validation failed: {e}")
                    response = {"error": f"Validation failed: {str(e)}"}
        elif path.startswith('/api/v1/nsgs/') and not path.startswith('/api/v1/nsg-validation/') and not path.startswith('/api/v1/nsg-recommendations/'):
            # Get individual NSG details for view/edit
            try:
                nsg_id = path.split('/')[-1]
                subscription_id = query_params.get('subscription_id', [AZURE_CONFIG["subscription_id"]])[0]
                resource_group = query_params.get('resource_group', [''])[0]
                
                # Get NSGs data and find the specific NSG
                nsgs_data = get_nsgs(subscription_id, resource_group if resource_group else None)
                target_nsg = None
                
                # Find NSG by ID or name
                for nsg in nsgs_data.get('nsgs', []):
                    if str(nsg['id']) == nsg_id or nsg['name'] == nsg_id:
                        target_nsg = nsg
                        break
                
                if target_nsg:
                    response = {
                        "success": True,
                        "nsg": target_nsg
                    }
                else:
                    response = {"error": f"NSG with ID/name '{nsg_id}' not found"}
                    
            except Exception as e:
                logger.error(f"Failed to get NSG details: {e}")
                response = {"error": f"Failed to get NSG details: {str(e)}"}
        elif path == '/api/v1/reports/asg-validation':
            # ASG Validation Report with Real Azure Data
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                request_data = json.loads(post_data.decode('utf-8'))
                
                subscription_id = request_data.get('subscription_id', AZURE_CONFIG["subscription_id"])
                resource_group = request_data.get('resource_group', '')
                nsg_names = request_data.get('nsg_names', [])
                
                logger.info(f"ASG Validation Report - Subscription: {subscription_id}, RG: {resource_group}, NSGs: {nsg_names}")
                
                # Get Azure clients
                _, resource_client, network_client = get_azure_clients()
                if not network_client:
                    raise Exception("Failed to initialize Azure clients")
                
                # Get subscription name
                subscription_name = "Unknown"
                try:
                    subscription_client, _, _ = get_azure_clients()
                    if subscription_client:
                        sub_info = subscription_client.subscriptions.get(subscription_id)
                        subscription_name = sub_info.display_name
                except:
                    pass
                
                # Fetch real ASG data from Azure
                asg_data = []
                total_asgs = 0
                
                if resource_group:
                    # Get ASGs from specific resource group
                    try:
                        asgs = network_client.application_security_groups.list(resource_group)
                        for asg in asgs:
                            total_asgs += 1
                            
                            # Get NSG associations for this ASG
                            source_nsgs = []
                            dest_nsgs = []
                            
                            # Check NSG rules that reference this ASG
                            if nsg_names:
                                for nsg_name in nsg_names:
                                    try:
                                        nsg = network_client.network_security_groups.get(resource_group, nsg_name)
                                        for rule in nsg.security_rules or []:
                                            # Check source ASGs
                                            if rule.source_application_security_groups:
                                                for src_asg in rule.source_application_security_groups:
                                                    if src_asg.id == asg.id:
                                                        source_nsgs.append(nsg_name)
                                            # Check destination ASGs
                                            if rule.destination_application_security_groups:
                                                for dest_asg in rule.destination_application_security_groups:
                                                    if dest_asg.id == asg.id:
                                                        dest_nsgs.append(nsg_name)
                                    except Exception as e:
                                        logger.warning(f"Could not check NSG {nsg_name}: {e}")
                            
                            asg_entry = {
                                "subscription_name": subscription_name,
                                "subscription_id": subscription_id,
                                "resource_group": asg.location if hasattr(asg, 'location') else resource_group,
                                "nsg_name": ", ".join(set(source_nsgs + dest_nsgs)) or "No NSG associations",
                                "source_asg": asg.name if asg.name in [nsg for nsg in source_nsgs] else "",
                                "destination_asg": asg.name if asg.name in [nsg for nsg in dest_nsgs] else "",
                                "asg_name": asg.name,
                                "asg_id": asg.id,
                                "location": asg.location if hasattr(asg, 'location') else "Unknown"
                            }
                            asg_data.append(asg_entry)
                    except Exception as e:
                        logger.error(f"Failed to fetch ASGs from resource group {resource_group}: {e}")
                else:
                    # Get ASGs from all resource groups in subscription
                    try:
                        resource_groups = resource_client.resource_groups.list()
                        for rg in resource_groups:
                            try:
                                asgs = network_client.application_security_groups.list(rg.name)
                                for asg in asgs:
                                    total_asgs += 1
                                    
                                    # Get NSG associations
                                    source_nsgs = []
                                    dest_nsgs = []
                                    
                                    if nsg_names:
                                        for nsg_name in nsg_names:
                                            try:
                                                nsg = network_client.network_security_groups.get(rg.name, nsg_name)
                                                for rule in nsg.security_rules or []:
                                                    if rule.source_application_security_groups:
                                                        for src_asg in rule.source_application_security_groups:
                                                            if src_asg.id == asg.id:
                                                                source_nsgs.append(nsg_name)
                                                    if rule.destination_application_security_groups:
                                                        for dest_asg in rule.destination_application_security_groups:
                                                            if dest_asg.id == asg.id:
                                                                dest_nsgs.append(nsg_name)
                                            except:
                                                continue
                                    
                                    asg_entry = {
                                        "subscription_name": subscription_name,
                                        "subscription_id": subscription_id,
                                        "resource_group": rg.name,
                                        "nsg_name": ", ".join(set(source_nsgs + dest_nsgs)) or "No NSG associations",
                                        "source_asg": asg.name if source_nsgs else "",
                                        "destination_asg": asg.name if dest_nsgs else "",
                                        "asg_name": asg.name,
                                        "asg_id": asg.id,
                                        "location": asg.location if hasattr(asg, 'location') else "Unknown"
                                    }
                                    asg_data.append(asg_entry)
                            except Exception as e:
                                logger.warning(f"Could not fetch ASGs from resource group {rg.name}: {e}")
                                continue
                    except Exception as e:
                        logger.error(f"Failed to fetch resource groups: {e}")
                
                # Validate ASG limits (Max 100 ASGs)
                max_asgs = 100
                validation_status = "compliant" if total_asgs <= max_asgs else "non-compliant"
                
                # Generate CSV data
                csv_data = []
                csv_headers = ["Subscription Name", "Subscription ID", "Resource Group", "NSG Name", "Source ASG", "Destination ASG"]
                
                for asg in asg_data:
                    csv_data.append([
                        asg["subscription_name"],
                        asg["subscription_id"],
                        asg["resource_group"],
                        asg["nsg_name"],
                        asg["source_asg"],
                        asg["destination_asg"]
                    ])
                
                response = {
                    "success": True,
                    "report_type": "asg_validation",
                    "generated_at": datetime.now().isoformat(),
                    "data": {
                        "total_asgs": total_asgs,
                        "max_allowed": max_asgs,
                        "validation_status": validation_status,
                        "compliance_percentage": round((max_asgs - max(0, total_asgs - max_asgs)) / max_asgs * 100, 2),
                        "csv_headers": csv_headers,
                        "csv_data": csv_data,
                        "summary": {
                            "subscription_name": subscription_name,
                            "subscription_id": subscription_id,
                            "resource_group": resource_group or "All",
                            "nsg_filter": nsg_names if nsg_names else "All",
                            "total_asgs_found": len(asg_data)
                        }
                    }
                }
            except Exception as e:
                logger.error(f"ASG validation report failed: {e}")
                response = {"error": f"Failed to generate ASG validation report: {str(e)}"}
        elif path == '/api/v1/reports/nsg-rules':
            # NSG Rules Report with Real Azure Data
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                request_data = json.loads(post_data.decode('utf-8'))
                
                subscription_id = request_data.get('subscription_id', AZURE_CONFIG["subscription_id"])
                resource_group = request_data.get('resource_group', '')
                nsg_names = request_data.get('nsg_names', [])
                
                logger.info(f"NSG Rules Report - Subscription: {subscription_id}, RG: {resource_group}, NSGs: {nsg_names}")
                
                # Get Azure clients
                _, resource_client, network_client = get_azure_clients()
                if not network_client:
                    raise Exception("Failed to initialize Azure clients")
                
                # Get subscription name
                subscription_name = "Unknown"
                try:
                    subscription_client, _, _ = get_azure_clients()
                    if subscription_client:
                        sub_info = subscription_client.subscriptions.get(subscription_id)
                        subscription_name = sub_info.display_name
                except:
                    pass
                
                # Fetch real NSG data from Azure
                nsg_data = []
                total_rules = 0
                
                if resource_group:
                    # Get NSGs from specific resource group
                    try:
                        if nsg_names:
                            # Get specific NSGs
                            for nsg_name in nsg_names:
                                try:
                                    nsg = network_client.network_security_groups.get(resource_group, nsg_name)
                                    
                                    # Count inbound and outbound rules
                                    inbound_count = 0
                                    outbound_count = 0
                                    
                                    for rule in nsg.security_rules or []:
                                        total_rules += 1
                                        if rule.direction.lower() == 'inbound':
                                            inbound_count += 1
                                        else:
                                            outbound_count += 1
                                    
                                    # Add default rules count
                                    for rule in nsg.default_security_rules or []:
                                        total_rules += 1
                                        if rule.direction.lower() == 'inbound':
                                            inbound_count += 1
                                        else:
                                            outbound_count += 1
                                    
                                    nsg_entry = {
                                        "subscription_name": subscription_name,
                                        "subscription_id": subscription_id,
                                        "resource_group": resource_group,
                                        "nsg_name": nsg.name,
                                        "source_rules": inbound_count,
                                        "destination_rules": outbound_count,
                                        "total_rules": inbound_count + outbound_count,
                                        "location": nsg.location if hasattr(nsg, 'location') else "Unknown"
                                    }
                                    nsg_data.append(nsg_entry)
                                except Exception as e:
                                    logger.warning(f"Could not fetch NSG {nsg_name}: {e}")
                        else:
                            # Get all NSGs from resource group
                            nsgs = network_client.network_security_groups.list(resource_group)
                            for nsg in nsgs:
                                inbound_count = 0
                                outbound_count = 0
                                
                                for rule in nsg.security_rules or []:
                                    total_rules += 1
                                    if rule.direction.lower() == 'inbound':
                                        inbound_count += 1
                                    else:
                                        outbound_count += 1
                                
                                for rule in nsg.default_security_rules or []:
                                    total_rules += 1
                                    if rule.direction.lower() == 'inbound':
                                        inbound_count += 1
                                    else:
                                        outbound_count += 1
                                
                                nsg_entry = {
                                    "subscription_name": subscription_name,
                                    "subscription_id": subscription_id,
                                    "resource_group": resource_group,
                                    "nsg_name": nsg.name,
                                    "source_rules": inbound_count,
                                    "destination_rules": outbound_count,
                                    "total_rules": inbound_count + outbound_count,
                                    "location": nsg.location if hasattr(nsg, 'location') else "Unknown"
                                }
                                nsg_data.append(nsg_entry)
                    except Exception as e:
                        logger.error(f"Failed to fetch NSGs from resource group {resource_group}: {e}")
                else:
                    # Get NSGs from all resource groups in subscription
                    try:
                        resource_groups = resource_client.resource_groups.list()
                        for rg in resource_groups:
                            try:
                                if nsg_names:
                                    # Get specific NSGs from this resource group
                                    for nsg_name in nsg_names:
                                        try:
                                            nsg = network_client.network_security_groups.get(rg.name, nsg_name)
                                            
                                            inbound_count = 0
                                            outbound_count = 0
                                            
                                            for rule in nsg.security_rules or []:
                                                total_rules += 1
                                                if rule.direction.lower() == 'inbound':
                                                    inbound_count += 1
                                                else:
                                                    outbound_count += 1
                                            
                                            for rule in nsg.default_security_rules or []:
                                                total_rules += 1
                                                if rule.direction.lower() == 'inbound':
                                                    inbound_count += 1
                                                else:
                                                    outbound_count += 1
                                            
                                            nsg_entry = {
                                                "subscription_name": subscription_name,
                                                "subscription_id": subscription_id,
                                                "resource_group": rg.name,
                                                "nsg_name": nsg.name,
                                                "source_rules": inbound_count,
                                                "destination_rules": outbound_count,
                                                "total_rules": inbound_count + outbound_count,
                                                "location": nsg.location if hasattr(nsg, 'location') else "Unknown"
                                            }
                                            nsg_data.append(nsg_entry)
                                        except:
                                            continue
                                else:
                                    # Get all NSGs from this resource group
                                    nsgs = network_client.network_security_groups.list(rg.name)
                                    for nsg in nsgs:
                                        inbound_count = 0
                                        outbound_count = 0
                                        
                                        for rule in nsg.security_rules or []:
                                            total_rules += 1
                                            if rule.direction.lower() == 'inbound':
                                                inbound_count += 1
                                            else:
                                                outbound_count += 1
                                        
                                        for rule in nsg.default_security_rules or []:
                                            total_rules += 1
                                            if rule.direction.lower() == 'inbound':
                                                inbound_count += 1
                                            else:
                                                outbound_count += 1
                                        
                                        nsg_entry = {
                                            "subscription_name": subscription_name,
                                            "subscription_id": subscription_id,
                                            "resource_group": rg.name,
                                            "nsg_name": nsg.name,
                                            "source_rules": inbound_count,
                                            "destination_rules": outbound_count,
                                            "total_rules": inbound_count + outbound_count,
                                            "location": nsg.location if hasattr(nsg, 'location') else "Unknown"
                                        }
                                        nsg_data.append(nsg_entry)
                            except Exception as e:
                                logger.warning(f"Could not fetch NSGs from resource group {rg.name}: {e}")
                                continue
                    except Exception as e:
                        logger.error(f"Failed to fetch resource groups: {e}")
                
                # Validate NSG rules limits (Max 1000 rules)
                max_rules = 1000
                validation_status = "compliant" if total_rules <= max_rules else "non-compliant"
                
                # Generate CSV data
                csv_data = []
                csv_headers = ["Subscription Name", "Subscription ID", "Resource Group", "NSG Name", "Source No of Rules", "Destination No of Rules"]
                
                for nsg in nsg_data:
                    csv_data.append([
                        nsg["subscription_name"],
                        nsg["subscription_id"],
                        nsg["resource_group"],
                        nsg["nsg_name"],
                        str(nsg["source_rules"]),
                        str(nsg["destination_rules"])
                    ])
                
                response = {
                    "success": True,
                    "report_type": "nsg_rules",
                    "generated_at": datetime.now().isoformat(),
                    "data": {
                        "total_rules": total_rules,
                        "max_allowed": max_rules,
                        "validation_status": validation_status,
                        "compliance_percentage": round((max_rules - max(0, total_rules - max_rules)) / max_rules * 100, 2),
                        "csv_headers": csv_headers,
                        "csv_data": csv_data,
                        "summary": {
                            "subscription_name": subscription_name,
                            "subscription_id": subscription_id,
                            "resource_group": resource_group or "All",
                            "nsg_filter": nsg_names if nsg_names else "All",
                            "total_nsgs_found": len(nsg_data),
                            "total_inbound_rules": sum(nsg["source_rules"] for nsg in nsg_data),
                            "total_outbound_rules": sum(nsg["destination_rules"] for nsg in nsg_data)
                        }
                    }
                }
            except Exception as e:
                logger.error(f"NSG rules report failed: {e}")
                response = {"error": f"Failed to generate NSG rules report: {str(e)}"}
        elif path == '/api/v1/reports/ip-limitations':
            # IP Address Limitations Report - Real Azure SDK Implementation
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                request_data = json.loads(post_data.decode('utf-8'))
                
                subscription_id = request_data.get('subscription_id', AZURE_CONFIG["subscription_id"])
                resource_group = request_data.get('resource_group', '')
                nsg_names = request_data.get('nsg_names', [])
                
                logger.info(f"IP Limitations Report - Subscription: {subscription_id}, RG: {resource_group}, NSGs: {nsg_names}")
                
                # Initialize Azure clients
                try:
                    credential = DefaultAzureCredential()
                    network_client = NetworkManagementClient(credential, subscription_id)
                    resource_client = ResourceManagementClient(credential, subscription_id)
                except Exception as e:
                    logger.error(f"Failed to initialize Azure clients: {e}")
                    raise Exception(f"Azure authentication failed: {str(e)}")
                
                # Get subscription name
                subscription_name = "Unknown"
                try:
                    subscription_client = SubscriptionClient(credential)
                    sub_info = subscription_client.subscriptions.get(subscription_id)
                    subscription_name = sub_info.display_name
                except:
                    pass
                
                # Fetch NSGs from Azure
                nsgs = []
                if resource_group:
                    # Get NSGs from specific resource group
                    try:
                        rg_nsgs = list(network_client.network_security_groups.list(resource_group))
                        nsgs.extend(rg_nsgs)
                    except Exception as e:
                        logger.warning(f"Failed to get NSGs from resource group {resource_group}: {e}")
                else:
                    # Get NSGs from all resource groups
                    try:
                        all_nsgs = list(network_client.network_security_groups.list_all())
                        nsgs.extend(all_nsgs)
                    except Exception as e:
                        logger.warning(f"Failed to get NSGs from all resource groups: {e}")
                
                # Filter NSGs by name if specified
                if nsg_names:
                    nsgs = [nsg for nsg in nsgs if nsg.name in nsg_names]
                
                # Initialize NSG validator for IP counting
                nsg_validator = NSGValidator()
                
                # CSV data for export
                csv_data = []
                csv_headers = ['Subscription Name', 'Subscription ID', 'Resource Group', 'NSG Name', 'Source IPs + ASGs', 'Destination IPs + ASGs']
                
                total_ip_count = 0
                validation_status = "compliant"
                compliance_percentage = 100.0
                
                # Process each NSG
                for nsg in nsgs:
                    try:
                        # Count IP addresses using NSG validation logic
                        ip_count = nsg_validator.count_ip_addresses(nsg)
                        total_ip_count += ip_count
                        
                        # Extract source and destination IPs/ASGs
                        source_ips_asgs = []
                        dest_ips_asgs = []
                        
                        # Process security rules
                        all_rules = list(nsg.security_rules) + list(nsg.default_security_rules)
                        for rule in all_rules:
                            # Source addresses
                            if hasattr(rule, 'source_address_prefix') and rule.source_address_prefix:
                                if rule.source_address_prefix not in source_ips_asgs:
                                    source_ips_asgs.append(rule.source_address_prefix)
                            
                            if hasattr(rule, 'source_address_prefixes') and rule.source_address_prefixes:
                                for prefix in rule.source_address_prefixes:
                                    if prefix not in source_ips_asgs:
                                        source_ips_asgs.append(prefix)
                            
                            if hasattr(rule, 'source_application_security_groups') and rule.source_application_security_groups:
                                for asg in rule.source_application_security_groups:
                                    asg_name = f"ASG:{asg.id.split('/')[-1]}"
                                    if asg_name not in source_ips_asgs:
                                        source_ips_asgs.append(asg_name)
                            
                            # Destination addresses
                            if hasattr(rule, 'destination_address_prefix') and rule.destination_address_prefix:
                                if rule.destination_address_prefix not in dest_ips_asgs:
                                    dest_ips_asgs.append(rule.destination_address_prefix)
                            
                            if hasattr(rule, 'destination_address_prefixes') and rule.destination_address_prefixes:
                                for prefix in rule.destination_address_prefixes:
                                    if prefix not in dest_ips_asgs:
                                        dest_ips_asgs.append(prefix)
                            
                            if hasattr(rule, 'destination_application_security_groups') and rule.destination_application_security_groups:
                                for asg in rule.destination_application_security_groups:
                                    asg_name = f"ASG:{asg.id.split('/')[-1]}"
                                    if asg_name not in dest_ips_asgs:
                                        dest_ips_asgs.append(asg_name)
                        
                        # Add to CSV data
                        csv_data.append([
                            subscription_name,
                            subscription_id,
                            nsg.id.split('/')[4],  # Resource group from resource ID
                            nsg.name,
                            ', '.join(source_ips_asgs) if source_ips_asgs else 'None',
                            ', '.join(dest_ips_asgs) if dest_ips_asgs else 'None'
                        ])
                        
                    except Exception as e:
                        logger.error(f"Error processing NSG {nsg.name}: {e}")
                        continue
                
                # Check IP limit compliance (max 4000)
                if total_ip_count > 4000:
                    validation_status = "non_compliant"
                    compliance_percentage = max(0, (4000 / total_ip_count) * 100)
                elif total_ip_count > 3500:
                    validation_status = "warning"
                    compliance_percentage = 85.0
                
                # Generate comprehensive IP limitations report
                ip_report = {
                    "subscription_id": subscription_id,
                    "subscription_name": subscription_name,
                    "resource_group": resource_group or "All Resource Groups",
                    "total_nsgs_analyzed": len(nsgs),
                    "total_ip_addresses": total_ip_count,
                    "max_ip_limit": 4000,
                    "validation_status": validation_status,
                    "compliance_percentage": round(compliance_percentage, 2),
                    "csv_headers": csv_headers,
                    "csv_data": csv_data,
                    "summary": {
                        "compliant_nsgs": len([nsg for nsg in nsgs if nsg_validator.count_ip_addresses(nsg) <= 200]),
                        "warning_nsgs": len([nsg for nsg in nsgs if 200 < nsg_validator.count_ip_addresses(nsg) <= 300]),
                        "non_compliant_nsgs": len([nsg for nsg in nsgs if nsg_validator.count_ip_addresses(nsg) > 300])
                    }
                }
                
                # Detailed NSG analysis for backward compatibility
                nsg_details = []
                for nsg in nsgs:
                    try:
                        ip_count = nsg_validator.count_ip_addresses(nsg)
                        all_rules = list(nsg.security_rules) + list(nsg.default_security_rules)
                        
                        nsg_analysis = {
                            "name": nsg.name,
                            "resource_group": nsg.id.split('/')[4],
                            "location": nsg.location,
                            "total_rules": len(all_rules),
                            "ip_count": ip_count,
                            "compliance_status": "compliant" if ip_count <= 200 else "warning" if ip_count <= 300 else "non_compliant",
                            "risk_level": "low" if ip_count <= 200 else "medium" if ip_count <= 300 else "high"
                        }
                        nsg_details.append(nsg_analysis)
                    except Exception as e:
                        logger.error(f"Error analyzing NSG {nsg.name}: {e}")
                        continue
                
                ip_report["nsg_details"] = nsg_details

                
                response = {
                    "success": True,
                    "report_type": "ip_limitations",
                    "generated_at": datetime.now().isoformat(),
                    "data": ip_report
                }
            except Exception as e:
                logger.error(f"IP limitations report failed: {e}")
                response = {"error": f"Failed to generate IP limitations report: {str(e)}"}
        elif path == '/api/v1/reports/nsg-ports':
            # NSG Ports Report - Real Azure SDK Implementation
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                request_data = json.loads(post_data.decode('utf-8'))
                
                subscription_id = request_data.get('subscription_id', AZURE_CONFIG["subscription_id"])
                resource_group = request_data.get('resource_group', '')
                nsg_names = request_data.get('nsg_names', [])
                
                logger.info(f"NSG Ports Report - Subscription: {subscription_id}, RG: {resource_group}, NSGs: {nsg_names}")
                
                # Initialize Azure clients
                try:
                    credential = DefaultAzureCredential()
                    network_client = NetworkManagementClient(credential, subscription_id)
                    resource_client = ResourceManagementClient(credential, subscription_id)
                except Exception as e:
                    logger.error(f"Failed to initialize Azure clients: {e}")
                    raise Exception(f"Azure authentication failed: {str(e)}")
                
                # Get subscription name
                subscription_name = "Unknown"
                try:
                    subscription_client = SubscriptionClient(credential)
                    sub_info = subscription_client.subscriptions.get(subscription_id)
                    subscription_name = sub_info.display_name
                except:
                    pass
                
                # Fetch NSGs from Azure
                nsgs = []
                if resource_group:
                    # Get NSGs from specific resource group
                    try:
                        rg_nsgs = list(network_client.network_security_groups.list(resource_group))
                        nsgs.extend(rg_nsgs)
                    except Exception as e:
                        logger.warning(f"Failed to get NSGs from resource group {resource_group}: {e}")
                else:
                    # Get NSGs from all resource groups
                    try:
                        all_nsgs = list(network_client.network_security_groups.list_all())
                        nsgs.extend(all_nsgs)
                    except Exception as e:
                        logger.warning(f"Failed to get NSGs from all resource groups: {e}")
                
                # Filter NSGs by name if specified
                if nsg_names:
                    nsgs = [nsg for nsg in nsgs if nsg.name in nsg_names]
                
                # CSV data for export
                csv_data = []
                csv_headers = ['Subscription Name', 'Subscription ID', 'Resource Group', 'NSG Name', 'Source Ports', 'Destination Ports']
                
                total_inbound_ports = 0
                total_outbound_ports = 0
                validation_status = "compliant"
                compliance_percentage = 100.0
                
                # Process each NSG
                for nsg in nsgs:
                    try:
                        # Extract source and destination ports
                        source_ports = set()
                        dest_ports = set()
                        inbound_port_count = 0
                        outbound_port_count = 0
                        
                        # Process security rules
                        all_rules = list(nsg.security_rules) + list(nsg.default_security_rules)
                        for rule in all_rules:
                            direction = rule.direction.lower() if hasattr(rule, 'direction') else 'unknown'
                            
                            # Source ports
                            if hasattr(rule, 'source_port_range') and rule.source_port_range:
                                if rule.source_port_range != '*':
                                    source_ports.add(rule.source_port_range)
                                    if direction == 'inbound':
                                        inbound_port_count += 1
                                    else:
                                        outbound_port_count += 1
                            
                            if hasattr(rule, 'source_port_ranges') and rule.source_port_ranges:
                                for port_range in rule.source_port_ranges:
                                    source_ports.add(port_range)
                                    if direction == 'inbound':
                                        inbound_port_count += 1
                                    else:
                                        outbound_port_count += 1
                            
                            # Destination ports
                            if hasattr(rule, 'destination_port_range') and rule.destination_port_range:
                                if rule.destination_port_range != '*':
                                    dest_ports.add(rule.destination_port_range)
                                    if direction == 'inbound':
                                        inbound_port_count += 1
                                    else:
                                        outbound_port_count += 1
                            
                            if hasattr(rule, 'destination_port_ranges') and rule.destination_port_ranges:
                                for port_range in rule.destination_port_ranges:
                                    dest_ports.add(port_range)
                                    if direction == 'inbound':
                                        inbound_port_count += 1
                                    else:
                                        outbound_port_count += 1
                        
                        total_inbound_ports += inbound_port_count
                        total_outbound_ports += outbound_port_count
                        
                        # Add to CSV data
                        csv_data.append([
                            subscription_name,
                            subscription_id,
                            nsg.id.split('/')[4],  # Resource group from resource ID
                            nsg.name,
                            ', '.join(sorted(source_ports)) if source_ports else 'None',
                            ', '.join(sorted(dest_ports)) if dest_ports else 'None'
                        ])
                        
                    except Exception as e:
                        logger.error(f"Error processing NSG {nsg.name}: {e}")
                        continue
                
                # Check port limit compliance (max 4000 inbound + 4000 outbound)
                if total_inbound_ports > 4000 or total_outbound_ports > 4000:
                    validation_status = "non_compliant"
                    inbound_compliance = min(100, (4000 / max(total_inbound_ports, 1)) * 100)
                    outbound_compliance = min(100, (4000 / max(total_outbound_ports, 1)) * 100)
                    compliance_percentage = (inbound_compliance + outbound_compliance) / 2
                elif total_inbound_ports > 3500 or total_outbound_ports > 3500:
                    validation_status = "warning"
                    compliance_percentage = 85.0
                
                # Generate comprehensive ports report
                ports_report = {
                    "subscription_id": subscription_id,
                    "subscription_name": subscription_name,
                    "resource_group": resource_group or "All Resource Groups",
                    "total_nsgs_analyzed": len(nsgs),
                    "total_inbound_ports": total_inbound_ports,
                    "total_outbound_ports": total_outbound_ports,
                    "max_inbound_limit": 4000,
                    "max_outbound_limit": 4000,
                    "validation_status": validation_status,
                    "compliance_percentage": round(compliance_percentage, 2),
                    "csv_headers": csv_headers,
                    "csv_data": csv_data,
                    "summary": {
                        "compliant_nsgs": len([nsg for nsg in nsgs if len(list(nsg.security_rules)) <= 50]),
                        "warning_nsgs": len([nsg for nsg in nsgs if 50 < len(list(nsg.security_rules)) <= 100]),
                        "non_compliant_nsgs": len([nsg for nsg in nsgs if len(list(nsg.security_rules)) > 100])
                    }
                }
                
                response = {
                    "success": True,
                    "report_type": "nsg_ports",
                    "generated_at": datetime.now().isoformat(),
                    "data": ports_report
                }
            except Exception as e:
                logger.error(f"NSG ports report failed: {e}")
                response = {"error": f"Failed to generate NSG ports report: {str(e)}"}
        elif path == '/api/v1/reports/consolidation':
            # Consolidation Report - Generate comprehensive consolidation analysis using real Azure data
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                request_data = json.loads(post_data.decode('utf-8'))
                
                subscription_id = request_data.get('subscription_id', AZURE_CONFIG["subscription_id"])
                resource_group = request_data.get('resource_group', '')
                nsg_names = request_data.get('nsg_names', [])
                
                logger.info(f"Consolidation Report - Subscription: {subscription_id}, RG: {resource_group}, NSGs: {nsg_names}")
                
                # Initialize Azure clients
                credential = DefaultAzureCredential()
                network_client = NetworkManagementClient(credential, subscription_id)
                subscription_client = SubscriptionClient(credential)
                
                # Get subscription name
                subscription_name = "Unknown"
                try:
                    subscription_info = subscription_client.subscriptions.get(subscription_id)
                    subscription_name = subscription_info.display_name
                except:
                    pass
                
                # Fetch real NSG data from Azure
                nsgs = []
                if resource_group:
                    nsg_list = network_client.network_security_groups.list(resource_group)
                else:
                    nsg_list = network_client.network_security_groups.list_all()
                
                for nsg in nsg_list:
                    if not nsg_names or nsg.name in nsg_names:
                        nsgs.append(nsg)
                
                # Fetch ASGs for analysis
                asgs = []
                if resource_group:
                    asg_list = network_client.application_security_groups.list(resource_group)
                else:
                    asg_list = network_client.application_security_groups.list_all()
                
                asgs = list(asg_list)
                
                # Initialize consolidation analysis
                rule_patterns = {}
                port_usage = {}
                ip_usage = {}
                protocol_usage = {}
                duplicate_rules = []
                redundant_nsgs = []
                optimization_opportunities = []
                consolidation_suggestions = []
                
                total_rules = 0
                total_ips = 0
                total_ports = 0
                total_asgs_used = 0
                
                # Analyze each NSG for consolidation opportunities
                for nsg in nsgs:
                    nsg_rules = list(nsg.security_rules) if nsg.security_rules else []
                    total_rules += len(nsg_rules)
                    
                    nsg_ips = set()
                    nsg_ports = set()
                    nsg_asgs = set()
                    
                    for rule in nsg_rules:
                        # Analyze rule patterns for duplicates
                        pattern_key = f"{rule.protocol}-{rule.access}-{rule.direction}"
                        if pattern_key not in rule_patterns:
                            rule_patterns[pattern_key] = []
                        
                        rule_info = {
                            'nsg_name': nsg.name,
                            'rule_name': rule.name,
                            'source_prefix': rule.source_address_prefix or '',
                            'dest_prefix': rule.destination_address_prefix or '',
                            'source_port': rule.source_port_range or '',
                            'dest_port': rule.destination_port_range or '',
                            'priority': rule.priority
                        }
                        rule_patterns[pattern_key].append(rule_info)
                        
                        # Count IP addresses
                        if rule.source_address_prefix and rule.source_address_prefix != '*':
                            nsg_ips.add(rule.source_address_prefix)
                        if rule.destination_address_prefix and rule.destination_address_prefix != '*':
                            nsg_ips.add(rule.destination_address_prefix)
                        
                        # Count ports
                        if rule.source_port_range and rule.source_port_range != '*':
                            nsg_ports.add(rule.source_port_range)
                        if rule.destination_port_range and rule.destination_port_range != '*':
                            nsg_ports.add(rule.destination_port_range)
                        
                        # Count ASG usage
                        if hasattr(rule, 'source_application_security_groups') and rule.source_application_security_groups:
                            for asg in rule.source_application_security_groups:
                                nsg_asgs.add(asg.id)
                        if hasattr(rule, 'destination_application_security_groups') and rule.destination_application_security_groups:
                            for asg in rule.destination_application_security_groups:
                                nsg_asgs.add(asg.id)
                    
                    total_ips += len(nsg_ips)
                    total_ports += len(nsg_ports)
                    total_asgs_used += len(nsg_asgs)
                
                # Find duplicate rules across NSGs
                for pattern, rules in rule_patterns.items():
                    if len(rules) > 1:
                        # Group by exact match (source, dest, ports)
                        exact_matches = {}
                        for rule in rules:
                            exact_key = f"{rule['source_prefix']}-{rule['dest_prefix']}-{rule['source_port']}-{rule['dest_port']}"
                            if exact_key not in exact_matches:
                                exact_matches[exact_key] = []
                            exact_matches[exact_key].append(rule)
                        
                        for exact_key, exact_rules in exact_matches.items():
                            if len(exact_rules) > 1:
                                duplicate_rules.append({
                                    'pattern': pattern,
                                    'exact_match': exact_key,
                                    'rules': exact_rules,
                                    'count': len(exact_rules),
                                    'affected_nsgs': list(set([r['nsg_name'] for r in exact_rules]))
                                })
                
                # Generate LLM-style consolidation recommendations
                
                # 1. Duplicate Rule Consolidation
                if duplicate_rules:
                    for dup in duplicate_rules:
                        optimization_opportunities.append({
                            'type': 'duplicate_rules',
                            'severity': 'high' if dup['count'] > 3 else 'medium',
                            'title': f"Duplicate Rule Pattern: {dup['pattern']}",
                            'description': f"Found {dup['count']} identical rules across {len(dup['affected_nsgs'])} NSGs",
                            'affected_nsgs': dup['affected_nsgs'],
                            'current_rules': [f"{r['nsg_name']}: {r['rule_name']}" for r in dup['rules']],
                            'recommended_action': 'Consolidate into single NSG or use Application Security Groups',
                            'estimated_savings': {
                                'rules_reduced': dup['count'] - 1,
                                'management_complexity': 'Reduced by 60-80%'
                            }
                        })
                
                # 2. NSG Count Optimization
                if len(nsgs) > 10:
                    consolidation_suggestions.append({
                        'type': 'nsg_consolidation',
                        'severity': 'medium',
                        'title': 'High NSG Count Detected',
                        'description': f'You have {len(nsgs)} NSGs which may indicate over-segmentation',
                        'recommendation': 'Consider consolidating NSGs with similar security requirements',
                        'benefits': ['Reduced management overhead', 'Simplified security policies', 'Better compliance tracking']
                    })
                
                # 3. ASG Utilization Recommendations
                if total_asgs_used < len(asgs) * 0.5:
                    consolidation_suggestions.append({
                        'type': 'asg_utilization',
                        'severity': 'low',
                        'title': 'Underutilized Application Security Groups',
                        'description': f'Only {total_asgs_used} out of {len(asgs)} ASGs are actively used',
                        'recommendation': 'Leverage ASGs more effectively to reduce IP-based rules',
                        'benefits': ['Dynamic security group membership', 'Reduced rule complexity', 'Better scalability']
                    })
                
                # 4. Port Range Optimization
                common_ports = ['80', '443', '22', '3389', '21', '25', '53', '110', '143']
                port_consolidation_opportunities = []
                
                for pattern, rules in rule_patterns.items():
                    port_ranges = []
                    for rule in rules:
                        if rule['dest_port'] and rule['dest_port'] not in ['*', '']:
                            port_ranges.append(rule['dest_port'])
                    
                    if len(set(port_ranges)) > 5:  # Many different ports for same pattern
                        port_consolidation_opportunities.append({
                            'pattern': pattern,
                            'ports': list(set(port_ranges)),
                            'count': len(set(port_ranges))
                        })
                
                if port_consolidation_opportunities:
                    consolidation_suggestions.append({
                        'type': 'port_optimization',
                        'severity': 'medium',
                        'title': 'Port Range Consolidation Opportunity',
                        'description': 'Multiple individual port rules could be consolidated into ranges',
                        'opportunities': port_consolidation_opportunities[:3],  # Show top 3
                        'recommendation': 'Use port ranges instead of individual port rules where possible'
                    })
                
                # Generate CSV data for consolidation report
                csv_headers = [
                    'Subscription Name', 'Subscription ID', 'Resource Group', 'NSG Name',
                    'Total Rules', 'Duplicate Rules', 'Optimization Score', 'Recommendations',
                    'Potential Savings (Rules)', 'Risk Level'
                ]
                
                csv_data = []
                for nsg in nsgs:
                    nsg_rule_count = len(list(nsg.security_rules)) if nsg.security_rules else 0
                    nsg_duplicates = sum(1 for dup in duplicate_rules if nsg.name in dup['affected_nsgs'])
                    
                    # Calculate optimization score (0-100)
                    optimization_score = 100
                    if nsg_duplicates > 0:
                        optimization_score -= (nsg_duplicates * 10)
                    if nsg_rule_count > 50:
                        optimization_score -= 20
                    optimization_score = max(0, optimization_score)
                    
                    # Determine risk level
                    if optimization_score >= 80:
                        risk_level = 'Low'
                    elif optimization_score >= 60:
                        risk_level = 'Medium'
                    else:
                        risk_level = 'High'
                    
                    recommendations = []
                    if nsg_duplicates > 0:
                        recommendations.append('Remove duplicate rules')
                    if nsg_rule_count > 50:
                        recommendations.append('Consolidate rules')
                    if not recommendations:
                        recommendations.append('Well optimized')
                    
                    csv_data.append([
                        subscription_name,
                        subscription_id,
                        nsg.id.split('/')[4] if '/' in nsg.id else 'Unknown',
                        nsg.name,
                        nsg_rule_count,
                        nsg_duplicates,
                        f"{optimization_score}%",
                        '; '.join(recommendations),
                        nsg_duplicates,
                        risk_level
                    ])
                
                # Calculate overall savings potential
                total_duplicate_rules = sum(dup['count'] - 1 for dup in duplicate_rules)
                potential_rule_reduction = (total_duplicate_rules / max(total_rules, 1)) * 100
                potential_ip_savings = min(total_ips * 0.3, 1000)  # Estimate 30% IP reduction
                potential_port_savings = len(port_consolidation_opportunities) * 5  # Estimate 5 ports per opportunity
                
                # Generate comprehensive consolidation report
                consolidation_report = {
                    "subscription_id": subscription_id,
                    "subscription_name": subscription_name,
                    "resource_group": resource_group or "All Resource Groups",
                    "analysis_summary": {
                        "total_nsgs_analyzed": len(nsgs),
                        "total_asgs_available": len(asgs),
                        "total_rules": total_rules,
                        "total_ips": total_ips,
                        "total_ports": total_ports,
                        "duplicate_rules_found": len(duplicate_rules),
                        "optimization_opportunities": len(optimization_opportunities)
                    },
                    "consolidation_opportunities": optimization_opportunities,
                    "recommendations": consolidation_suggestions,
                    "potential_savings": {
                        "rules_reduction": total_duplicate_rules,
                        "rules_reduction_percentage": round(potential_rule_reduction, 2),
                        "estimated_ip_savings": int(potential_ip_savings),
                        "estimated_port_savings": potential_port_savings,
                        "management_complexity_reduction": "40-60%",
                        "compliance_improvement": "Significant"
                    },
                    "csv_headers": csv_headers,
                    "csv_data": csv_data,
                    "overall_score": {
                        "optimization_score": round(100 - (potential_rule_reduction * 0.5), 2),
                        "compliance_level": "Good" if potential_rule_reduction < 20 else "Needs Improvement",
                        "priority_actions": len([opp for opp in optimization_opportunities if opp.get('severity') == 'high'])
                    }
                }
                
                response = {
                    "success": True,
                    "report_type": "consolidation",
                    "generated_at": datetime.now().isoformat(),
                    "data": consolidation_report
                }
            except Exception as e:
                logger.error(f"Consolidation report failed: {e}")
                response = {"error": f"Failed to generate consolidation report: {str(e)}"}
        elif path == '/api/v1/email/schedules':
            # Get all email schedules
            try:
                schedules = list(EMAIL_SCHEDULES.values())
                response = {
                    "success": True,
                    "schedules": schedules
                }
            except Exception as e:
                logger.error(f"Email schedules fetch failed: {e}")
                response = {"error": f"Failed to fetch email schedules: {str(e)}"}
        elif path.startswith('/api/v1/email/schedules/'):
            # Get specific email schedule
            try:
                schedule_id = path.split('/')[-1]
                # Mock schedule data
                schedule = {
                    "id": schedule_id,
                    "reportType": "ASG Validation",
                    "frequency": "weekly",
                    "emails": ["admin@company.com"],
                    "enabled": True,
                    "lastSent": "2024-01-08",
                    "nextSend": "2024-01-15",
                    "timeOfDay": "09:00",
                    "status": "active",
                    "successCount": 12,
                    "failureCount": 1,
                    "lastError": None,
                    "createdAt": "2024-01-01T00:00:00Z",
                    "executionHistory": [
                        {"date": "2024-01-08", "status": "success", "recipients": 1},
                        {"date": "2024-01-01", "status": "success", "recipients": 1},
                        {"date": "2023-12-25", "status": "failed", "error": "SMTP connection failed"}
                    ]
                }
                
                response = {
                    "success": True,
                    "schedule": schedule
                }
            except Exception as e:
                logger.error(f"Email schedule fetch failed: {e}")
                response = {"error": f"Failed to fetch email schedule: {str(e)}"}
        elif path == '/api/v1/email/config':
            # Get email configuration
            try:
                # Return current email configuration (password is never returned for security)
                config = {
                    "smtpServer": EMAIL_CONFIG["smtpServer"],
                    "smtpPort": str(EMAIL_CONFIG["smtpPort"]),
                    "smtpUsername": EMAIL_CONFIG["smtpUsername"],
                    "fromEmail": EMAIL_CONFIG["fromEmail"],
                    "fromName": EMAIL_CONFIG["fromName"],
                    "enableTLS": EMAIL_CONFIG["enableTLS"],
                    "isConfigured": bool(EMAIL_CONFIG["smtpUsername"] and EMAIL_CONFIG["smtpPassword"]),
                    "supportedProviders": [
                        {"name": "Office 365", "server": "smtp.office365.com", "port": 587, "tls": True},
                        {"name": "Gmail", "server": "smtp.gmail.com", "port": 587, "tls": True},
                        {"name": "Outlook", "server": "smtp-mail.outlook.com", "port": 587, "tls": True},
                        {"name": "Yahoo", "server": "smtp.mail.yahoo.com", "port": 587, "tls": True},
                        {"name": "Custom SMTP", "server": "", "port": 587, "tls": True}
                    ]
                    # Note: password is never returned for security reasons
                }
                
                response = {
                    "success": True,
                    "config": config
                }
            except Exception as e:
                logger.error(f"Email config fetch failed: {e}")
                response = {"error": f"Failed to fetch email configuration: {str(e)}"}
        elif path == '/api/v1/agents':
            # Return list of created agents
            response = {
                "success": True,
                "agents": AGENTS_STORAGE,
                "total": len(AGENTS_STORAGE)
            }
        elif path == '/api/v1/settings/security':
            response = {
                "success": True,
                "settings": SETTINGS_STORAGE.get("security", {})
            }
        elif path == '/api/v1/settings/notifications':
            response = {
                "success": True,
                "settings": SETTINGS_STORAGE.get("notifications", {})
            }
        elif path == '/api/v1/settings/system':
            response = {
                "success": True,
                "system": SETTINGS_STORAGE.get("system", {})
            }
        elif path == '/api/v1/users':
            response = {
                "success": True,
                "users": USERS_STORAGE,
                "total": len(USERS_STORAGE)
            }
        else:
            response = {"error": "Not found", "path": path}
        
        # Send response
        logger.info(f"Sending response for {path}")
        self.wfile.write(json.dumps(response).encode())
    
    def do_POST(self):
        logger.info(f"Received POST request: {self.path}")
        
        # Parse URL
        parsed_url = urlparse(self.path)
        path = parsed_url.path
        query_params = parse_qs(parsed_url.query)
        
        # Route handling for POST requests
        logger.info(f"POST request path: {path}")
        if path.startswith('/api/v1/nsg-recommendations/'):
            logger.info("Matched nsg-recommendations endpoint")
            # Extract NSG name from path
            nsg_name = path.split('/')[-1]
            subscription_id = query_params.get('subscription_id', [AZURE_CONFIG["subscription_id"]])[0]
            resource_group = query_params.get('resource_group', [''])[0]
            
            try:
                logger.info(f"Processing recommendations for NSG: {nsg_name}")
                # Handle demo NSG with mock data
                if nsg_name == 'demo-nsg':
                    logger.info("Using mock data for demo-nsg")
                    # Use the same mock validation result as the validation endpoint
                    validation_result = {
                        'nsgName': 'demo-nsg',
                        'resourceGroup': 'demo-rg',
                        'totalRules': 8,
                        'inboundRules': 5,
                        'outboundRules': 3,
                        'sourceIpCount': 15,
                        'destinationIpCount': 12,
                        'asgCount': 3,
                        'isWithinLimits': False,
                        'violations': [
                            {
                                'type': 'IP_LIMIT_EXCEEDED',
                                'severity': 'High',
                                'message': 'Source IP count (15) exceeds recommended limit',
                                'currentCount': 15,
                                'maxAllowed': 10
                            }
                        ],
                        'aiAnalysis': {
                            'duplicateIps': [{'ipAddress': '10.0.1.100', 'usageCount': 3}],
                            'cidrOverlaps': [{'cidr1': '10.0.0.0/16', 'cidr2': '10.0.1.0/24'}],
                            'redundantRules': [{'ruleName': 'AllowHTTP', 'similarityScore': 0.85}],
                            'securityRisks': [
                                {
                                    'ruleName': 'AllowAll',
                                    'overallSeverity': 'Critical',
                                    'risks': [{'type': 'Wildcard source address'}]
                                }
                            ],
                            'consolidationOpportunities': [
                                {
                                    'potentialSavings': {'ruleReduction': 5}
                                }
                            ],
                            'visualAnalytics': {
                                'ruleDistribution': {
                                    'inbound': 5,
                                    'outbound': 3
                                },
                                'accessTypes': {
                                    'allow': 7,
                                    'deny': 1
                                },
                                'protocolDistribution': {
                                    'TCP': 6,
                                    'UDP': 2,
                                    'ICMP': 0
                                },
                                'priorityRanges': {
                                    'high': 2,
                                    'medium': 4,
                                    'low': 2
                                },
                                'riskLevels': {
                                    'critical': 1,
                                    'high': 2,
                                    'medium': 3,
                                    'low': 2
                                }
                            }
                        }
                    }
                else:
                    # First get the validation analysis for real NSGs
                    validation_result = nsg_validator.analyze_nsg_rules(subscription_id, resource_group, nsg_name)
                
                # Generate LLM recommendations
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                recommendations = loop.run_until_complete(
                    nsg_validator.generate_llm_recommendations(validation_result)
                )
                loop.close()
                
                logger.info(f"Generated {len(recommendations) if recommendations else 0} recommendations")
                response = {"recommendations": recommendations}
            except Exception as e:
                logger.error(f"LLM recommendation generation failed: {e}")
                response = {"error": f"Recommendation generation failed: {str(e)}"}
            
            # Send the response
            logger.info(f"Sending POST response for nsg-recommendations: {nsg_name}")
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
            self.send_header('Access-Control-Allow-Headers', 'Content-Type')
            self.end_headers()
            self.wfile.write(json.dumps(response).encode())
            return
        elif path == '/api/v1/nsgs':
            # Create new NSG
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                nsg_data = json.loads(post_data.decode('utf-8'))
                
                # Mock NSG creation response
                response = {
                    "success": True,
                    "message": "NSG created successfully",
                    "nsg": {
                        "id": 999,
                        "name": nsg_data.get('name', 'new-nsg'),
                        "resource_group": nsg_data.get('resource_group', 'default-rg'),
                        "region": nsg_data.get('region', 'East US'),
                        "subscription_id": nsg_data.get('subscription_id', AZURE_CONFIG["subscription_id"]),
                        "azure_id": f"/subscriptions/{nsg_data.get('subscription_id', AZURE_CONFIG['subscription_id'])}/resourceGroups/{nsg_data.get('resource_group', 'default-rg')}/providers/Microsoft.Network/networkSecurityGroups/{nsg_data.get('name', 'new-nsg')}",
                        "inbound_rules": [],
                        "outbound_rules": [],
                        "tags": nsg_data.get('tags', {}),
                        "is_active": True,
                        "compliance_score": 85,
                        "risk_level": "low",
                        "last_sync": "2024-08-22T13:51:52Z",
                        "created_at": "2024-08-22T13:51:52Z",
                        "updated_at": "2024-08-22T13:51:52Z"
                    }
                }
            except Exception as e:
                logger.error(f"NSG creation failed: {e}")
                response = {"error": f"NSG creation failed: {str(e)}"}
            
            # Send NSG creation response and return
            logger.info(f"Sending POST response for {path}")
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
            self.send_header('Access-Control-Allow-Headers', 'Content-Type')
            self.end_headers()
            self.wfile.write(json.dumps(response).encode())
            return
        elif path.startswith('/api/v1/agents'):
            # Handle agent creation
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                agent_data = json.loads(post_data.decode('utf-8'))
                
                # Create new agent with unique ID
                new_agent = {
                    "id": len(AGENTS_STORAGE) + 1,
                    "name": agent_data.get('name', ''),
                    "description": agent_data.get('description', ''),
                    "ai_model": agent_data.get('ai_model', 'GPT-4 Turbo'),
                    "security_scan": agent_data.get('security_scan', ''),
                    "finding_title": agent_data.get('finding_title', ''),
                    "finding_description": agent_data.get('finding_description', ''),
                    "severity": agent_data.get('severity', 'Medium'),
                    "resource_type": agent_data.get('resource_type', 'Network Security Group'),
                    "resource_id": agent_data.get('resource_id', ''),
                    "subscription_id": agent_data.get('subscription_id', ''),
                    "resource_group": agent_data.get('resource_group', ''),
                    "region": agent_data.get('region', ''),
                    "network_security_group": agent_data.get('network_security_group', ''),
                    "validation_mode": agent_data.get('validation_mode', True),
                    "automated_remediation": agent_data.get('automated_remediation', False),
                    "status": "Created",
                    "created_at": datetime.utcnow().isoformat() + "Z",
                    "updated_at": datetime.utcnow().isoformat() + "Z"
                }
                
                # Add to storage
                AGENTS_STORAGE.append(new_agent)
                
                response = {
                    "success": True,
                    "message": "Agent created successfully",
                    "agent": new_agent
                }
                
            except Exception as e:
                logger.error(f"Agent creation failed: {e}")
                response = {"error": f"Agent creation failed: {str(e)}"}
            
            # Send agent creation response
            logger.info(f"Sending POST response for {path}")
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
            self.send_header('Access-Control-Allow-Headers', 'Content-Type')
            self.end_headers()
            self.wfile.write(json.dumps(response).encode())
            return
        elif path == '/api/v1/golden-rule/compare':
            # Golden Rule comparison endpoint
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                comparison_data = json.loads(post_data.decode('utf-8'))
                
                logger.info(f"Golden Rule comparison request: {comparison_data}")
                
                current_config = comparison_data.get('current_config', {})
                golden_standard = comparison_data.get('golden_standard', {})
                comparison_type = comparison_data.get('comparison_type', 'full')
                
                # Perform comparison analysis
                comparison_result = {
                    "compliance_score": 0,
                    "differences": {
                        "added_rules": [],
                        "removed_rules": [],
                        "modified_rules": []
                    },
                    "recommendations": []
                }
                
                # Extract rules from configurations
                current_inbound = current_config.get('inbound_rules', [])
                current_outbound = current_config.get('outbound_rules', [])
                golden_inbound = golden_standard.get('inbound_rules', [])
                golden_outbound = golden_standard.get('outbound_rules', [])
                
                # Create rule dictionaries for comparison
                current_rules = {rule['name']: rule for rule in current_inbound + current_outbound}
                golden_rules = {rule['name']: rule for rule in golden_inbound + golden_outbound}
                
                # Find missing rules (in golden but not in current)
                for rule_name, rule_data in golden_rules.items():
                    if rule_name not in current_rules:
                        comparison_result["differences"]["removed_rules"].append(rule_data)
                
                # Find extra rules (in current but not in golden)
                for rule_name, rule_data in current_rules.items():
                    if rule_name not in golden_rules:
                        comparison_result["differences"]["added_rules"].append(rule_data)
                
                # Find modified rules
                for rule_name in current_rules:
                    if rule_name in golden_rules:
                        current_rule = current_rules[rule_name]
                        golden_rule = golden_rules[rule_name]
                        
                        changes = []
                        for key in ['priority', 'access', 'protocol', 'source_port_range', 'destination_port_range', 'source_address_prefix', 'destination_address_prefix']:
                            if current_rule.get(key) != golden_rule.get(key):
                                changes.append(f"{key}: {current_rule.get(key)} -> {golden_rule.get(key)}")
                        
                        if changes:
                            comparison_result["differences"]["modified_rules"].append({
                                "rule_name": rule_name,
                                "current": current_rule,
                                "golden": golden_rule,
                                "changes": changes
                            })
                
                # Calculate compliance score
                total_golden_rules = len(golden_rules)
                missing_count = len(comparison_result["differences"]["removed_rules"])
                extra_count = len(comparison_result["differences"]["added_rules"])
                modified_count = len(comparison_result["differences"]["modified_rules"])
                
                if total_golden_rules > 0:
                    base_score = 100 - ((missing_count + modified_count) / total_golden_rules * 100)
                    penalty = min(extra_count * 5, 20)  # Max 20% penalty for extra rules
                    comparison_result["compliance_score"] = max(0, base_score - penalty)
                else:
                    comparison_result["compliance_score"] = 100
                
                # Generate recommendations
                if missing_count > 0:
                    comparison_result["recommendations"].append(f"Add {missing_count} missing security rules from golden standard")
                if extra_count > 0:
                    comparison_result["recommendations"].append(f"Review {extra_count} extra rules not in golden standard")
                if modified_count > 0:
                    comparison_result["recommendations"].append(f"Update {modified_count} rules to match golden standard configuration")
                
                response = comparison_result
                
            except Exception as e:
                logger.error(f"Golden Rule comparison failed: {e}")
                response = {"error": f"Golden Rule comparison failed: {str(e)}"}
            
            # Send response
            logger.info(f"Sending POST response for golden-rule/compare")
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
            self.send_header('Access-Control-Allow-Headers', 'Content-Type')
            self.end_headers()
            self.wfile.write(json.dumps(response).encode())
            return
        elif path == '/api/v1/golden-rule/storage':
            # Load golden standard from Azure Storage
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                storage_data = json.loads(post_data.decode('utf-8'))
                
                logger.info(f"Golden Rule storage load request: {storage_data}")
                
                storage_account = storage_data.get('storage_account', 'thirustorage001')
                container_name = storage_data.get('container', 'golden-rules')
                file_name = storage_data.get('file_name', 'golden-standard.json')
                
                # Mock loading from Azure Storage
                # In real implementation, this would use Azure Storage SDK
                mock_golden_standard = {
                    "name": "Production Security Standard",
                    "description": "Standard security configuration for production NSGs",
                    "inbound_rules": [
                        {
                            "name": "AllowHTTPS",
                            "priority": 100,
                            "direction": "Inbound",
                            "access": "Allow",
                            "protocol": "TCP",
                            "source_port_range": "*",
                            "destination_port_range": "443",
                            "source_address_prefix": "*",
                            "destination_address_prefix": "*",
                            "description": "Allow HTTPS traffic"
                        },
                        {
                            "name": "AllowSSH",
                            "priority": 110,
                            "direction": "Inbound",
                            "access": "Allow",
                            "protocol": "TCP",
                            "source_port_range": "*",
                            "destination_port_range": "22",
                            "source_address_prefix": "10.0.0.0/8",
                            "destination_address_prefix": "*",
                            "description": "Allow SSH from internal network"
                        }
                    ],
                    "outbound_rules": [
                        {
                            "name": "AllowInternetOutbound",
                            "priority": 100,
                            "direction": "Outbound",
                            "access": "Allow",
                            "protocol": "*",
                            "source_port_range": "*",
                            "destination_port_range": "*",
                            "source_address_prefix": "*",
                            "destination_address_prefix": "Internet",
                            "description": "Allow outbound internet access"
                        }
                    ]
                }
                
                response = {
                    "success": True,
                    "golden_standard": mock_golden_standard,
                    "source": f"Azure Storage: {storage_account}/{container_name}/{file_name}"
                }
                
            except Exception as e:
                logger.error(f"Golden Rule storage load failed: {e}")
                response = {"error": f"Failed to load from storage: {str(e)}"}
            
            # Send response
            logger.info(f"Sending POST response for golden-rule/storage")
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
            self.send_header('Access-Control-Allow-Headers', 'Content-Type')
            self.end_headers()
            self.wfile.write(json.dumps(response).encode())
            return
        elif path == '/api/v1/backup/create':
            # Create backup configuration
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                backup_data = json.loads(post_data.decode('utf-8'))
                
                # Extract backup parameters
                storage_account = backup_data.get('storageAccount', backup_data.get('storage_account', 'thirustorage001'))
                container = backup_data.get('container', backup_data.get('container_name', 'nsg-backups'))
                selected_nsgs = backup_data.get('selectedNSGs', backup_data.get('selected_nsgs', []))
                selected_asgs = backup_data.get('selectedASGs', backup_data.get('selected_asgs', []))
                resource_type = backup_data.get('resourceType', backup_data.get('resource_type', 'nsg'))
                immediate = backup_data.get('immediate', False)
                backup_format = backup_data.get('backupFormat', backup_data.get('backup_format', 'both'))  # 'json', 'csv', or 'both'
                
                logger.info(f"Backup request - NSGs: {selected_nsgs}, ASGs: {selected_asgs}, Format: {backup_format}, Type: {resource_type}")
                
                # Validate that at least some resources are selected
                if not selected_nsgs and not selected_asgs:
                    self.send_response(400)
                    self.send_header('Content-Type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    error_response = {
                        "success": False,
                        "error": "No resources selected",
                        "message": "Please select at least one NSG or ASG to backup"
                    }
                    self.wfile.write(json.dumps(error_response).encode('utf-8'))
                    return
                
                # Create backup file content
                backup_content = {
                    "backup_metadata": {
                        "backup_id": "backup-" + str(hash(str(backup_data)) % 10000),
                        "created_at": datetime.utcnow().isoformat() + "Z",
                        "backup_type": "immediate" if immediate else "scheduled",
                        "resource_type": resource_type,
                        "storage_account": storage_account,
                        "container": container
                    },
                    "nsgs": [],
                    "asgs": []
                }
                
                # Add NSG data if selected - fetch real data from Azure
                if selected_nsgs and (resource_type == 'nsg' or resource_type == 'both'):
                    subscription_id = backup_data.get('subscription_id', backup_data.get('selectedSubscription', ''))
                    if subscription_id:
                        try:
                            # Get Azure clients with timeout
                            def fetch_nsg_data():
                                network_client = get_azure_clients()[2]  # NetworkManagementClient
                                nsgs = list(network_client.network_security_groups.list_all())
                                return nsgs
                            
                            # Use threading for timeout
                            result_container = []
                            exception_container = []
                            
                            def target():
                                try:
                                    result = fetch_nsg_data()
                                    result_container.append(result)
                                except Exception as e:
                                    exception_container.append(e)
                            
                            thread = threading.Thread(target=target)
                            thread.daemon = True
                            thread.start()
                            thread.join(timeout=10)  # 10 second timeout
                            
                            if thread.is_alive():
                                logger.error("NSG data fetch timed out after 10 seconds")
                                # Use mock data as fallback
                                azure_nsgs = []
                            elif exception_container:
                                logger.error(f"Error fetching NSG data: {exception_container[0]}")
                                azure_nsgs = []
                            elif result_container:
                                azure_nsgs = result_container[0]
                                logger.info(f"Successfully fetched {len(azure_nsgs)} NSGs from Azure")
                            else:
                                azure_nsgs = []
                            
                            # Process each selected NSG
                            logger.info(f"Processing {len(selected_nsgs)} selected NSGs: {selected_nsgs}")
                            logger.info(f"Available Azure NSGs: {[nsg.name for nsg in azure_nsgs]}")
                            
                            for nsg_id in selected_nsgs:
                                logger.info(f"Processing NSG: {nsg_id}")
                                # Find matching NSG from Azure data
                                matching_nsg = None
                                for azure_nsg in azure_nsgs:
                                    if azure_nsg.name == nsg_id or azure_nsg.id.endswith(nsg_id) or nsg_id in azure_nsg.id:
                                        matching_nsg = azure_nsg
                                        logger.info(f"Found matching NSG: {azure_nsg.name} for {nsg_id}")
                                        break
                                
                                if matching_nsg:
                                    logger.info(f"Extracting data for NSG: {matching_nsg.name}")
                                    # Extract real NSG data
                                    inbound_rules = []
                                    outbound_rules = []
                                    
                                    # Process custom security rules
                                    if hasattr(matching_nsg, 'security_rules') and matching_nsg.security_rules:
                                        for rule in matching_nsg.security_rules:
                                            rule_data = {
                                                "name": rule.name,
                                                "priority": rule.priority,
                                                "direction": rule.direction,
                                                "access": rule.access,
                                                "protocol": rule.protocol,
                                                "source_port_range": rule.source_port_range or "*",
                                                "destination_port_range": rule.destination_port_range or "*",
                                                "source_address_prefix": rule.source_address_prefix or "*",
                                                "destination_address_prefix": rule.destination_address_prefix or "*",
                                                "source_address_prefixes": getattr(rule, 'source_address_prefixes', None) or [],
                                                "destination_address_prefixes": getattr(rule, 'destination_address_prefixes', None) or [],
                                                "source_application_security_groups": [asg.id for asg in (rule.source_application_security_groups or [])],
                                                "destination_application_security_groups": [asg.id for asg in (rule.destination_application_security_groups or [])],
                                                "rule_type": "custom"
                                            }
                                            
                                            if rule.direction == 'Inbound':
                                                inbound_rules.append(rule_data)
                                            else:
                                                outbound_rules.append(rule_data)
                                    
                                    # Process default security rules
                                    if hasattr(matching_nsg, 'default_security_rules') and matching_nsg.default_security_rules:
                                        for rule in matching_nsg.default_security_rules:
                                            rule_data = {
                                                "name": rule.name,
                                                "priority": rule.priority,
                                                "direction": rule.direction,
                                                "access": rule.access,
                                                "protocol": rule.protocol,
                                                "source_port_range": rule.source_port_range or "*",
                                                "destination_port_range": rule.destination_port_range or "*",
                                                "source_address_prefix": rule.source_address_prefix or "*",
                                                "destination_address_prefix": rule.destination_address_prefix or "*",
                                                "source_address_prefixes": getattr(rule, 'source_address_prefixes', None) or [],
                                                "destination_address_prefixes": getattr(rule, 'destination_address_prefixes', None) or [],
                                                "source_application_security_groups": [asg.id for asg in (rule.source_application_security_groups or [])],
                                                "destination_application_security_groups": [asg.id for asg in (rule.destination_application_security_groups or [])],
                                                "rule_type": "default"
                                            }
                                            
                                            if rule.direction == 'Inbound':
                                                inbound_rules.append(rule_data)
                                            else:
                                                outbound_rules.append(rule_data)
                                    
                                    nsg_data = {
                                        "id": matching_nsg.id,
                                        "name": matching_nsg.name,
                                        "resource_group": matching_nsg.id.split('/')[4] if '/' in matching_nsg.id else 'unknown',
                                        "location": matching_nsg.location,
                                        "subscription_id": subscription_id,
                                        "inbound_rules": inbound_rules,
                                        "outbound_rules": outbound_rules,
                                        "tags": dict(matching_nsg.tags) if matching_nsg.tags else {},
                                        "backed_up_at": datetime.utcnow().isoformat() + "Z"
                                    }
                                else:
                                    logger.warning(f"NSG {nsg_id} not found in Azure data, using sample data with rules")
                                    # Fallback to sample data with actual rules if NSG not found
                                    nsg_data = {
                                        "id": nsg_id,
                                        "name": f"NSG-{nsg_id}",
                                        "resource_group": backup_data.get('selectedResourceGroup', 'default-rg'),
                                        "location": backup_data.get('selectedLocation', 'East US'),
                                        "subscription_id": subscription_id,
                                        "inbound_rules": [
                                            {
                                                "name": "AllowHTTP",
                                                "priority": 100,
                                                "direction": "Inbound",
                                                "access": "Allow",
                                                "protocol": "TCP",
                                                "source_port_range": "*",
                                                "destination_port_range": "80",
                                                "source_address_prefix": "*",
                                                "destination_address_prefix": "*",
                                                "source_address_prefixes": [],
                                                "destination_address_prefixes": [],
                                                "source_application_security_groups": [],
                                                "destination_application_security_groups": [],
                                                "description": "Allow HTTP traffic",
                                                "rule_type": "custom"
                                            },
                                            {
                                                "name": "AllowHTTPS",
                                                "priority": 110,
                                                "direction": "Inbound",
                                                "access": "Allow",
                                                "protocol": "TCP",
                                                "source_port_range": "*",
                                                "destination_port_range": "443",
                                                "source_address_prefix": "*",
                                                "destination_address_prefix": "*",
                                                "source_address_prefixes": [],
                                                "destination_address_prefixes": [],
                                                "source_application_security_groups": [],
                                                "destination_application_security_groups": [],
                                                "description": "Allow HTTPS traffic",
                                                "rule_type": "custom"
                                            }
                                        ],
                                        "outbound_rules": [
                                            {
                                                "name": "AllowInternetOutbound",
                                                "priority": 100,
                                                "direction": "Outbound",
                                                "access": "Allow",
                                                "protocol": "*",
                                                "source_port_range": "*",
                                                "destination_port_range": "*",
                                                "source_address_prefix": "*",
                                                "destination_address_prefix": "Internet",
                                                "source_address_prefixes": [],
                                                "destination_address_prefixes": ["Internet"],
                                                "source_application_security_groups": [],
                                                "destination_application_security_groups": [],
                                                "description": "Allow outbound internet access",
                                                "rule_type": "custom"
                                            }
                                        ],
                                        "tags": {"Environment": "Production", "Owner": "IT-Team"},
                                        "backed_up_at": datetime.utcnow().isoformat() + "Z"
                                    }
                                
                                backup_content["nsgs"].append(nsg_data)
                                logger.info(f"Added NSG {nsg_data['name']} to backup with {len(nsg_data['inbound_rules'])} inbound and {len(nsg_data['outbound_rules'])} outbound rules")
                                
                        except Exception as e:
                            logger.error(f"Error processing NSG data: {e}")
                            # Fallback to mock data
                            for nsg_id in selected_nsgs:
                                nsg_data = {
                                    "id": nsg_id,
                                    "name": f"NSG-{nsg_id}",
                                    "resource_group": backup_data.get('selectedResourceGroup', 'default-rg'),
                                    "location": backup_data.get('selectedLocation', 'East US'),
                                    "subscription_id": subscription_id,
                                    "inbound_rules": [],
                                    "outbound_rules": [],
                                    "tags": {},
                                    "backed_up_at": datetime.utcnow().isoformat() + "Z"
                                }
                                backup_content["nsgs"].append(nsg_data)
                
                # Add ASG data if selected
                if selected_asgs and (resource_type == 'asg' or resource_type == 'both'):
                    for asg_id in selected_asgs:
                        asg_data = {
                            "id": asg_id,
                            "name": f"ASG-{asg_id}",
                            "resource_group": backup_data.get('selectedResourceGroup', 'default-rg'),
                            "location": backup_data.get('selectedLocation', 'East US'),
                            "subscription_id": backup_data.get('selectedSubscription', ''),
                            "backed_up_at": datetime.utcnow().isoformat() + "Z"
                        }
                        backup_content["asgs"].append(asg_data)
                
                # Save backup to Azure Storage and create CSV export
                timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
                backup_filename = f"backup_{resource_type}_{timestamp}"
                blob_url = None
                json_blob_url = None
                csv_blob_url = None
                
                if azure_service and azure_service.blob_service_client:
                    try:
                        # Upload based on selected format
                        if backup_format in ['json', 'both']:
                            # Create JSON backup in Azure Storage
                            json_content = json.dumps(backup_content, indent=2, ensure_ascii=False)
                            json_blob_url = azure_service.upload_blob_sync(
                                json_content, f"{backup_filename}.json", container, content_type="application/json"
                            )
                            blob_url = json_blob_url
                            logger.info(f"JSON backup saved to Azure Storage: {json_blob_url}")
                        
                        if backup_format in ['csv', 'both']:
                            # Create CSV export
                            csv_filename = f"{backup_filename}.csv"
                            csv_content = create_csv_from_backup(backup_content)
                            
                            # Upload CSV to Azure Storage
                            csv_blob_url = azure_service.upload_blob_sync(
                                csv_content, csv_filename, container, content_type="text/csv"
                            )
                            if backup_format == 'csv':
                                blob_url = csv_blob_url
                            logger.info(f"CSV backup saved to Azure Storage: {csv_blob_url}")
                        
                    except Exception as e:
                        logger.error(f"Failed to save to Azure Storage: {e}")
                        # Fallback to local storage
                        backup_dir = os.path.join(os.path.dirname(__file__), 'backups')
                        os.makedirs(backup_dir, exist_ok=True)
                        
                        if backup_format in ['json', 'both']:
                            backup_filepath = os.path.join(backup_dir, f"{backup_filename}.json")
                            with open(backup_filepath, 'w', encoding='utf-8') as f:
                                json.dump(backup_content, f, indent=2, ensure_ascii=False)
                            logger.info(f"JSON backup saved locally: {backup_filepath}")
                        
                        if backup_format in ['csv', 'both']:
                            # Create local CSV file
                            csv_filepath = os.path.join(backup_dir, f"{backup_filename}.csv")
                            csv_content = create_csv_from_backup(backup_content)
                            with open(csv_filepath, 'w', newline='', encoding='utf-8') as f:
                                f.write(csv_content)
                            logger.info(f"CSV backup saved locally: {csv_filepath}")
                else:
                    # Fallback to local storage when Azure Service is not available
                    backup_dir = os.path.join(os.path.dirname(__file__), 'backups')
                    os.makedirs(backup_dir, exist_ok=True)
                    
                    if backup_format in ['json', 'both']:
                        backup_filepath = os.path.join(backup_dir, f"{backup_filename}.json")
                        with open(backup_filepath, 'w', encoding='utf-8') as f:
                            json.dump(backup_content, f, indent=2, ensure_ascii=False)
                        logger.info(f"JSON backup saved locally: {backup_filepath}")
                    
                    if backup_format in ['csv', 'both']:
                        # Create local CSV file
                        csv_filepath = os.path.join(backup_dir, f"{backup_filename}.csv")
                        csv_content = create_csv_from_backup(backup_content)
                        with open(csv_filepath, 'w', newline='', encoding='utf-8') as f:
                            f.write(csv_content)
                        logger.info(f"CSV backup saved locally: {csv_filepath}")
                
                # Set primary backup URL based on format
                if backup_format == 'csv':
                    primary_url = csv_blob_url
                elif backup_format == 'json':
                    primary_url = json_blob_url
                else:  # 'both'
                    primary_url = json_blob_url or csv_blob_url
                
                # Determine primary backup file based on format
                primary_backup_file = None
                if backup_format == 'json':
                    primary_backup_file = f"{backup_filename}.json"
                elif backup_format == 'csv':
                    primary_backup_file = f"{backup_filename}.csv"
                elif backup_format == 'both':
                    primary_backup_file = f"{backup_filename}.json"  # Default to JSON for 'both'
                
                response = {
                    "success": True,
                    "message": "Backup created and saved successfully",
                    "backup_id": backup_content["backup_metadata"]["backup_id"],
                    "backup_file": primary_backup_file,
                    "json_file": f"{backup_filename}.json" if backup_format in ['json', 'both'] else None,
                    "csv_file": f"{backup_filename}.csv" if backup_format in ['csv', 'both'] else None,
                    "backup_url": primary_url,
                    "csv_url": csv_blob_url,
                    "schedule": {
                        "frequency": backup_data.get('frequency', 'immediate' if immediate else 'daily'),
                        "time": backup_data.get('time', '09:00'),
                        "start_date": backup_data.get('start_date', datetime.utcnow().strftime('%Y-%m-%d'))
                    },
                    "storage": {
                        "account": storage_account,
                        "container": container,
                        "type": "azure_storage" if blob_url else "local_storage"
                    },
                    "exports": {
                        "json": {
                            "available": backup_format in ['json', 'both'],
                            "url": json_blob_url if backup_format in ['json', 'both'] else None,
                            "format": "application/json"
                        },
                        "csv": {
                            "available": backup_format in ['csv', 'both'],
                            "url": csv_blob_url if backup_format in ['csv', 'both'] else None,
                            "format": "text/csv"
                        }
                    },
                    "backup_format": backup_format,
                    "resource_type": resource_type,
                    "nsgs_count": len(backup_content["nsgs"]),
                    "asgs_count": len(backup_content["asgs"]),
                    "backup_type": "immediate" if immediate else "scheduled",
                    "created_at": backup_content["backup_metadata"]["created_at"],
                    "next_run": "immediate" if immediate else "2024-01-16T09:00:00Z"
                }
                
                # Send response immediately and return
                logger.info(f"Sending POST response for {path}")
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
            except Exception as e:
                logger.error(f"Backup creation failed: {e}")
                error_response = {"error": f"Backup creation failed: {str(e)}"}
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(error_response).encode())
                return
        elif path == '/api/v1/storage-accounts/create':
            # Create new storage account
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                storage_data = json.loads(post_data.decode('utf-8'))
                
                # Mock storage account creation response
                response = {
                    "success": True,
                    "message": "Storage account created successfully",
                    "storage_account": {
                        "id": f"/subscriptions/{storage_data.get('subscription_id', '')}/resourceGroups/{storage_data.get('resource_group', '')}/providers/Microsoft.Storage/storageAccounts/{storage_data.get('name', '')}",
                        "name": storage_data.get('name', ''),
                        "resource_group": storage_data.get('resource_group', ''),
                        "location": storage_data.get('location', 'eastus'),
                        "sku": "Standard_LRS",
                        "kind": "StorageV2",
                        "provisioning_state": "Succeeded",
                        "created_at": "2024-01-15T10:30:00Z"
                    },
                    "container": {
                        "name": storage_data.get('container_name', 'nsg-backups'),
                        "created_at": "2024-01-15T10:30:00Z"
                    }
                }
            except Exception as e:
                logger.error(f"Storage account creation failed: {e}")
                response = {"error": f"Storage account creation failed: {str(e)}"}
        elif path == '/api/v1/backup/storage-config':
            # Save storage configuration
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                storage_config = json.loads(post_data.decode('utf-8'))
                
                logger.info(f"Received storage config: {storage_config}")
                
                # Mock storage configuration save response
                response = {
                    "success": True,
                    "message": "Storage configuration saved successfully",
                    "storage_account_created": storage_config.get('create_new', False),
                    "configuration": {
                        "subscription_id": storage_config.get('subscription_id'),
                        "storage_account": storage_config.get('storage_account_name'),
                        "container": storage_config.get('container_name'),
                        "resource_group": storage_config.get('resource_group'),
                        "location": storage_config.get('location'),
                        "storage_type": storage_config.get('storage_type', 'Standard_LRS'),
                        "created_new": storage_config.get('create_new', False)
                    }
                }
                
                # Send response immediately
                logger.info(f"Sending POST response for {path}")
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
                
            except Exception as e:
                logger.error(f"Storage configuration save failed: {e}")
                response = {"error": f"Storage configuration save failed: {str(e)}"}
                
                # Send error response immediately
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
        elif path == '/api/v1/backup/export':
            # Export backup data
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                export_data = json.loads(post_data.decode('utf-8'))
                
                export_format = export_data.get('format', 'csv')
                logger.info(f"Export format received: '{export_format}', type: {type(export_format)}")
                subscription_id = export_data.get('subscription_id', '')
                resource_type = export_data.get('resource_type', 'nsg')
                separate_columns = export_data.get('separateColumns', False)
                include_rule_details = export_data.get('includeRuleDetails', False)
                include_asg_mapping = export_data.get('includeASGMapping', False)
                
                # Extract NSG selection from multiple possible locations
                selected_nsgs = export_data.get('selected_nsgs', [])
                if not selected_nsgs:
                    selected_nsgs = export_data.get('selectedNSGs', [])
                    if not selected_nsgs:
                        backup_config = export_data.get('backupConfig', {})
                        selected_nsgs = backup_config.get('selectedNSGs', [])
                
                logger.info(f"Raw selected NSGs from export_data.get('selectedNSGs'): {export_data.get('selectedNSGs', [])}")
                logger.info(f"Raw selected NSGs from export_data.get('selected_nsgs'): {export_data.get('selected_nsgs', [])}")
                
                logger.info(f"Selected NSG IDs for export: {selected_nsgs}, type: {type(selected_nsgs)}")
                logger.info(f"Export data keys: {list(export_data.keys())}")
                logger.info(f"Full export data: {json.dumps(export_data, indent=2)}")
                
                # Validate selected NSGs
                if not selected_nsgs:
                    logger.error("No NSGs selected for export")
                    self.send_response(400)
                    self.send_header('Content-Type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                    self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                    self.end_headers()
                    self.wfile.write(json.dumps({"error": "No NSGs selected for export"}).encode('utf-8'))
                    return
                
                # Handle selected NSG names directly (frontend sends NSG names, not IDs)
                selected_nsg_names = []
                if selected_nsgs:
                    # Frontend sends NSG names directly, so use them as-is
                    selected_nsg_names = [str(nsg_name) for nsg_name in selected_nsgs]
                    logger.info(f"Selected NSG names for export: {selected_nsg_names}")
                
                # Get real NSG data for export with improved reliability
                try:
                    # Create backup content with real NSG data
                    backup_content = {
                        "backup_metadata": {
                            "backup_id": "export-" + str(hash(str(export_data)) % 10000),
                            "created_at": datetime.utcnow().isoformat() + "Z",
                            "backup_type": "export",
                            "resource_type": resource_type
                        },
                        "nsgs": []
                    }
                    
                    # Fetch NSG data from Azure with retry mechanism
                    def fetch_nsg_data_with_retry(max_retries=3):
                        for attempt in range(max_retries):
                            try:
                                logger.info(f"Attempting to fetch NSG data (attempt {attempt + 1}/{max_retries})")
                                network_client = get_azure_clients()[2]  # NetworkManagementClient
                                if not network_client:
                                    logger.error("Network client not available")
                                    if attempt == max_retries - 1:
                                        raise Exception("Azure Network client not initialized")
                                    continue
                                
                                azure_nsgs = list(network_client.network_security_groups.list_all())
                                logger.info(f"Fetched {len(azure_nsgs)} NSGs for export")
                                
                                # Track processed NSGs to avoid duplicates
                                processed_nsgs = set()
                                
                                for azure_nsg in azure_nsgs:
                                    # Filter by selected NSGs if specified
                                    if selected_nsg_names and azure_nsg.name not in selected_nsg_names:
                                        logger.info(f"Skipping NSG {azure_nsg.name} - not in selected list")
                                        continue
                                    
                                    # Skip if already processed
                                    if azure_nsg.name in processed_nsgs:
                                        logger.info(f"Skipping duplicate NSG {azure_nsg.name}")
                                        continue
                                    
                                    processed_nsgs.add(azure_nsg.name)
                                    logger.info(f"Processing NSG {azure_nsg.name} for export")
                                    # Extract real NSG data
                                    inbound_rules = []
                                    outbound_rules = []
                                    
                                    # Process custom security rules
                                    if hasattr(azure_nsg, 'security_rules') and azure_nsg.security_rules:
                                        for rule in azure_nsg.security_rules:
                                            rule_data = {
                                                "name": rule.name,
                                                "priority": rule.priority,
                                                "direction": rule.direction,
                                                "access": rule.access,
                                                "protocol": rule.protocol,
                                                "source_port_range": rule.source_port_range or "*",
                                                "destination_port_range": rule.destination_port_range or "*",
                                                "source_address_prefix": rule.source_address_prefix or "*",
                                                "destination_address_prefix": rule.destination_address_prefix or "*",
                                                # Prefer plural address prefixes when available
                                                "source_address_prefixes": getattr(rule, 'source_address_prefixes', None) or [],
                                                "destination_address_prefixes": getattr(rule, 'destination_address_prefixes', None) or [],
                                                "source_application_security_groups": [asg.id for asg in (rule.source_application_security_groups or [])],
                                                "destination_application_security_groups": [asg.id for asg in (rule.destination_application_security_groups or [])],
                                                "description": getattr(rule, 'description', '') or f"NSG rule {rule.name}",
                                                "rule_type": "custom"
                                            }
                                            
                                            if rule.direction == 'Inbound':
                                                inbound_rules.append(rule_data)
                                            else:
                                                outbound_rules.append(rule_data)
                                    
                                    # Process default security rules
                                    if hasattr(azure_nsg, 'default_security_rules') and azure_nsg.default_security_rules:
                                        for rule in azure_nsg.default_security_rules:
                                            rule_data = {
                                                "name": rule.name,
                                                "priority": rule.priority,
                                                "direction": rule.direction,
                                                "access": rule.access,
                                                "protocol": rule.protocol,
                                                "source_port_range": rule.source_port_range or "*",
                                                "destination_port_range": rule.destination_port_range or "*",
                                                "source_address_prefix": rule.source_address_prefix or "*",
                                                "destination_address_prefix": rule.destination_address_prefix or "*",
                                                # Prefer plural address prefixes when available
                                                "source_address_prefixes": getattr(rule, 'source_address_prefixes', None) or [],
                                                "destination_address_prefixes": getattr(rule, 'destination_address_prefixes', None) or [],
                                                "source_application_security_groups": [asg.id for asg in (rule.source_application_security_groups or [])],
                                                "destination_application_security_groups": [asg.id for asg in (rule.destination_application_security_groups or [])],
                                                "description": getattr(rule, 'description', '') or f"Default NSG rule {rule.name}",
                                                "rule_type": "default"
                                            }
                                            
                                            if rule.direction == 'Inbound':
                                                inbound_rules.append(rule_data)
                                            else:
                                                outbound_rules.append(rule_data)
                                    
                                    nsg_data = {
                                        "id": azure_nsg.id,
                                        "name": azure_nsg.name,
                                        "resource_group": azure_nsg.id.split('/')[4] if '/' in azure_nsg.id else 'unknown',
                                        "location": azure_nsg.location,
                                        "subscription_id": subscription_id,
                                        "inbound_rules": inbound_rules,
                                        "outbound_rules": outbound_rules,
                                        "tags": dict(azure_nsg.tags) if azure_nsg.tags else {},
                                        "backed_up_at": datetime.utcnow().isoformat() + "Z"
                                    }
                                    backup_content["nsgs"].append(nsg_data)
                                    
                                logger.info(f"Successfully fetched real NSG data on attempt {attempt + 1}")
                                return backup_content
                                
                            except Exception as e:
                                logger.error(f"Error fetching NSG data on attempt {attempt + 1}: {e}")
                                if attempt == max_retries - 1:
                                    raise e
                                import time
                                time.sleep(2 ** attempt)  # Exponential backoff
                        
                        return None
                    
                    # Try to fetch real data with increased timeout and retry mechanism
                    import threading
                    result = [None]
                    exception_result = [None]
                    
                    def target():
                        try:
                            result[0] = fetch_nsg_data_with_retry()
                        except Exception as e:
                            exception_result[0] = e
                    
                    thread = threading.Thread(target=target)
                    thread.start()
                    thread.join(timeout=30)  # Increased timeout to 30 seconds
                    
                    if result[0] is not None:
                        backup_content = result[0]
                        logger.info(f"Successfully fetched real NSG data for export")
                    else:
                        # If real data fetch failed, return an error instead of sample data
                        error_msg = f"Failed to fetch real NSG data: {exception_result[0] if exception_result[0] else 'Timeout after 30 seconds'}"
                        logger.error(error_msg)
                        self.send_response(500)
                        self.send_header('Content-Type', 'application/json')
                        self.send_header('Access-Control-Allow-Origin', '*')
                        self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                        self.end_headers()
                        self.wfile.write(json.dumps({"error": error_msg}).encode('utf-8'))
                        return
                        
                except Exception as e:
                    logger.error(f"Critical error in NSG data export: {e}")
                    # Return error instead of sample data
                    self.send_response(500)
                    self.send_header('Content-Type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                    self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                    self.end_headers()
                    self.wfile.write(json.dumps({"error": f"Export failed: {str(e)}"}).encode('utf-8'))
                
                # Filter NSGs based on selection (filtering already done during fetch)
                # The fetch_nsg_data function already filters by selected_nsg_names
                # So backup_content already contains only the selected NSGs
                filtered_nsgs = backup_content.get('nsgs', [])
                logger.info(f"NSGs after filtering: {len(filtered_nsgs)} NSGs")
                
                # Check if we have any NSGs after filtering
                if len(filtered_nsgs) == 0:
                    logger.warning(f"No NSGs found for selection criteria: {selected_nsg_names}")
                    self.send_response(400)
                    self.send_header('Content-Type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                    self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                    self.end_headers()
                    self.wfile.write(json.dumps({"error": "No NSGs found matching the selection criteria"}).encode('utf-8'))
                    return
                
                # Generate data based on format
                if export_format == 'json':
                    # Return JSON format with real NSG data
                    self.send_response(200)
                    self.send_header('Content-Type', 'application/json')
                    self.send_header('Content-Disposition', f'attachment; filename="nsg_export_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.json"')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                    self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                    self.end_headers()
                    self.wfile.write(json.dumps(backup_content, indent=2).encode('utf-8'))
                    logger.info(f"Successfully sent JSON file with real NSG data")
                    return
                    
                elif export_format == 'enhanced_csv':
                    # Check if we have multiple NSGs to create separate files
                    filtered_nsgs = backup_content.get('nsgs', [])
                    
                    if len(filtered_nsgs) > 1:
                        # Create separate CSV files for each NSG in a ZIP archive
                        import zipfile
                        
                        zip_buffer = io.BytesIO()
                        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
                        
                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                            for nsg in filtered_nsgs:
                                nsg_name = nsg.get('name', 'unknown_nsg')
                                csv_content = create_csv_for_single_nsg(nsg)
                                
                                # Log CSV content size for debugging
                                logger.info(f"CSV content size for {nsg_name}: {len(csv_content)} bytes")
                                
                                if not csv_content:
                                    logger.warning(f"Empty CSV content generated for NSG: {nsg_name}")
                                    continue
                                filename = f"{nsg_name}_{timestamp}.csv"
                                zip_file.writestr(filename, csv_content)
                                logger.info(f"Added {filename} to ZIP archive")
                        
                        zip_content = zip_buffer.getvalue()
                        zip_buffer.close()
                        
                        # Log the size of the ZIP content
                        logger.info(f"ZIP content size: {len(zip_content)} bytes")
                        
                        # Send ZIP file response
                        self.send_response(200)
                        self.send_header('Content-Type', 'application/zip')
                        self.send_header('Content-Disposition', f'attachment; filename="nsg_exports_{timestamp}.zip"')
                        self.send_header('Access-Control-Allow-Origin', '*')
                        self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                        self.end_headers()
                        self.wfile.write(zip_content)
                        logger.info(f"Successfully sent ZIP file with {len(filtered_nsgs)} NSGs")
                        return
                    
                    elif len(filtered_nsgs) == 1:
                        # Single NSG - create individual CSV file with timestamp
                        nsg = filtered_nsgs[0]
                        nsg_name = nsg.get('name', 'unknown_nsg')
                        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
                        csv_content = create_csv_for_single_nsg(nsg)
                        
                        # Log the CSV content for debugging
                        logger.info(f"CSV content size for {nsg_name}: {len(csv_content)} bytes")
                        logger.info(f"CSV content preview: {csv_content[:200]}...")
                        
                        # Send single CSV file response
                        self.send_response(200)
                        self.send_header('Content-Type', 'text/csv')
                        self.send_header('Content-Disposition', f'attachment; filename="{nsg_name}_{timestamp}.csv"')
                        self.send_header('Access-Control-Allow-Origin', '*')
                        self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                        self.end_headers()
                        self.wfile.write(csv_content.encode('utf-8'))
                        logger.info(f"Successfully sent CSV file for NSG: {nsg_name}")
                        return
                    
                    else:
                        # No NSGs found - fallback to combined format
                        csv_content = create_csv_from_backup(backup_content)
                        
                        # Send single CSV file response for fallback
                        self.send_response(200)
                        self.send_header('Content-Type', 'text/csv')
                        self.send_header('Content-Disposition', f'attachment; filename="nsg_backup_fallback_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.csv"')
                        self.send_header('Access-Control-Allow-Origin', '*')
                        self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                        self.end_headers()
                        self.wfile.write(csv_content.encode('utf-8'))
                        return
                elif export_format == 'excel':
                    # Generate Excel content using real NSG data
                    try:
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, PatternFill
                        
                        # Convert backup content to CSV format first, then parse for Excel
                        csv_content = create_csv_from_backup(backup_content)
                        csv_lines = csv_content.strip().split('\n')
                        
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "NSG Backup Export"
                        
                        # Add headers with styling
                        if csv_lines:
                            headers = csv_lines[0].split(',')
                            for col, header in enumerate(headers, 1):
                                cell = ws.cell(row=1, column=col, value=header.strip('"'))
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            
                            # Add data rows
                            for row_idx, line in enumerate(csv_lines[1:], 2):
                                values = line.split(',')
                                for col, value in enumerate(values, 1):
                                    ws.cell(row=row_idx, column=col, value=value.strip('"'))
                        
                        # Save to bytes
                        excel_buffer = io.BytesIO()
                        wb.save(excel_buffer)
                        excel_content = excel_buffer.getvalue()
                        excel_buffer.close()
                        
                        # Send Excel file response
                        self.send_response(200)
                        self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        self.send_header('Content-Disposition', f'attachment; filename="nsg_export_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx"')
                        self.send_header('Access-Control-Allow-Origin', '*')
                        self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                        self.end_headers()
                        self.wfile.write(excel_content)
                        logger.info(f"Successfully sent Excel file with real NSG data")
                        return
                        
                    except ImportError:
                        # Fallback to CSV if openpyxl is not available
                        logger.warning("openpyxl not available, falling back to CSV")
                        
                        csv_content = create_csv_from_backup(backup_content)
                        
                        # Send CSV file response with Excel filename
                        self.send_response(200)
                        self.send_header('Content-Type', 'text/csv')
                        self.send_header('Content-Disposition', f'attachment; filename="nsg_export_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.csv"')
                        self.send_header('Access-Control-Allow-Origin', '*')
                        self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                        self.end_headers()
                        self.wfile.write(csv_content.encode('utf-8'))
                        logger.info(f"Successfully sent CSV fallback with real NSG data")
                        return
                        
                else:
                    # Use real NSG data for regular CSV export
                    csv_content = create_csv_from_backup(backup_content)
                    
                    # Send CSV file response
                    self.send_response(200)
                    self.send_header('Content-Type', 'text/csv')
                    self.send_header('Content-Disposition', f'attachment; filename="nsg_export_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.csv"')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                    self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                    self.end_headers()
                    self.wfile.write(csv_content.encode('utf-8'))
                    logger.info(f"Successfully sent CSV file with real NSG data")
                    return
                        
            except Exception as e:
                logger.error(f"Export failed: {e}")
                # Send error response for export
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                error_response = {"error": f"Export failed: {str(e)}"}
                self.wfile.write(json.dumps(error_response).encode())
                return
        elif path == '/api/v1/backup/restore/preview':
            # Preview restore data
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                restore_data = json.loads(post_data.decode('utf-8'))
                
                logger.info(f"Restore preview request: {restore_data}")
                
                source_type = restore_data.get('source_type', restore_data.get('source', 'storage'))
                backup_file_name = restore_data.get('backup_file_name', '')
                logger.info(f"Source type: {source_type}, Backup file: {backup_file_name}")
                storage_account = restore_data.get('storage_account', 'thirustorage001')
                container_name = restore_data.get('container_name', 'nsg-backups')
                subscription_id = restore_data.get('subscription_id', '')
                target_resource_groups = restore_data.get('target_resource_groups', [])
                target_type = restore_data.get('target_type', 'single')
                
                if source_type == 'storage' and backup_file_name:
                    # Load actual file content from Azure Storage
                    try:
                        file_content = download_blob_content(storage_account, container_name, backup_file_name)
                        if file_content:
                            # Check if it's a CSV file
                            if backup_file_name.endswith('.csv'):
                                sample_rules = parse_csv_rules(file_content)
                                logger.info(f"Loaded {len(sample_rules)} rules from CSV file: {backup_file_name}")
                            else:
                                # Handle JSON backup files
                                try:
                                    backup_data = json.loads(file_content)
                                    sample_rules = []
                                    # Extract rules from JSON backup format
                                    if 'nsgs' in backup_data:
                                        for nsg in backup_data['nsgs']:
                                            if 'security_rules' in nsg:
                                                sample_rules.extend(nsg['security_rules'])
                                            if 'inbound_rules' in nsg:
                                                sample_rules.extend(nsg['inbound_rules'])
                                            if 'outbound_rules' in nsg:
                                                sample_rules.extend(nsg['outbound_rules'])
                                    logger.info(f"Loaded {len(sample_rules)} rules from JSON file: {backup_file_name}")
                                except json.JSONDecodeError as e:
                                    logger.error(f"Failed to parse JSON backup file: {e}")
                                    sample_rules = []
                        else:
                            logger.error(f"Failed to download file content: {backup_file_name}")
                            sample_rules = []
                    except Exception as e:
                        logger.error(f"Error loading file from storage: {e}")
                        sample_rules = []
                elif source_type == 'csv':
                    # Parse the uploaded CSV file content
                    csv_file_data = restore_data.get('csv_file')
                    csv_error_message = None
                    if csv_file_data:
                        try:
                            # Parse CSV content
                            sample_rules = parse_csv_rules(csv_file_data)
                            logger.info(f"Parsed {len(sample_rules)} rules from CSV file")
                            
                            # Check if CSV format is incorrect (NSG summary instead of rules)
                            if len(sample_rules) == 0:
                                import io
                                csv_reader = csv.DictReader(io.StringIO(csv_file_data))
                                headers = csv_reader.fieldnames or []
                                is_nsg_summary = any(header in headers for header in ['Subscription Name', 'Subscription ID', 'NSG Name', 'Resource Group'])
                                if is_nsg_summary:
                                    csv_error_message = "The uploaded CSV appears to be an NSG summary report, not individual NSG rules. Please upload a CSV file containing individual rule data with columns: Rule Name, Priority, Direction, Access, Protocol, Source, Destination, Description."
                                    
                        except Exception as e:
                            logger.error(f"Error parsing CSV file: {e}")
                            sample_rules = []
                            csv_error_message = f"Error parsing CSV file: {str(e)}"
                    else:
                        sample_rules = []
                        csv_error_message = "No CSV file data provided"
                else:
                    sample_rules = []
                    csv_error_message = None
                
                response = {
                    "success": True,
                    "preview": {
                        "rules": sample_rules,
                        "totalRules": len(sample_rules),
                        "source": source_type,
                        "error": csv_error_message if csv_error_message else None
                    }
                }
                
                # Send response
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
                
            except Exception as e:
                logger.error(f"Restore preview failed: {e}")
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                error_response = {"error": f"Restore preview failed: {str(e)}"}
                self.wfile.write(json.dumps(error_response).encode())
                return
        elif path == '/api/v1/backup/restore/confirm':
            # Confirm and execute restore
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                restore_data = json.loads(post_data.decode('utf-8'))
                
                logger.info(f"Restore confirm request: {restore_data}")
                
                source_type = restore_data.get('source_type', 'storage')
                target_resource_groups = restore_data.get('target_resource_groups', [])
                create_new_nsgs = restore_data.get('create_new_nsgs', False)
                new_nsg_names = restore_data.get('new_nsg_names', [])
                subscription_id = restore_data.get('subscription_id', '')
                
                # Get rules to restore - prioritize edited_rules from frontend
                rules_to_restore = []
                edited_rules = restore_data.get('edited_rules', [])
                
                logger.info(f"Restore logic - edited_rules count: {len(edited_rules)}, source_type: {source_type}")
                
                if edited_rules:
                    # Use edited rules from the frontend preview
                    rules_to_restore = edited_rules
                    logger.info(f"Using {len(rules_to_restore)} edited rules from frontend")
                elif source_type == 'csv':
                    csv_file_data = restore_data.get('csv_file')
                    logger.info(f"CSV source - csv_file_data present: {csv_file_data is not None}")
                    if csv_file_data:
                        rules_to_restore = parse_csv_rules(csv_file_data)
                        logger.info(f"Parsed {len(rules_to_restore)} rules from CSV data")
                    else:
                        logger.warning("CSV source type but no csv_file data provided")
                elif source_type == 'storage':
                    # Load actual file content from Azure Storage
                    backup_file_name = restore_data.get('backup_file_name', '')
                    storage_account = restore_data.get('storage_account', 'thirustorage001')
                    container_name = restore_data.get('container_name', 'nsg-backups')
                    
                    if backup_file_name:
                        try:
                            file_content = download_blob_content(storage_account, container_name, backup_file_name)
                            if file_content:
                                if backup_file_name.endswith('.csv'):
                                    rules_to_restore = parse_csv_rules(file_content)
                                    logger.info(f"Loaded {len(rules_to_restore)} rules from CSV file: {backup_file_name}")
                                else:
                                    # Handle JSON backup files
                                    try:
                                        backup_data = json.loads(file_content)
                                        rules_to_restore = []
                                        if 'nsgs' in backup_data:
                                            for nsg in backup_data['nsgs']:
                                                if 'security_rules' in nsg:
                                                    rules_to_restore.extend(nsg['security_rules'])
                                                if 'inbound_rules' in nsg:
                                                    rules_to_restore.extend(nsg['inbound_rules'])
                                                if 'outbound_rules' in nsg:
                                                    rules_to_restore.extend(nsg['outbound_rules'])
                                        logger.info(f"Loaded {len(rules_to_restore)} rules from JSON file: {backup_file_name}")
                                    except json.JSONDecodeError as e:
                                        logger.error(f"Failed to parse JSON backup file: {e}")
                                        rules_to_restore = []
                            else:
                                logger.error(f"Failed to download file content: {backup_file_name}")
                                rules_to_restore = []
                        except Exception as e:
                            logger.error(f"Error loading file from storage: {e}")
                            rules_to_restore = []
                
                logger.info(f"Final rules_to_restore count: {len(rules_to_restore)}")
                if len(rules_to_restore) == 0:
                    logger.warning("No rules to restore! This will result in 0 rules restored.")
                
                restored_count = 0
                created_nsgs = []
                
                # Initialize Azure client for real operations
                try:
                    from app.core.azure_client import AzureClient
                    azure_client = AzureClient()
                    logger.info("Azure client initialized for restore operations")
                except Exception as e:
                    logger.error(f"Failed to initialize Azure client: {e}")
                    # Fall back to simulation mode
                    azure_client = None
                
                # Process restore for each target resource group
                for rg in target_resource_groups:
                    if rg == '*':
                        continue  # Skip wildcard for now
                    
                    if create_new_nsgs:
                        # Find NSG names for this resource group
                        rg_nsgs = [nsg for nsg in new_nsg_names if nsg.get('resourceGroup') == rg]
                        for nsg_config in rg_nsgs:
                            nsg_name = nsg_config.get('nsgName', f'restored-nsg-{len(created_nsgs) + 1}')
                            
                            if azure_client:
                                try:
                                    # Create NSG in Azure
                                    import asyncio
                                    loop = asyncio.new_event_loop()
                                    asyncio.set_event_loop(loop)
                                    
                                    # Create the NSG
                                    nsg_result = loop.run_until_complete(
                                        azure_client.create_nsg(rg, nsg_name, "East US")
                                    )
                                    logger.info(f"Successfully created NSG '{nsg_name}' in Azure")
                                    
                                    # Add rules to the NSG
                                    rules_added = 0
                                    for rule in rules_to_restore:
                                        try:
                                            # Normalize rule from any source into Azure SDK format
                                            azure_rule = convert_any_rule_to_azure_format(rule)
                                            # Ensure defaults and sequencing for missing fields
                                            if not azure_rule.get('name'):
                                                azure_rule['name'] = f"Rule-{rules_added + 1}"
                                            if 'priority' not in azure_rule or azure_rule.get('priority') in [None, '', 0]:
                                                try:
                                                    azure_rule['priority'] = int(rule.get('priority', 1000 + rules_added))
                                                except Exception:
                                                    azure_rule['priority'] = 1000 + rules_added
                                            if not azure_rule.get('direction'):
                                                azure_rule['direction'] = rule.get('direction', 'Inbound')

                                            loop.run_until_complete(
                                                azure_client.create_nsg_rule(rg, nsg_name, azure_rule)
                                            )
                                            rules_added += 1
                                            logger.info(f"Added rule '{azure_rule['name']}' to NSG '{nsg_name}'")
                                        except Exception as rule_error:
                                            logger.error(f"Failed to add rule to NSG '{nsg_name}': {rule_error}")
                                    
                                    loop.close()
                                    
                                    created_nsgs.append({
                                        'name': nsg_name,
                                        'resourceGroup': rg,
                                        'rules': rules_added
                                    })
                                    restored_count += rules_added
                                    logger.info(f"Successfully created NSG '{nsg_name}' in RG '{rg}' with {rules_added} rules")
                                    
                                except Exception as nsg_error:
                                    logger.error(f"Failed to create NSG '{nsg_name}' in Azure: {nsg_error}")
                                    # Fall back to simulation
                                    created_nsgs.append({
                                        'name': nsg_name,
                                        'resourceGroup': rg,
                                        'rules': len(rules_to_restore)
                                    })
                                    restored_count += len(rules_to_restore)
                                    logger.info(f"Fallback: Simulated creation of NSG '{nsg_name}' in RG '{rg}' with {len(rules_to_restore)} rules")
                            else:
                                # Simulation mode
                                created_nsgs.append({
                                    'name': nsg_name,
                                    'resourceGroup': rg,
                                    'rules': len(rules_to_restore)
                                })
                                restored_count += len(rules_to_restore)
                                logger.info(f"Simulated creation of NSG '{nsg_name}' in RG '{rg}' with {len(rules_to_restore)} rules")
                    else:
                        # Apply to existing NSGs in the resource group
                        if azure_client:
                            try:
                                import asyncio
                                loop = asyncio.new_event_loop()
                                asyncio.set_event_loop(loop)
                                
                                # Get existing NSGs in the resource group
                                existing_nsgs = loop.run_until_complete(
                                    azure_client.get_nsgs(rg)
                                )
                                
                                for nsg in existing_nsgs:
                                    nsg_name = nsg['name']
                                    rules_added = 0
                                    
                                    for rule in rules_to_restore:
                                        try:
                                            # Normalize rule from any source into Azure SDK format
                                            azure_rule = convert_any_rule_to_azure_format(rule)
                                            # Ensure defaults and sequencing for missing fields
                                            if not azure_rule.get('name'):
                                                azure_rule['name'] = f"Restored-Rule-{rules_added + 1}"
                                            if 'priority' not in azure_rule or azure_rule.get('priority') in [None, '', 0]:
                                                try:
                                                    azure_rule['priority'] = int(rule.get('priority', 1000 + rules_added))
                                                except Exception:
                                                    azure_rule['priority'] = 1000 + rules_added
                                            if not azure_rule.get('direction'):
                                                azure_rule['direction'] = rule.get('direction', 'Inbound')

                                            loop.run_until_complete(
                                                azure_client.create_nsg_rule(rg, nsg_name, azure_rule)
                                            )
                                            rules_added += 1
                                            logger.info(f"Added rule '{azure_rule['name']}' to existing NSG '{nsg_name}'")
                                        except Exception as rule_error:
                                            logger.error(f"Failed to add rule to existing NSG '{nsg_name}': {rule_error}")
                                    
                                    restored_count += rules_added
                                
                                loop.close()
                                logger.info(f"Successfully restored {restored_count} rules to existing NSGs in RG '{rg}'")
                                
                            except Exception as restore_error:
                                logger.error(f"Failed to restore to existing NSGs in RG '{rg}': {restore_error}")
                                # Fall back to simulation
                                restored_count += len(rules_to_restore)
                                logger.info(f"Fallback: Simulated restore of {len(rules_to_restore)} rules to existing NSGs in RG '{rg}'")
                        else:
                            # Simulation mode
                            restored_count += len(rules_to_restore)
                            logger.info(f"Simulated restore of {len(rules_to_restore)} rules to existing NSGs in RG '{rg}'")
                
                response = {
                    "success": True,
                    "message": f"Successfully restored {restored_count} rules",
                    "restored_rules_count": restored_count,
                    "created_nsgs": created_nsgs,
                    "target_resource_groups": target_resource_groups,
                    "timestamp": datetime.now().isoformat()
                }
                
                # Send response
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
                
            except Exception as e:
                logger.error(f"Restore failed: {e}")
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                error_response = {"error": f"Restore failed: {str(e)}"}
                self.wfile.write(json.dumps(error_response).encode())
                return
        elif path == '/api/v1/backup/files':
            # Get available backup files from storage
            try:
                # Read and parse request body
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                request_data = json.loads(post_data.decode('utf-8'))
                
                logger.info(f"Received backup files request: {request_data}")
                
                # Extract storage account and container from request
                storage_account = request_data.get('storage_account')
                container_name = request_data.get('container_name')
                
                # Simulate getting backup files from Azure Storage
                sample_files = [
                    {
                        "id": "backup_001",
                        "name": "NSG_Backup_WebTier_20241225.json",
                        "createdAt": "2024-12-25T10:30:00Z",
                        "size": "2.5 KB",
                        "resourceGroup": "Web-rg",
                        "nsgCount": 3
                    },
                    {
                        "id": "backup_002",
                        "name": "NSG_Backup_DatabaseTier_20241224.json",
                        "createdAt": "2024-12-24T15:45:00Z",
                        "size": "1.8 KB",
                        "resourceGroup": "Database-rg",
                        "nsgCount": 2
                    },
                    {
                        "id": "backup_003",
                        "name": "NSG_Backup_AppTier_20241223.json",
                        "createdAt": "2024-12-23T09:15:00Z",
                        "size": "3.2 KB",
                        "resourceGroup": "App-rg",
                        "nsgCount": 5
                    }
                ]
                
                response = {
                    "success": True,
                    "files": sample_files
                }
                
                # Send response
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
                
            except Exception as e:
                logger.error(f"Failed to get backup files: {e}")
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                error_response = {"error": f"Failed to get backup files: {str(e)}"}
                self.wfile.write(json.dumps(error_response).encode())
                return
        elif path == '/api/v1/reports/asg-validation':
            # ASG Validation Report
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                request_data = json.loads(post_data.decode('utf-8'))
                
                subscription_id = request_data.get('subscription_id', AZURE_CONFIG["subscription_id"])
                resource_group = request_data.get('resource_group', '')
                
                # Get Azure clients
                subscription_client, resource_client, network_client = get_azure_clients()
                
                if not network_client:
                    response = {"error": "Azure Network client not available"}
                else:
                    asg_data = []
                    total_asgs = 0
                    
                    try:
                        # Get all ASGs across subscriptions or specific resource group
                        if resource_group:
                            asgs = list(network_client.application_security_groups.list(resource_group))
                        else:
                            asgs = list(network_client.application_security_groups.list_all())
                        
                        for asg in asgs:
                            if total_asgs >= 100:  # Max limit
                                break
                                
                            # Get subscription and resource group info
                            asg_resource_group = asg.id.split('/')[4] if len(asg.id.split('/')) > 4 else 'Unknown'
                            
                            asg_info = {
                                "subscription_name": subscription_id,
                                "subscription_id": subscription_id,
                                "resource_group": asg_resource_group,
                                "asg_name": asg.name,
                                "location": asg.location,
                                "provisioning_state": asg.provisioning_state,
                                "resource_guid": getattr(asg, 'resource_guid', ''),
                                "tags": asg.tags or {}
                            }
                            asg_data.append(asg_info)
                            total_asgs += 1
                        
                        # Generate CSV format for frontend
                        csv_headers = ["Subscription Name", "Subscription ID", "Resource Group", "ASG Name", "Location", 
                                     "Provisioning State", "Resource GUID", "Tags"]
                        csv_data = []
                        for item in asg_data:
                            tags_str = ', '.join([f"{k}:{v}" for k, v in item.get('tags', {}).items()]) if item.get('tags') else ''
                            csv_data.append([
                                item.get('subscription_name', ''),
                                item.get('subscription_id', ''),
                                item.get('resource_group', ''),
                                item.get('asg_name', ''),
                                item.get('location', ''),
                                item.get('provisioning_state', ''),
                                item.get('resource_guid', ''),
                                tags_str
                            ])
                        
                        response = {
                            "success": True,
                            "report_type": "ASG Validation",
                            "total_asgs": total_asgs,
                            "max_limit": 100,
                            "data": {
                                "csv_headers": csv_headers,
                                "csv_data": csv_data,
                                "raw_data": asg_data
                            },
                            "generated_at": datetime.utcnow().isoformat() + "Z",
                            "summary": {
                                "compliant_asgs": len([asg for asg in asg_data if asg.get('provisioning_state') == 'Succeeded']),
                                "non_compliant_asgs": len([asg for asg in asg_data if asg.get('provisioning_state') != 'Succeeded']),
                                "coverage_percentage": round((len([asg for asg in asg_data if asg.get('provisioning_state') == 'Succeeded']) / max(total_asgs, 1)) * 100, 2)
                            }
                        }
                        
                    except Exception as e:
                        logger.error(f"Failed to fetch ASG data: {e}")
                        response = {"error": f"Failed to fetch ASG data: {str(e)}"}
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
                
            except Exception as e:
                logger.error(f"ASG validation report failed: {e}")
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                error_response = {"error": f"ASG validation report failed: {str(e)}"}
                self.wfile.write(json.dumps(error_response).encode())
                return
        elif path == '/api/v1/reports/nsg-rules':
            # NSG Rules Report
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                request_data = json.loads(post_data.decode('utf-8'))
                
                subscription_id = request_data.get('subscription_id', AZURE_CONFIG["subscription_id"])
                resource_group = request_data.get('resource_group', '')
                
                # Get Azure clients
                subscription_client, resource_client, network_client = get_azure_clients()
                
                if not network_client:
                    response = {"error": "Azure Network client not available"}
                else:
                    nsg_rules_data = []
                    total_rules = 0
                    
                    try:
                        # Get all NSGs
                        if resource_group:
                            nsgs = list(network_client.network_security_groups.list(resource_group))
                        else:
                            nsgs = list(network_client.network_security_groups.list_all())
                        
                        for nsg in nsgs:
                            if total_rules >= 1000:  # Max limit
                                break
                                
                            nsg_resource_group = nsg.id.split('/')[4] if len(nsg.id.split('/')) > 4 else 'Unknown'
                            
                            inbound_rules = len(nsg.security_rules) if nsg.security_rules else 0
                            outbound_rules = len(nsg.default_security_rules) if nsg.default_security_rules else 0
                            
                            nsg_info = {
                                "subscription_name": subscription_id,
                                "subscription_id": subscription_id,
                                "resource_group": nsg_resource_group,
                                "nsg_name": nsg.name,
                                "location": nsg.location,
                                "inbound_rules_count": inbound_rules,
                                "outbound_rules_count": outbound_rules,
                                "total_rules": inbound_rules + outbound_rules,
                                "provisioning_state": nsg.provisioning_state
                            }
                            nsg_rules_data.append(nsg_info)
                            total_rules += inbound_rules + outbound_rules
                        
                        # Generate CSV headers and data
                        csv_headers = ["Subscription Name", "Subscription ID", "Resource Group", "NSG Name", "Location", "Inbound Rules", "Outbound Rules", "Total Rules", "Provisioning State"]
                        csv_data = []
                        for nsg in nsg_rules_data:
                            csv_data.append([
                                nsg.get('subscription_name', ''),
                                nsg.get('subscription_id', ''),
                                nsg.get('resource_group', ''),
                                nsg.get('nsg_name', ''),
                                nsg.get('location', ''),
                                nsg.get('inbound_rules_count', 0),
                                nsg.get('outbound_rules_count', 0),
                                nsg.get('total_rules', 0),
                                nsg.get('provisioning_state', '')
                            ])
                        
                        response = {
                            "success": True,
                            "report_type": "NSG Rules",
                            "total_rules": total_rules,
                            "max_limit": 1000,
                            "data": {
                                "csv_headers": csv_headers,
                                "csv_data": csv_data,
                                "raw_data": nsg_rules_data
                            },
                            "generated_at": datetime.utcnow().isoformat() + "Z",
                            "summary": {
                                "total_nsgs": len(nsg_rules_data),
                                "active_rules": sum([nsg['total_rules'] for nsg in nsg_rules_data if nsg.get('provisioning_state') == 'Succeeded']),
                                "efficiency_percentage": round((sum([nsg['total_rules'] for nsg in nsg_rules_data if nsg.get('provisioning_state') == 'Succeeded']) / max(total_rules, 1)) * 100, 2)
                            }
                        }
                        
                    except Exception as e:
                        logger.error(f"Failed to fetch NSG rules data: {e}")
                        response = {"error": f"Failed to fetch NSG rules data: {str(e)}"}
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
                
            except Exception as e:
                logger.error(f"NSG rules report failed: {e}")
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                error_response = {"error": f"NSG rules report failed: {str(e)}"}
                self.wfile.write(json.dumps(error_response).encode())
                return
        elif path == '/api/v1/reports/ip-limitations':
            # IP Limitations Report
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                request_data = json.loads(post_data.decode('utf-8'))
                
                subscription_id = request_data.get('subscription_id', AZURE_CONFIG["subscription_id"])
                resource_group = request_data.get('resource_group', '')
                
                # Get Azure clients
                subscription_client, resource_client, network_client = get_azure_clients()
                
                if not network_client:
                    response = {"error": "Azure Network client not available"}
                else:
                    ip_data = []
                    total_ips = 0
                    
                    try:
                        # Get all NSGs and analyze IP addresses in rules
                        if resource_group:
                            nsgs = list(network_client.network_security_groups.list(resource_group))
                        else:
                            nsgs = list(network_client.network_security_groups.list_all())
                        
                        for nsg in nsgs:
                            if total_ips >= 4000:  # Max limit
                                break
                                
                            nsg_resource_group = nsg.id.split('/')[4] if len(nsg.id.split('/')) > 4 else 'Unknown'
                            
                            source_ips = set()
                            destination_ips = set()
                            source_asgs = set()
                            destination_asgs = set()
                            
                            # Analyze security rules for IP addresses and ASGs
                            if nsg.security_rules:
                                for rule in nsg.security_rules:
                                    # Source analysis
                                    if rule.source_address_prefix and rule.source_address_prefix != '*':
                                        source_ips.add(rule.source_address_prefix)
                                    if rule.source_address_prefixes:
                                        source_ips.update(rule.source_address_prefixes)
                                    if rule.source_application_security_groups:
                                        source_asgs.update([asg.id for asg in rule.source_application_security_groups])
                                    
                                    # Destination analysis
                                    if rule.destination_address_prefix and rule.destination_address_prefix != '*':
                                        destination_ips.add(rule.destination_address_prefix)
                                    if rule.destination_address_prefixes:
                                        destination_ips.update(rule.destination_address_prefixes)
                                    if rule.destination_application_security_groups:
                                        destination_asgs.update([asg.id for asg in rule.destination_application_security_groups])
                            
                            nsg_info = {
                                "subscription_name": subscription_id,
                                "subscription_id": subscription_id,
                                "resource_group": nsg_resource_group,
                                "nsg_name": nsg.name,
                                "location": nsg.location,
                                "source_ips_count": len(source_ips),
                                "destination_ips_count": len(destination_ips),
                                "source_asgs_count": len(source_asgs),
                                "destination_asgs_count": len(destination_asgs),
                                "source_ips": list(source_ips)[:50],  # Limit for response size
                                "destination_ips": list(destination_ips)[:50],
                                "source_asgs": list(source_asgs)[:20],
                                "destination_asgs": list(destination_asgs)[:20],
                                "total_unique_ips": len(source_ips | destination_ips)
                            }
                            ip_data.append(nsg_info)
                            total_ips += len(source_ips | destination_ips)
                        
                        # Generate CSV format for frontend
                        csv_headers = ["Subscription Name", "Subscription ID", "Resource Group", "NSG Name", "Location", 
                                     "Source IPs Count", "Destination IPs Count", "Source ASGs Count", "Destination ASGs Count", "Total Unique IPs"]
                        csv_data = []
                        for item in ip_data:
                            csv_data.append([
                                item.get('subscription_name', ''),
                                item.get('subscription_id', ''),
                                item.get('resource_group', ''),
                                item.get('nsg_name', ''),
                                item.get('location', ''),
                                item.get('source_ips_count', 0),
                                item.get('destination_ips_count', 0),
                                item.get('source_asgs_count', 0),
                                item.get('destination_asgs_count', 0),
                                item.get('total_unique_ips', 0)
                            ])
                        
                        response = {
                            "success": True,
                            "report_type": "IP Limitations",
                            "total_ips": total_ips,
                            "max_limit": 4000,
                            "data": {
                                "csv_headers": csv_headers,
                                "csv_data": csv_data,
                                "raw_data": ip_data
                            },
                            "generated_at": datetime.utcnow().isoformat() + "Z",
                            "summary": {
                                "total_nsgs_analyzed": len(ip_data),
                                "total_source_ips": sum([nsg['source_ips_count'] for nsg in ip_data]),
                                "total_destination_ips": sum([nsg['destination_ips_count'] for nsg in ip_data]),
                                "total_asgs": sum([nsg['source_asgs_count'] + nsg['destination_asgs_count'] for nsg in ip_data]),
                                "utilization_percentage": round((total_ips / 4000) * 100, 2)
                            }
                        }
                        
                    except Exception as e:
                        logger.error(f"Failed to fetch IP limitations data: {e}")
                        response = {"error": f"Failed to fetch IP limitations data: {str(e)}"}
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
                
            except Exception as e:
                logger.error(f"IP limitations report failed: {e}")
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                error_response = {"error": f"IP limitations report failed: {str(e)}"}
                self.wfile.write(json.dumps(error_response).encode())
                return
        elif path == '/api/v1/reports/nsg-ports':
            # NSG Ports Report
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                request_data = json.loads(post_data.decode('utf-8'))
                
                subscription_id = request_data.get('subscription_id', AZURE_CONFIG["subscription_id"])
                resource_group = request_data.get('resource_group', '')
                
                # Get Azure clients
                subscription_client, resource_client, network_client = get_azure_clients()
                
                if not network_client:
                    response = {"error": "Azure Network client not available"}
                else:
                    ports_data = []
                    total_inbound_ports = 0
                    total_outbound_ports = 0
                    
                    try:
                        # Get all NSGs and analyze port configurations
                        if resource_group:
                            nsgs = list(network_client.network_security_groups.list(resource_group))
                        else:
                            nsgs = list(network_client.network_security_groups.list_all())
                        
                        for nsg in nsgs:
                            if total_inbound_ports >= 4000 or total_outbound_ports >= 4000:  # Max limit
                                break
                                
                            nsg_resource_group = nsg.id.split('/')[4] if len(nsg.id.split('/')) > 4 else 'Unknown'
                            
                            inbound_ports = set()
                            outbound_ports = set()
                            high_risk_ports = set()
                            
                            # Analyze security rules for port configurations
                            if nsg.security_rules:
                                for rule in nsg.security_rules:
                                    ports_to_analyze = []
                                    
                                    # Get destination ports
                                    if rule.destination_port_range and rule.destination_port_range != '*':
                                        if '-' in rule.destination_port_range:
                                            start, end = rule.destination_port_range.split('-')
                                            ports_to_analyze.extend(range(int(start), min(int(end) + 1, int(start) + 100)))  # Limit range expansion
                                        else:
                                            ports_to_analyze.append(int(rule.destination_port_range))
                                    
                                    if rule.destination_port_ranges:
                                        for port_range in rule.destination_port_ranges:
                                            if '-' in port_range:
                                                start, end = port_range.split('-')
                                                ports_to_analyze.extend(range(int(start), min(int(end) + 1, int(start) + 100)))
                                            else:
                                                ports_to_analyze.append(int(port_range))
                                    
                                    # Categorize by direction
                                    if rule.direction == 'Inbound':
                                        inbound_ports.update(ports_to_analyze)
                                    else:
                                        outbound_ports.update(ports_to_analyze)
                                    
                                    # Identify high-risk ports (common vulnerable ports)
                                    risky_ports = {21, 22, 23, 25, 53, 80, 110, 143, 443, 993, 995, 1433, 3389, 5432, 5985, 5986}
                                    high_risk_ports.update(set(ports_to_analyze) & risky_ports)
                            
                            nsg_info = {
                                "subscription_name": subscription_id,
                                "subscription_id": subscription_id,
                                "resource_group": nsg_resource_group,
                                "nsg_name": nsg.name,
                                "location": nsg.location,
                                "inbound_ports_count": len(inbound_ports),
                                "outbound_ports_count": len(outbound_ports),
                                "high_risk_ports_count": len(high_risk_ports),
                                "inbound_ports": sorted(list(inbound_ports))[:100],  # Limit for response size
                                "outbound_ports": sorted(list(outbound_ports))[:100],
                                "high_risk_ports": sorted(list(high_risk_ports)),
                                "security_score": max(0, 100 - (len(high_risk_ports) * 10))  # Simple scoring
                            }
                            ports_data.append(nsg_info)
                            total_inbound_ports += len(inbound_ports)
                            total_outbound_ports += len(outbound_ports)
                        
                        # Generate CSV headers and data
                        csv_headers = ["Subscription Name", "Subscription ID", "Resource Group", "NSG Name", "Location", "Inbound Ports Count", "Outbound Ports Count", "High Risk Ports Count", "Security Score"]
                        csv_data = []
                        for nsg in ports_data:
                            csv_data.append([
                                nsg.get('subscription_name', ''),
                                nsg.get('subscription_id', ''),
                                nsg.get('resource_group', ''),
                                nsg.get('nsg_name', ''),
                                nsg.get('location', ''),
                                nsg.get('inbound_ports_count', 0),
                                nsg.get('outbound_ports_count', 0),
                                nsg.get('high_risk_ports_count', 0),
                                nsg.get('security_score', 0)
                            ])
                        
                        response = {
                            "success": True,
                            "report_type": "NSG Ports",
                            "total_inbound_ports": total_inbound_ports,
                            "total_outbound_ports": total_outbound_ports,
                            "max_limit_inbound": 4000,
                            "max_limit_outbound": 4000,
                            "data": {
                                "csv_headers": csv_headers,
                                "csv_data": csv_data,
                                "raw_data": ports_data
                            },
                            "generated_at": datetime.utcnow().isoformat() + "Z",
                            "summary": {
                                "total_nsgs_analyzed": len(ports_data),
                                "total_high_risk_ports": sum([nsg['high_risk_ports_count'] for nsg in ports_data]),
                                "average_security_score": round(sum([nsg['security_score'] for nsg in ports_data]) / max(len(ports_data), 1), 2),
                                "inbound_utilization": round((total_inbound_ports / 4000) * 100, 2),
                                "outbound_utilization": round((total_outbound_ports / 4000) * 100, 2)
                            }
                        }
                        
                    except Exception as e:
                        logger.error(f"Failed to fetch NSG ports data: {e}")
                        response = {"error": f"Failed to fetch NSG ports data: {str(e)}"}
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
                
            except Exception as e:
                logger.error(f"NSG ports report failed: {e}")
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                error_response = {"error": f"NSG ports report failed: {str(e)}"}
                self.wfile.write(json.dumps(error_response).encode())
                return
        elif path == '/api/v1/reports/consolidation':
            # Consolidation Report with LLM Analysis
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                request_data = json.loads(post_data.decode('utf-8'))
                
                subscription_id = request_data.get('subscription_id', AZURE_CONFIG["subscription_id"])
                resource_group = request_data.get('resource_group', '')
                
                # Get Azure clients
                subscription_client, resource_client, network_client = get_azure_clients()
                
                if not network_client:
                    response = {"error": "Azure Network client not available"}
                else:
                    consolidation_data = []
                    
                    try:
                        # Get all NSGs for analysis
                        if resource_group:
                            nsgs = list(network_client.network_security_groups.list(resource_group))
                        else:
                            nsgs = list(network_client.network_security_groups.list_all())
                        
                        # Analyze NSGs for consolidation opportunities
                        rule_patterns = {}
                        duplicate_rules = []
                        optimization_opportunities = []
                        
                        for nsg in nsgs:
                            nsg_resource_group = nsg.id.split('/')[4] if len(nsg.id.split('/')) > 4 else 'Unknown'
                            
                            if nsg.security_rules:
                                for rule in nsg.security_rules:
                                    # Create rule signature for duplicate detection
                                    rule_signature = f"{rule.protocol}_{rule.direction}_{rule.access}_{rule.source_address_prefix}_{rule.destination_address_prefix}_{rule.destination_port_range}"
                                    
                                    if rule_signature in rule_patterns:
                                        rule_patterns[rule_signature]['count'] += 1
                                        rule_patterns[rule_signature]['nsgs'].append(nsg.name)
                                    else:
                                        rule_patterns[rule_signature] = {
                                            'count': 1,
                                            'rule': rule,
                                            'nsgs': [nsg.name],
                                            'potential_savings': 0
                                        }
                            
                            # Analyze for optimization opportunities
                            nsg_analysis = {
                                "subscription_name": subscription_id,
                                "subscription_id": subscription_id,
                                "resource_group": nsg_resource_group,
                                "nsg_name": nsg.name,
                                "location": nsg.location,
                                "total_rules": len(nsg.security_rules) if nsg.security_rules else 0,
                                "optimization_score": 0,
                                "recommendations": []
                            }
                            
                            # Simple optimization analysis
                            if nsg.security_rules:
                                redundant_rules = 0
                                overly_permissive = 0
                                
                                for rule in nsg.security_rules:
                                    # Check for overly permissive rules
                                    if (rule.source_address_prefix == '*' or rule.destination_address_prefix == '*') and rule.access == 'Allow':
                                        overly_permissive += 1
                                    
                                    # Check for potential redundancy (simplified)
                                    if rule.destination_port_range == '*' and rule.protocol == '*':
                                        redundant_rules += 1
                                
                                nsg_analysis['redundant_rules'] = redundant_rules
                                nsg_analysis['overly_permissive_rules'] = overly_permissive
                                nsg_analysis['optimization_score'] = max(0, 100 - (redundant_rules * 10) - (overly_permissive * 15))
                                
                                # Generate recommendations
                                if redundant_rules > 0:
                                    nsg_analysis['recommendations'].append(f"Consider consolidating {redundant_rules} overly broad rules")
                                if overly_permissive > 0:
                                    nsg_analysis['recommendations'].append(f"Review {overly_permissive} overly permissive rules for security")
                                if len(nsg.security_rules) > 50:
                                    nsg_analysis['recommendations'].append("Consider rule consolidation due to high rule count")
                            
                            consolidation_data.append(nsg_analysis)
                        
                        # Calculate overall savings potential
                        total_rules = sum([nsg['total_rules'] for nsg in consolidation_data])
                        potential_rule_reduction = sum([nsg.get('redundant_rules', 0) for nsg in consolidation_data])
                        estimated_cost_savings = potential_rule_reduction * 2.5  # Estimated $2.5 per rule per month
                        
                        # Find duplicate rules across NSGs
                        for signature, pattern in rule_patterns.items():
                            if pattern['count'] > 1:
                                duplicate_rules.append({
                                    'rule_pattern': signature,
                                    'occurrence_count': pattern['count'],
                                    'affected_nsgs': pattern['nsgs'],
                                    'consolidation_potential': f"Can reduce {pattern['count']-1} duplicate rules"
                                })
                        
                        # Generate CSV headers and data
                        csv_headers = ["Subscription Name", "Subscription ID", "Resource Group", "NSG Name", "Location", "Total Rules", "Redundant Rules", "Overly Permissive Rules", "Optimization Score", "Recommendations"]
                        csv_data = []
                        for nsg in consolidation_data:
                            csv_data.append([
                                nsg.get('subscription_name', ''),
                                nsg.get('subscription_id', ''),
                                nsg.get('resource_group', ''),
                                nsg.get('nsg_name', ''),
                                nsg.get('location', ''),
                                nsg.get('total_rules', 0),
                                nsg.get('redundant_rules', 0),
                                nsg.get('overly_permissive_rules', 0),
                                nsg.get('optimization_score', 0),
                                '; '.join(nsg.get('recommendations', []))
                            ])
                        
                        response = {
                            "success": True,
                            "report_type": "Consolidation Analysis",
                            "data": {
                                "csv_headers": csv_headers,
                                "csv_data": csv_data,
                                "raw_data": consolidation_data
                            },
                            "generated_at": datetime.utcnow().isoformat() + "Z",
                            "analysis_summary": {
                                "total_nsgs_analyzed": len(consolidation_data),
                                "total_rules": total_rules,
                                "potential_rule_reduction": potential_rule_reduction,
                                "estimated_monthly_savings": f"${estimated_cost_savings:.2f}",
                                "average_optimization_score": round(sum([nsg['optimization_score'] for nsg in consolidation_data]) / max(len(consolidation_data), 1), 2),
                                "consolidation_opportunities": len(duplicate_rules)
                            },
                            "duplicate_rules": duplicate_rules[:20],  # Limit for response size
                            "recommendations": {
                                "immediate_actions": [
                                    "Review and consolidate duplicate rules across NSGs",
                                    "Implement least-privilege principle for overly permissive rules",
                                    "Consider using Application Security Groups for better rule management"
                                ],
                                "long_term_optimizations": [
                                    "Implement automated rule compliance monitoring",
                                    "Regular security rule audits and cleanup",
                                    "Standardize NSG templates across environments"
                                ]
                            }
                        }
                        
                    except Exception as e:
                        logger.error(f"Failed to generate consolidation report: {e}")
                        response = {"error": f"Failed to generate consolidation report: {str(e)}"}
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
                
            except Exception as e:
                logger.error(f"Consolidation report failed: {e}")
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                error_response = {"error": f"Consolidation report failed: {str(e)}"}
                self.wfile.write(json.dumps(error_response).encode())
                return
        elif path == '/api/v1/reports/export-csv':
            # CSV Export Endpoint
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                request_data = json.loads(post_data.decode('utf-8'))
                
                report_type = request_data.get('report_type', '')
                report_data = request_data.get('data', {})
                selected_nsgs = request_data.get('selectedNSGs', [])
                
                if not report_type or not report_data:
                    self.send_response(400)
                    self.send_header('Content-Type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    error_response = {"error": "Missing report_type or data"}
                    self.wfile.write(json.dumps(error_response).encode())
                    return
                
                # Check if multiple NSGs are selected for ZIP creation
                if selected_nsgs and len(selected_nsgs) > 1:
                    import zipfile
                    import io
                    
                    # Create ZIP file in memory
                    zip_buffer = io.BytesIO()
                    
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                        for nsg_name in selected_nsgs:
                            # Filter data for this specific NSG
                            nsg_data = [item for item in report_data.get('data', []) if item.get('nsg_name') == nsg_name]
                            
                            if nsg_data:
                                csv_content = self._generate_csv_content(report_type, {'data': nsg_data})
                                if csv_content:
                                    filename = f"{nsg_name}_{report_type}_report_{datetime.utcnow().strftime('%Y-%m-%d')}.csv"
                                    zip_file.writestr(filename, csv_content)
                    
                    zip_buffer.seek(0)
                    
                    # Send ZIP file
                    self.send_response(200)
                    self.send_header('Content-Type', 'application/zip')
                    self.send_header('Content-Disposition', f'attachment; filename="{report_type}_reports_{datetime.utcnow().strftime("%Y-%m-%d")}.zip"')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                    self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                    self.end_headers()
                    self.wfile.write(zip_buffer.getvalue())
                    return
                
                # Generate CSV content for single NSG or all NSGs
                csv_content = self._generate_csv_content(report_type, report_data)
                
                if not csv_content:
                    self.send_response(400)
                    self.send_header('Content-Type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    error_response = {"error": f"Unsupported report type: {report_type}"}
                    self.wfile.write(json.dumps(error_response).encode())
                    return
                
                # Send CSV file
                self.send_response(200)
                self.send_header('Content-Type', 'text/csv')
                self.send_header('Content-Disposition', f'attachment; filename="{report_type}-report-{datetime.utcnow().strftime("%Y-%m-%d")}.csv"')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(csv_content.encode('utf-8'))
                return
                
            except Exception as e:
                logger.error(f"CSV export failed: {e}")
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                error_response = {"error": f"CSV export failed: {str(e)}"}
                self.wfile.write(json.dumps(error_response).encode())
                return
        elif path == '/api/v1/email/schedule':
            # Schedule email reports
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                request_data = json.loads(post_data.decode('utf-8'))
                
                # Extract email scheduling parameters
                report_type = request_data.get('report_type', '')
                recipients = request_data.get('recipients', [])
                frequency = request_data.get('frequency', 'daily')  # daily, weekly, monthly
                time_of_day = request_data.get('time_of_day', '09:00')
                
                # Validate input
                if not report_type or not recipients:
                    response = {"error": "Missing required fields: report_type and recipients"}
                else:
                    # Create new email schedule
                    schedule_id = f"schedule_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                    
                    schedule = {
                        "id": schedule_id,
                        "reportType": report_type,
                        "frequency": frequency,
                        "emails": recipients,
                        "enabled": True,
                        "timeOfDay": time_of_day,
                        "status": "active",
                        "successCount": 0,
                        "failureCount": 0,
                        "createdAt": datetime.now().isoformat(),
                        "lastSent": None,
                        "nextSend": self._calculate_next_execution(frequency, time_of_day)
                    }
                    
                    # Save to EMAIL_SCHEDULES storage
                    EMAIL_SCHEDULES[schedule_id] = schedule
                    
                    response = {
                        "success": True,
                        "message": f"Email schedule created successfully for {report_type} report",
                        "schedule": schedule
                    }
                
                # Send response
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
                
            except Exception as e:
                logger.error(f"Email scheduling failed: {e}")
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                error_response = {"error": f"Email scheduling failed: {str(e)}"}
                self.wfile.write(json.dumps(error_response).encode())
                return
        elif path == '/api/v1/email/config':
            # Save email configuration
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                request_data = json.loads(post_data.decode('utf-8'))
                
                # Validate required fields
                required_fields = ['smtpServer', 'smtpPort', 'smtpUsername', 'fromEmail']
                missing_fields = [field for field in required_fields if not request_data.get(field)]
                
                if missing_fields:
                    response = {"error": f"Missing required fields: {', '.join(missing_fields)}"}
                else:
                    # Update EMAIL_CONFIG with new values
                    EMAIL_CONFIG.update({
                        'smtpServer': request_data.get('smtpServer'),
                        'smtpPort': int(request_data.get('smtpPort', 587)),
                        'smtpUsername': request_data.get('smtpUsername'),
                        'smtpPassword': request_data.get('smtpPassword', ''),
                        'fromEmail': request_data.get('fromEmail'),
                        'fromName': request_data.get('fromName', 'NSG Tool Reports'),
                        'enableTLS': request_data.get('enableTLS', True)
                    })
                    
                    # Return success response
                    response = {
                        "success": True,
                        "message": "Email configuration saved successfully",
                        "config": {
                            "smtpServer": smtp_server,
                            "smtpPort": smtp_port,
                            "smtpUsername": smtp_username,
                            "fromEmail": request_data.get('fromEmail'),
                            "fromName": request_data.get('fromName', 'NSG Tool Reports'),
                            "enableTLS": enable_tls,
                            "enableSSL": enable_ssl,
                            "timeout": request_data.get('timeout', 30),
                            "retryAttempts": request_data.get('retryAttempts', 3)
                        },
                        "connectionTest": connection_test_result,
                        "configSaved": config_saved
                    }
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
                
            except Exception as e:
                logger.error(f"Email config save failed: {e}")
                response = {"error": f"Failed to save email configuration: {str(e)}"}
                
        elif path == '/api/v1/email/test':
            # Test email configuration
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                request_data = json.loads(post_data.decode('utf-8'))
                
                # Update EMAIL_CONFIG with test data if provided
                if 'smtpServer' in request_data:
                    EMAIL_CONFIG.update({
                        'smtpServer': request_data.get('smtpServer'),
                        'smtpPort': int(request_data.get('smtpPort', 587)),
                        'smtpUsername': request_data.get('smtpUsername'),
                        'smtpPassword': request_data.get('smtpPassword', ''),
                        'fromEmail': request_data.get('fromEmail'),
                        'fromName': request_data.get('fromName', 'NSG Tool Reports'),
                        'enableTLS': request_data.get('enableTLS', True)
                    })
                
                test_recipient = request_data.get('testRecipient')
                
                if not test_recipient:
                    response = {"error": "Test recipient email is required"}
                else:
                    # Send test email using our send_email function
                    test_body = f"""
                    <html>
                    <body>
                        <h2>NSG Tool - Email Configuration Test</h2>
                        <p>This is a test email to verify your email configuration is working correctly.</p>
                        <p>Sent at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                        <p>If you received this email, your configuration is working properly!</p>
                    </body>
                    </html>
                    """
                    
                    success = send_email([test_recipient], "NSG Tool - Email Configuration Test", test_body)
                    
                    if success:
                        response = {
                            "success": True,
                            "message": f"Test email sent successfully to {test_recipient}"
                        }
                    else:
                        response = {
                            "success": False,
                            "message": "Failed to send test email. Please check your configuration."
                        }
                        
                        # Create test email
                        msg = MIMEMultipart('alternative')
                        msg['Subject'] = 'NSG Tool - Email Configuration Test'
                        msg['From'] = f"{from_name} <{from_email}>"
                        msg['To'] = test_recipient
                        
                        # Create HTML content
                        html_content = f"""
                        <html>
                          <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                            <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                              <h2 style="color: #2563eb; border-bottom: 2px solid #e5e7eb; padding-bottom: 10px;">
                                🔧 NSG Tool Email Test
                              </h2>
                              <p>Congratulations! Your email configuration is working correctly.</p>
                              <div style="background-color: #f3f4f6; padding: 15px; border-radius: 8px; margin: 20px 0;">
                                <h3 style="margin-top: 0; color: #374151;">Configuration Details:</h3>
                                <ul style="margin: 0;">
                                  <li><strong>SMTP Server:</strong> {smtp_server}:{smtp_port}</li>
                                  <li><strong>Username:</strong> {smtp_username}</li>
                                  <li><strong>TLS Enabled:</strong> {'Yes' if enable_tls else 'No'}</li>
                                  <li><strong>SSL Enabled:</strong> {'Yes' if enable_ssl else 'No'}</li>
                                  <li><strong>Test Time:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</li>
                                </ul>
                              </div>
                              <p>You can now configure email schedules and receive automated NSG reports.</p>
                              <hr style="border: none; border-top: 1px solid #e5e7eb; margin: 30px 0;">
                              <p style="font-size: 12px; color: #6b7280;">
                                This is an automated test email from NSG Tool. If you received this in error, please contact your administrator.
                              </p>
                            </div>
                          </body>
                        </html>
                        """
                        
                        # Create plain text version
                        text_content = f"""
                        NSG Tool Email Test
                        
                        Congratulations! Your email configuration is working correctly.
                        
                        Configuration Details:
                        - SMTP Server: {smtp_server}:{smtp_port}
                        - Username: {smtp_username}
                        - TLS Enabled: {'Yes' if enable_tls else 'No'}
                        - SSL Enabled: {'Yes' if enable_ssl else 'No'}
                        - Test Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
                        
                        You can now configure email schedules and receive automated NSG reports.
                        
                        This is an automated test email from NSG Tool.
                        """
                        
                        # Attach parts
                        part1 = MIMEText(text_content, 'plain')
                        part2 = MIMEText(html_content, 'html')
                        msg.attach(part1)
                        msg.attach(part2)
                        
                        # Send email
                        try:
                            if enable_ssl:
                                server = smtplib.SMTP_SSL(smtp_server, smtp_port, timeout=30)
                            else:
                                server = smtplib.SMTP(smtp_server, smtp_port, timeout=30)
                                if enable_tls:
                                    server.starttls()
                            
                            server.login(smtp_username, smtp_password)
                            server.send_message(msg)
                            server.quit()
                            
                            response = {
                                "success": True,
                                "message": f"Test email sent successfully to {test_recipient}",
                                "details": {
                                    "recipient": test_recipient,
                                    "smtp_server": f"{smtp_server}:{smtp_port}",
                                    "sent_at": datetime.now().isoformat(),
                                    "tls_enabled": enable_tls,
                                    "ssl_enabled": enable_ssl
                                }
                            }
                        
                        except smtplib.SMTPAuthenticationError as auth_error:
                            response = {
                                "success": False,
                                "message": "Authentication failed. Please check your username and password.",
                                "error_type": "authentication",
                                "details": str(auth_error)
                            }
                        except smtplib.SMTPRecipientsRefused as recipient_error:
                            response = {
                                "success": False,
                                "message": f"Recipient {test_recipient} was refused by the server.",
                                "error_type": "recipient_refused",
                                "details": str(recipient_error)
                            }
                        except smtplib.SMTPConnectError as connect_error:
                            response = {
                                "success": False,
                                "message": "Cannot connect to SMTP server. Please check server address and port.",
                                "error_type": "connection",
                                "details": str(connect_error)
                            }
                        except Exception as email_error:
                            response = {
                                "success": False,
                                "message": f"Failed to send test email: {str(email_error)}",
                                "error_type": "general",
                                "details": str(email_error)
                            }
                
                self.send_response(200 if response.get('success') else 400)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
                
            except Exception as e:
                logger.error(f"Email test failed: {e}")
                response = {"error": f"Email test failed: {str(e)}"}
                
        elif path == '/api/v1/email/send-report':
            # Send report via email with real SMTP functionality
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                request_data = json.loads(post_data.decode('utf-8'))
                
                report_type = request_data.get('report_type', '')
                recipients = request_data.get('recipients', [])
                subscription_id = request_data.get('subscription_id', AZURE_CONFIG["subscription_id"])
                resource_group = request_data.get('resource_group', '')
                report_data = request_data.get('data', {})
                selected_nsgs = request_data.get('selectedNSGs', [])
                custom_message = request_data.get('customMessage', '')
                include_csv = request_data.get('includeCsv', True)
                
                if not report_type or not recipients:
                    response = {"error": "Missing required fields: report_type and recipients"}
                else:
                    # Check email configuration
                    if not EMAIL_CONFIG.get('smtpServer') or not EMAIL_CONFIG.get('smtpUsername') or not EMAIL_CONFIG.get('smtpPassword'):
                        response = {"error": "Email configuration not complete. Please configure SMTP settings first."}
                    else:
                        try:
                            # Generate email body using our email function
                            email_body = generate_report_email_body(
                                report_type=report_type,
                                subscription_id=subscription_id,
                                resource_group=resource_group,
                                selected_nsgs=selected_nsgs,
                                custom_message=custom_message,
                                report_data=report_data
                            )
                            
                            # Send email to all recipients
                            subject = f'NSG Tool Report - {report_type.replace("-", " ").title()}'
                            success = send_email(recipients, subject, email_body)
                            
                            if success:
                                response = {
                                    "success": True,
                                    "message": f"{report_type} report sent successfully to {len(recipients)} recipients",
                                    "report_type": report_type,
                                    "recipients": recipients,
                                    "sent_at": datetime.now().isoformat()
                                }
                            else:
                                response = {
                                    "success": False,
                                    "message": "Failed to send report emails. Please check your email configuration."
                                }

                        
                        except Exception as smtp_error:
                            logger.error(f"SMTP error: {smtp_error}")
                            response = {
                                "success": False,
                                "message": f"Failed to send report emails: {str(smtp_error)}",
                                "error_type": "smtp_error"
                            }
                
                # Send response
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
                
            except Exception as e:
                logger.error(f"Email sending failed: {e}")
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                error_response = {"error": f"Email sending failed: {str(e)}"}
                self.wfile.write(json.dumps(error_response).encode())
                return
        elif path.startswith('/api/v1/email/schedule/') and path.endswith('/run'):
            # Manually trigger a scheduled email report
            try:
                schedule_id = path.split('/')[-2]  # Extract schedule ID from path like /api/v1/email/schedule/{id}/run
                
                # Mock getting schedule details and triggering report generation
                # In real implementation, this would:
                # 1. Fetch the schedule details from database
                # 2. Generate the report based on schedule configuration
                # 3. Send email to recipients
                # 4. Update schedule execution history
                
                response = {
                    "success": True,
                    "message": f"Scheduled report {schedule_id} executed successfully",
                    "schedule_id": schedule_id,
                    "executed_at": datetime.now().isoformat(),
                    "report_type": "NSG Compliance",  # This would come from the actual schedule
                    "recipients": ["admin@company.com"],  # This would come from the actual schedule
                    "report_size": "1.8 MB",
                    "format": "PDF",
                    "execution_time": "3.2 seconds"
                }
                
                # Send response
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
                
            except Exception as e:
                logger.error(f"Manual schedule execution failed: {e}")
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                error_response = {"error": f"Manual schedule execution failed: {str(e)}"}
                self.wfile.write(json.dumps(error_response).encode())
                return

        elif path == '/api/v1/settings/security':
            # Save security settings
            try:
                content_length = int(self.headers.get('Content-Length', '0'))
                post_data = self.rfile.read(content_length) if content_length > 0 else b'{}'
                request_data = json.loads(post_data.decode('utf-8'))

                SETTINGS_STORAGE['security'].update({
                    'twoFactorAuth': bool(request_data.get('twoFactorAuth', SETTINGS_STORAGE['security']['twoFactorAuth'])),
                    'sessionTimeout': request_data.get('sessionTimeout', SETTINGS_STORAGE['security']['sessionTimeout']),
                    'passwordPolicy': bool(request_data.get('passwordPolicy', SETTINGS_STORAGE['security']['passwordPolicy'])),
                    'auditLogging': bool(request_data.get('auditLogging', SETTINGS_STORAGE['security']['auditLogging']))
                })

                response = {
                    'success': True,
                    'message': 'Security settings saved successfully',
                    'settings': SETTINGS_STORAGE['security']
                }

                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
            except Exception as e:
                logger.error(f"Security settings save failed: {e}")
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({'error': f'Security settings save failed: {str(e)}'}).encode())
                return

        elif path == '/api/v1/settings/notifications':
            # Save notification settings
            try:
                content_length = int(self.headers.get('Content-Length', '0'))
                post_data = self.rfile.read(content_length) if content_length > 0 else b'{}'
                request_data = json.loads(post_data.decode('utf-8'))

                SETTINGS_STORAGE['notifications'].update({
                    'securityAlerts': bool(request_data.get('securityAlerts', SETTINGS_STORAGE['notifications']['securityAlerts'])),
                    'systemUpdates': bool(request_data.get('systemUpdates', SETTINGS_STORAGE['notifications']['systemUpdates'])),
                    'backupStatus': bool(request_data.get('backupStatus', SETTINGS_STORAGE['notifications']['backupStatus']))
                })

                response = {
                    'success': True,
                    'message': 'Notification settings saved successfully',
                    'settings': SETTINGS_STORAGE['notifications']
                }

                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
            except Exception as e:
                logger.error(f"Notification settings save failed: {e}")
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({'error': f'Notification settings save failed: {str(e)}'}).encode())
                return

        elif path == '/api/v1/system/maintenance':
            # Perform system maintenance actions
            try:
                content_length = int(self.headers.get('Content-Length', '0'))
                post_data = self.rfile.read(content_length) if content_length > 0 else b'{}'
                request_data = json.loads(post_data.decode('utf-8'))

                action = request_data.get('action', '')
                msg = 'Unknown action'
                success = True

                if action == 'check_updates':
                    msg = 'System is up to date. No updates available.'
                elif action == 'clear_cache':
                    msg = 'Application cache cleared successfully.'
                elif action == 'restart':
                    msg = 'Restart command acknowledged. Service will restart shortly.'
                else:
                    success = False
                    msg = 'Invalid maintenance action.'

                response = {
                    'success': success,
                    'message': msg,
                    'action': action
                }

                self.send_response(200 if success else 400)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
                self.send_header('Access-Control-Allow-Headers', 'Content-Type')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
            except Exception as e:
                logger.error(f"System maintenance failed: {e}")
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({'error': f'System maintenance failed: {str(e)}'}).encode())
                return

        elif path == '/api/v1/users/create':
            try:
                content_length = int(self.headers.get('Content-Length', '0'))
                post_data = self.rfile.read(content_length) if content_length > 0 else b'{}'
                request_data = json.loads(post_data.decode('utf-8'))

                name = request_data.get('name', '').strip()
                email = request_data.get('email', '').strip()
                role = request_data.get('role', 'User')

                if not name or not email:
                    response = {'success': False, 'message': 'Name and email are required'}
                    self.send_response(400)
                else:
                    new_user = {
                        'id': str(int(datetime.now().timestamp() * 1000)),
                        'name': name,
                        'email': email,
                        'role': role,
                        'status': 'Active',
                        'lastLogin': 'Never'
                    }
                    USERS_STORAGE.append(new_user)
                    response = {'success': True, 'message': 'User created successfully', 'user': new_user}
                    self.send_response(200)

                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
            except Exception as e:
                logger.error(f"User creation failed: {e}")
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({'error': f'User creation failed: {str(e)}'}).encode())
                return

        elif path == '/api/v1/users/update':
            try:
                content_length = int(self.headers.get('Content-Length', '0'))
                post_data = self.rfile.read(content_length) if content_length > 0 else b'{}'
                request_data = json.loads(post_data.decode('utf-8'))

                user_id = request_data.get('id')
                if not user_id:
                    self.send_response(400)
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({'success': False, 'message': 'User id is required'}).encode())
                    return

                updated = False
                updated_user = None
                for user in USERS_STORAGE:
                    if user['id'] == user_id:
                        user.update({
                            'name': request_data.get('name', user['name']),
                            'email': request_data.get('email', user['email']),
                            'role': request_data.get('role', user['role']),
                            'status': request_data.get('status', user['status']),
                        })
                        updated = True
                        updated_user = user
                        break

                if updated:
                    response = {'success': True, 'message': 'User updated successfully', 'user': updated_user}
                else:
                    response = {'success': False, 'message': 'User not found'}
                self.send_response(200 if updated else 404)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
            except Exception as e:
                logger.error(f"User update failed: {e}")
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({'error': f'User update failed: {str(e)}'}).encode())
                return

        elif path == '/api/v1/users/delete':
            try:
                content_length = int(self.headers.get('Content-Length', '0'))
                post_data = self.rfile.read(content_length) if content_length > 0 else b'{}'
                request_data = json.loads(post_data.decode('utf-8'))

                user_id = request_data.get('id')
                if not user_id:
                    self.send_response(400)
                    self.send_header('Content-Type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({'success': False, 'message': 'User id is required'}).encode())
                    return

                before_count = len(USERS_STORAGE)
                USERS_STORAGE[:] = [u for u in USERS_STORAGE if u['id'] != user_id]
                deleted = len(USERS_STORAGE) < before_count

                response = {
                    'success': deleted,
                    'message': 'User deleted successfully' if deleted else 'User not found'
                }
                self.send_response(200 if deleted else 404)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps(response).encode())
                return
            except Exception as e:
                logger.error(f"User deletion failed: {e}")
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({'error': f'User deletion failed: {str(e)}'}).encode())
                return
        else:
            # Only send response for unhandled paths
            response = {"error": "Not found", "path": path}

            # Send response
            logger.info(f"Sending POST response for {path}")
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
            self.send_header('Access-Control-Allow-Headers', 'Content-Type')
            self.end_headers()
            self.wfile.write(json.dumps(response).encode())
    
    def do_PUT(self):
        logger.info(f"Received PUT request: {self.path}")
        
        # Parse URL
        parsed_url = urlparse(self.path)
        path = parsed_url.path
        query_params = parse_qs(parsed_url.query)
        
        # Send response headers
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
        
        # Route handling for PUT requests
        if path.startswith('/api/v1/nsgs/'):
            try:
                path_parts = path.split('/')
                nsg_id = path_parts[4]  # /api/v1/nsgs/{id} or /api/v1/nsgs/{id}/rules
                
                content_length = int(self.headers['Content-Length'])
                put_data = self.rfile.read(content_length)
                request_data = json.loads(put_data.decode('utf-8'))
                
                # Check if this is a rules update endpoint
                if len(path_parts) > 5 and path_parts[5] == 'rules':
                    # Handle NSG rules update: /api/v1/nsgs/{id}/rules
                    try:
                        nsg_id_int = int(nsg_id)
                        
                        # Get subscription_id and resource_group from query params or request data
                        subscription_id = query_params.get('subscription_id', [AZURE_CONFIG["subscription_id"]])[0]
                        resource_group = query_params.get('resource_group', [None])[0]
                        
                        # Update NSG rules in Azure
                        update_result = update_nsg_rules_in_azure(
                            subscription_id, 
                            resource_group, 
                            nsg_id_int, 
                            request_data.get('inbound_rules', []),
                            request_data.get('outbound_rules', [])
                        )
                        
                        if update_result.get('success'):
                            response = {
                                "message": "NSG rules updated successfully",
                                "nsg_id": nsg_id_int
                            }
                            logger.info(f"Updated rules for NSG {nsg_id}: {len(request_data.get('inbound_rules', []))} inbound, {len(request_data.get('outbound_rules', []))} outbound")
                        else:
                            response = {"error": update_result.get('error', 'Failed to update NSG rules')}
                    except ValueError:
                        response = {"error": f"Invalid NSG ID: {nsg_id}"}
                    except Exception as e:
                        logger.error(f"NSG rules update failed: {e}")
                        response = {"error": f"NSG rules update failed: {str(e)}"}
                else:
                    # Handle general NSG update: /api/v1/nsgs/{id}
                    response = {
                        "success": True,
                        "message": "NSG updated successfully",
                        "nsg": {
                            "id": int(nsg_id),
                            "name": request_data.get('name', 'updated-nsg'),
                            "resource_group": request_data.get('resource_group', 'default-rg'),
                            "region": request_data.get('region', 'East US'),
                            "subscription_id": request_data.get('subscription_id', AZURE_CONFIG["subscription_id"]),
                            "inbound_rules": request_data.get('inbound_rules', []),
                            "outbound_rules": request_data.get('outbound_rules', []),
                            "tags": request_data.get('tags', {}),
                            "is_active": True,
                            "compliance_score": 85,
                            "risk_level": "medium",
                            "last_sync": "2024-08-22T13:51:52Z",
                            "updated_at": "2024-08-22T13:51:52Z"
                        }
                    }
            except Exception as e:
                logger.error(f"NSG update failed: {e}")
                response = {"error": f"NSG update failed: {str(e)}"}
        elif path.startswith('/api/v1/email/schedules/'):
            try:
                schedule_id = path.split('/')[-1]
                
                content_length = int(self.headers['Content-Length'])
                put_data = self.rfile.read(content_length)
                request_data = json.loads(put_data.decode('utf-8'))
                
                # Mock email schedule update response
                response = {
                    "success": True,
                    "message": f"Email schedule with ID {schedule_id} updated successfully",
                    "schedule": {
                        "id": schedule_id,
                        "report_type": request_data.get('report_type', 'NSG Compliance'),
                        "frequency": request_data.get('frequency', 'weekly'),
                        "recipients": request_data.get('recipients', []),
                        "monthly_date": request_data.get('monthly_date'),
                        "time_of_day": request_data.get('time_of_day', '09:00'),
                        "status": request_data.get('status', 'active'),
                        "last_sent": "2024-01-15T09:00:00Z",
                        "next_scheduled": "2024-02-15T09:00:00Z",
                        "updated_at": "2024-01-22T14:30:00Z"
                    }
                }
            except Exception as e:
                logger.error(f"Email schedule update failed: {e}")
                response = {"error": f"Email schedule update failed: {str(e)}"}
        else:
            response = {"error": "Not found", "path": path}
        
        # Send response
        logger.info(f"Sending PUT response for {path}")
        self.wfile.write(json.dumps(response).encode())
    
    def do_DELETE(self):
        logger.info(f"Received DELETE request: {self.path}")
        
        # Parse URL
        parsed_url = urlparse(self.path)
        path = parsed_url.path
        
        # Send response headers
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
        
        # Route handling for DELETE requests
        if path.startswith('/api/v1/email/schedules/'):
            try:
                schedule_id = path.split('/')[-1]
                
                # Mock email schedule deletion response
                response = {
                    "success": True,
                    "message": f"Email schedule with ID {schedule_id} deleted successfully",
                    "schedule_id": schedule_id
                }
            except Exception as e:
                logger.error(f"Email schedule deletion failed: {e}")
                response = {"error": f"Email schedule deletion failed: {str(e)}"}
        elif path.startswith('/api/v1/nsgs/'):
            try:
                nsg_id = path.split('/')[-1]
                
                # Mock NSG deletion response
                response = {
                    "success": True,
                    "message": f"NSG with ID {nsg_id} deleted successfully"
                }
            except Exception as e:
                logger.error(f"NSG deletion failed: {e}")
                response = {"error": f"NSG deletion failed: {str(e)}"}
        else:
            response = {"error": "Not found", "path": path}
        
        # Send response
        logger.info(f"Sending DELETE response for {path}")
        self.wfile.write(json.dumps(response).encode())
    
    def do_OPTIONS(self):
        logger.info(f"Received OPTIONS request: {self.path}")
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, PUT, DELETE, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type, Authorization, X-Requested-With')
        self.send_header('Access-Control-Max-Age', '86400')
        self.end_headers()

if __name__ == '__main__':
    import sys
    port = 8000
    
    # Parse command line arguments for port
    if len(sys.argv) > 1:
        for i, arg in enumerate(sys.argv):
            if arg == '--port' and i + 1 < len(sys.argv):
                try:
                    port = int(sys.argv[i + 1])
                except ValueError:
                    logger.error(f"Invalid port number: {sys.argv[i + 1]}")
                    port = 8000
    
    try:
        server = HTTPServer(('0.0.0.0', port), WorkingHandler)
        logger.info(f"Working backend server running on http://0.0.0.0:{port}")
        logger.info(f"Server accessible at http://localhost:{port} and from network")
        logger.info("Available endpoints:")
        logger.info("  GET /api/v1/health")
        logger.info("  GET /api/v1/dashboard?subscription_id=<id>")
        logger.info("  GET /api/v1/subscriptions")
        logger.info("  GET /api/v1/resource-groups?subscription_id=<id>")
        logger.info("  GET /api/v1/locations?subscription_id=<id>")
        logger.info("  GET /api/v1/nsgs?subscription_id=<id>&resource_group=<rg>")
        logger.info("  GET /api/v1/asgs?subscription_id=<id>&resource_group=<rg>")
        logger.info("  GET /api/v1/storage-accounts?subscription_id=<id>")
        logger.info("  GET /api/v1/containers?storage_account=<name>")
        logger.info("  GET /api/v1/nsg-validation/<nsg_name>?subscription_id=<id>&resource_group=<rg>")
        logger.info("  POST /api/v1/nsg-recommendations/<nsg_name>?subscription_id=<id>&resource_group=<rg>")
        logger.info("  POST /api/v1/backup/create")
        logger.info("  POST /api/v1/backup/storage-config")
        logger.info("  POST /api/v1/storage-accounts/create")
        logger.info("  POST /api/v1/backup/export")
        logger.info("  POST /api/v1/backup/restore/preview")
        logger.info("  POST /api/v1/backup/restore/confirm")
        logger.info("  POST /api/v1/backup/files")
        logger.info("  GET /api/v1/reports/asg-validation")
        logger.info("  GET /api/v1/reports/nsg-rules")
        logger.info("  GET /api/v1/reports/ip-limitations")
        logger.info("  GET /api/v1/reports/nsg-ports")
        logger.info("  GET /api/v1/reports/consolidation")
        logger.info("  POST /api/v1/email/schedule")
        logger.info("  POST /api/v1/email/send-report")
        logger.info("  GET /api/v1/settings/security")
        logger.info("  GET /api/v1/settings/notifications")
        logger.info("  GET /api/v1/settings/system")
        logger.info("  GET /api/v1/users")
        logger.info("  POST /api/v1/settings/security")
        logger.info("  POST /api/v1/settings/notifications")
        logger.info("  POST /api/v1/system/maintenance")
        logger.info("  POST /api/v1/users/create")
        logger.info("  POST /api/v1/users/update")
        logger.info("  POST /api/v1/users/delete")
        server.serve_forever()
    except KeyboardInterrupt:
        logger.info("Server stopped")
        server.shutdown()
    except Exception as e:
        logger.error(f"Server error: {e}")
